# postmarket_runner.py (BULLETPROOF - uses DAILY_LOG for BOTH premarket+midday + weekly excel attach + perf log hardening + debug artifacts)
from __future__ import annotations

import os
import re
import json
import html as _html
from pathlib import Path
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from typing import Optional, Tuple, Dict

import pandas as pd
from pandas import DatetimeTZDtype
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from llm.client import llm_text
from llm.explain import safe_postmarket_coach

from config import (
    APP_ENV,
    IS_LOCAL,
    SENDER_EMAIL,
    RECEIVER_EMAIL,
    LOCAL_RECEIVER_EMAIL,
    EMAIL_SUBJECT_PREFIX_LOCAL,
    EMAIL_SUBJECT_PREFIX_PROD,
    SCORE_COLORS,
    INTRADAY_INTERVAL,
    CONSERVATIVE_SAME_BAR_POLICY,
    EVAL_FALLBACK_STRICT,
    MARKET_OPEN_CT,
    MARKET_CLOSE_CT,
    POST_MARKET_START_CT,
)

from email_sender import send_email as _send_email
from performance_tracker import (
    PortfolioConfig,
    load_open_portfolio,
    save_open_portfolio,
    append_trade_history,
    update_and_close_positions,
    portfolio_summary,
)

import warnings
warnings.filterwarnings(
    "ignore",
    category=FutureWarning,
    message="The behavior of DataFrame concatenation with empty or all-NA entries is deprecated"
)

LOCAL_TZ = ZoneInfo("America/Chicago")
POST_MARKET_START = POST_MARKET_START_CT

# Debug artifacts (set POSTMARKET_DEBUG=1 to enable extra CSVs)
POSTMARKET_DEBUG = os.getenv("POSTMARKET_DEBUG", "0") == "1"
# Local testing override: clamp late reco timestamps to session_start so evaluation still runs
EVAL_CLAMP_LATE_RECO = os.getenv("EVAL_CLAMP_LATE_RECO", "0") == "1"


# -----------------------------
# Output helpers (env-aware)
# -----------------------------
def env_base_dir() -> Path:
    base = Path(__file__).resolve().parent / "outputs" / APP_ENV
    base.mkdir(parents=True, exist_ok=True)
    return base


def logs_dir() -> Path:
    p = env_base_dir() / "logs"
    p.mkdir(parents=True, exist_ok=True)
    return p


def run_dir(now: datetime, mode: str) -> Path:
    day = now.strftime("%Y%m%d")
    p = env_base_dir() / "runs" / day / mode
    p.mkdir(parents=True, exist_ok=True)
    return p


def out_path(filename: str, *, now: datetime | None = None, mode: str | None = None, kind: str = "runs") -> str:
    if kind == "logs":
        return str(logs_dir() / filename)
    if now is None or mode is None:
        return str(env_base_dir() / filename)
    return str(run_dir(now, mode) / filename)


def make_run_id(now: datetime) -> str:
    return now.strftime("%Y%m%d_%H%M%S")


def ensure_csv_exists(path: str, header_cols: list[str]) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    if (not p.exists()) or p.stat().st_size == 0:
        pd.DataFrame(columns=header_cols).to_csv(path, index=False)


# -----------------------------
# Email routing (env-aware)
# -----------------------------
EMAIL_SUBJECT_PREFIX = EMAIL_SUBJECT_PREFIX_LOCAL if IS_LOCAL else EMAIL_SUBJECT_PREFIX_PROD
EFFECTIVE_RECEIVER_EMAIL = (LOCAL_RECEIVER_EMAIL or RECEIVER_EMAIL) if IS_LOCAL else RECEIVER_EMAIL


def send_email(subject: str, html_body: str, attachment_path: str | None = None) -> bool:
    final_subject = f"{EMAIL_SUBJECT_PREFIX} {subject}"
    print(f"📧 Sending email to={EFFECTIVE_RECEIVER_EMAIL} from={SENDER_EMAIL} subject={final_subject} attach={bool(attachment_path)}")

    return _send_email(
        subject=final_subject,
        html_body=html_body,
        to_email=EFFECTIVE_RECEIVER_EMAIL,
        from_email=SENDER_EMAIL,
        attachment_path=attachment_path,
    )


def send_postmarket_email_once(now: datetime, subject: str, html_body: str, attachment_path: str | None = None) -> bool:
    sent_flag = Path(run_dir(now, "postmarket")) / f"email_sent_{now.strftime('%Y-%m-%d')}.txt"
    force_email = os.getenv("FORCE_EMAIL", "0") == "1"


    if sent_flag.exists() and not force_email:
        print("📩 Postmarket email already sent for this run_date — skipping resend.")
        return False

    ok = send_email(subject, html_body, attachment_path)

    if not ok:
        print("⚠️ Postmarket email failed — NOT writing sent-flag (will allow retry).")
        return False

    sent_flag.write_text(
        f"sent_ts={now.strftime('%Y-%m-%d %H:%M:%S')}\nrun_id={make_run_id(now)}\n",
        encoding="utf-8",
    )
    return True


# -----------------------------
# Logs
# -----------------------------
DAILY_LOG_CSV = out_path("daily_stock_log.csv", kind="logs")
RECO_LOG_CSV = out_path("recommendations_log.csv", kind="logs")
PERF_LOG_CSV = out_path("performance_log.csv", kind="logs")

PERF_COLS = [
    "run_date",
    "source_mode",
    "instrument_type",
    "symbol",
    "decision",
    "score",
    "confidence",
    "entry_price",
    "target_price",
    "stop_loss",
    "reco_ts",
    "session_start",
    "session_end",
    "day_high_after_reco",
    "day_low_after_reco",
    "close_price",
    "actual_change_pct_close",
    "target_hit",
    "stop_hit",
    "first_hit",
    "first_hit_time",
    "hit_latency_minutes",
    "target_overshoot_pct",
    "best_exit_price_after_target",
    "best_exit_time_after_target",
    "best_exit_from_entry_pct",
    "best_exit_from_target_pct",
    "best_exit_latency_minutes",
    "outcome",

    # NEW: midday next-session evaluation columns
    "next_session_end",
    "next_close_price",
    "next_actual_change_pct_close",
    "next_target_hit",
    "next_stop_hit",
    "next_first_hit",
    "next_first_hit_time",
    "next_hit_latency_minutes",
    "next_target_overshoot_pct",
    "next_best_exit_price_after_target",
    "next_best_exit_time_after_target",
    "next_best_exit_from_entry_pct",
    "next_best_exit_from_target_pct",
    "next_best_exit_latency_minutes",
    "next_outcome",
    "combined_outcome",
    "recovered_after_stop",
]
ensure_csv_exists(PERF_LOG_CSV, PERF_COLS)


# -----------------------------
# Excel styling
# -----------------------------
def normalize_color(color: str) -> str:
    if not color:
        color = "#FFFFFF"
    color = color.lstrip("#")
    if len(color) == 6:
        color = "FF" + color
    return color.upper()


def style_excel_sheet(sheet):
    if sheet.max_row < 1 or sheet.max_column < 1:
        return

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color=normalize_color("#2F5597"),
        end_color=normalize_color("#2F5597"),
        fill_type="solid",
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    sheet.freeze_panes = "A2"

    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = min(max_length + 2, 55)


# -----------------------------
# Helpers
# -----------------------------
def _safe_float(x) -> Optional[float]:
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return None
        return float(x)
    except Exception:
        return None


def _ensure_cols(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    out = df.copy() if df is not None else pd.DataFrame()
    for c in cols:
        if c not in out.columns:
            out[c] = pd.NA
    return out


def _parse_ts_maybe(val: object) -> Optional[datetime]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        if isinstance(val, datetime):
            dt = val
        else:
            dt = pd.to_datetime(str(val), errors="coerce").to_pydatetime()
        if dt is None:
            return None
        if dt.tzinfo is None:
            return dt.replace(tzinfo=LOCAL_TZ)
        return dt.astimezone(LOCAL_TZ)
    except Exception:
        return None


def _session_bounds(run_date: str) -> Tuple[datetime, datetime]:
    d = pd.to_datetime(run_date, errors="coerce")
    if pd.isna(d):
        d = pd.Timestamp(datetime.now(LOCAL_TZ).date())
    day = d.date()
    start_dt = datetime.combine(day, MARKET_OPEN_CT, tzinfo=LOCAL_TZ)
    end_dt = datetime.combine(day, MARKET_CLOSE_CT, tzinfo=LOCAL_TZ)
    return start_dt, end_dt


def _classify_instrument(symbol: str, row: Optional[pd.Series] = None) -> str:
    try:
        if row is not None:
            it = str(row.get("instrument_type") or row.get("asset_type") or "").strip().lower()
            if it in {"option", "options"}:
                return "options"
            if it in {"stock", "equity", "shares"}:
                return "stock"
    except Exception:
        pass

    s = (symbol or "").upper().strip()

    if re.fullmatch(r"[A-Z]{1,6}\d{6}[CP]\d{8}", s):
        return "options"
    if re.search(r"\d{6,8}[CP]\d+", s):
        return "options"
    return "stock"


def _intraday_after_reco(symbol: str, start_dt: datetime, end_dt: datetime) -> pd.DataFrame:
    interval = INTRADAY_INTERVAL if INTRADAY_INTERVAL else "5m"
    try:
        start_date = start_dt.date()
        end_date = (end_dt.date() + timedelta(days=1))

        h = yf.Ticker(symbol).history(
            start=str(start_date),
            end=str(end_date),
            interval=interval,
            auto_adjust=False,
            prepost=False,
        )
        if h is None or h.empty:
            return pd.DataFrame()

        try:
            idx = pd.to_datetime(h.index, errors="coerce")
            if getattr(idx, "tz", None) is None:
                idx = idx.tz_localize(LOCAL_TZ)
            else:
                idx = idx.tz_convert(LOCAL_TZ)
            h.index = idx
            h = h[~h.index.isna()].copy()
        except Exception:
            pass

        h = h.sort_index()
        h = h[(h.index >= start_dt) & (h.index <= end_dt)].copy()
        return h

    except Exception:
        return pd.DataFrame()


def _get_close_for_date(symbol: str, run_date: str) -> Optional[float]:
    """
    Returns the CLOSE price for the given run_date (local trading day).
    Uses 7d window and picks the row whose date matches run_date.
    """
    try:
        h = yf.Ticker(symbol).history(period="7d", auto_adjust=False)
        if h is None or h.empty or "Close" not in h.columns:
            return None

        idx = pd.to_datetime(h.index, errors="coerce")
        if getattr(idx, "tz", None) is not None:
            idx = idx.tz_convert(LOCAL_TZ)
        dates = idx.date

        target_day = pd.to_datetime(run_date, errors="coerce")
        if pd.isna(target_day):
            return None
        target_day = target_day.date()

        mask = dates == target_day
        if not mask.any():
            return None

        close = pd.to_numeric(h.loc[mask, "Close"], errors="coerce").dropna()
        if close.empty:
            return None
        return float(close.iloc[-1])
    except Exception:
        return None


def _compute_outcome_from_hits(target_hit: bool, stop_hit: bool) -> str:
    if target_hit:
        return "🏆 Target Hit"
    if stop_hit:
        return "🛑 Stop Hit"
    return "⏳ Not Hit"

def _compute_combined_midday_outcome(
    same_day_target_hit: bool,
    same_day_stop_hit: bool,
    next_target_hit: bool,
) -> str:
    """
    Midday combined interpretation:
    - same-day target = true win
    - same-day stop + next-session target = late win
    - same-day stop only = loss
    - otherwise not hit
    """
    if same_day_target_hit:
        return "🏆 Target Hit"
    if same_day_stop_hit and next_target_hit:
        return "🟡 Late Win"
    if same_day_stop_hit:
        return "🛑 Stop Hit"
    return "⏳ Not Hit"

def _next_trading_day(day) -> datetime.date:
    d = pd.to_datetime(day, errors="coerce")
    if pd.isna(d):
        d = pd.Timestamp(datetime.now(LOCAL_TZ).date())
    cur = d.date()
    while True:
        cur = cur + timedelta(days=1)
        if cur.weekday() < 5:  # Mon-Fri
            return cur


def _evaluate_window(
    symbol: str,
    entry: Optional[float],
    target: Optional[float],
    stop: Optional[float],
    reco_dt: datetime,
    window_end: datetime,
    close_run_date: str,
) -> dict:
    """
    Evaluate one row over a custom time window.
    Used for:
      - same-day evaluation
      - next-session evaluation
    """
    if reco_dt.tzinfo is None:
        reco_dt = reco_dt.replace(tzinfo=LOCAL_TZ)

    close_price = _get_close_for_date(symbol, close_run_date)
    actual_change_pct_close = pd.NA
    if close_price is not None and entry is not None and entry > 0:
        actual_change_pct_close = ((close_price - entry) / entry) * 100.0

    intraday = _intraday_after_reco(symbol, reco_dt, window_end)

    result = {
        "day_high_after_reco": pd.NA,
        "day_low_after_reco": pd.NA,
        "close_price": close_price,
        "actual_change_pct_close": actual_change_pct_close,
        "target_hit": False,
        "stop_hit": False,
        "first_hit": "",
        "first_hit_time": "",
        "hit_latency_minutes": pd.NA,
        "target_overshoot_pct": pd.NA,
        "best_exit_price_after_target": pd.NA,
        "best_exit_time_after_target": "",
        "best_exit_from_entry_pct": pd.NA,
        "best_exit_from_target_pct": pd.NA,
        "best_exit_latency_minutes": pd.NA,
        "outcome": "⏳ Not Hit",
    }

    if intraday is None or intraday.empty:
        if close_price is None:
            result["first_hit"] = "NO_DATA"
            result["first_hit_time"] = "no_intraday_no_close"
            result["outcome"] = "⛔ No Data (Market Closed / Holiday)"
            return result

        if not EVAL_FALLBACK_STRICT:
            if close_price is not None and target is not None and target > 0 and close_price >= target:
                result["target_hit"] = True
                result["first_hit"] = "TARGET"
                result["first_hit_time"] = "close_fallback"
            elif close_price is not None and stop is not None and stop > 0 and close_price <= stop:
                result["stop_hit"] = True
                result["first_hit"] = "STOP"
                result["first_hit_time"] = "close_fallback"

        result["outcome"] = _compute_outcome_from_hits(result["target_hit"], result["stop_hit"])
        return result

    hi = pd.to_numeric(intraday.get("High"), errors="coerce").dropna() if "High" in intraday.columns else pd.Series(dtype=float)
    lo = pd.to_numeric(intraday.get("Low"), errors="coerce").dropna() if "Low" in intraday.columns else pd.Series(dtype=float)

    max_high = float(hi.max()) if not hi.empty else None
    min_low = float(lo.min()) if not lo.empty else None

    result["day_high_after_reco"] = max_high if max_high is not None else pd.NA
    result["day_low_after_reco"] = min_low if min_low is not None else pd.NA

    has_target = target is not None and target > 0
    has_stop = stop is not None and stop > 0

    if not has_target and not has_stop:
        return result

    first_hit_dt: Optional[datetime] = None

    for ts, bar in intraday.iterrows():
        bar_high = _safe_float(bar.get("High"))
        bar_low = _safe_float(bar.get("Low"))

        hit_target_now = has_target and (bar_high is not None) and (bar_high >= target)
        hit_stop_now = has_stop and (bar_low is not None) and (bar_low <= stop)

        if hit_target_now and hit_stop_now:
            if str(CONSERVATIVE_SAME_BAR_POLICY).strip().lower() == "target_first":
                result["target_hit"] = True
                result["first_hit"] = "TARGET"
            else:
                result["stop_hit"] = True
                result["first_hit"] = "STOP"
            first_hit_dt = ts if isinstance(ts, datetime) else None
            break

        if hit_target_now:
            result["target_hit"] = True
            result["first_hit"] = "TARGET"
            first_hit_dt = ts if isinstance(ts, datetime) else None
            break

        if hit_stop_now:
            result["stop_hit"] = True
            result["first_hit"] = "STOP"
            first_hit_dt = ts if isinstance(ts, datetime) else None
            break

    if first_hit_dt is not None:
        try:
            first_hit_dt = first_hit_dt.astimezone(LOCAL_TZ)
            result["first_hit_time"] = first_hit_dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            result["first_hit_time"] = str(first_hit_dt)

        try:
            latency = round((first_hit_dt - reco_dt).total_seconds() / 60.0, 2)
            result["hit_latency_minutes"] = max(latency, 0.0)
        except Exception:
            result["hit_latency_minutes"] = pd.NA

    if result["target_hit"] and has_target and max_high is not None and target and target > 0:
        result["target_overshoot_pct"] = round(((max_high - target) / target) * 100.0, 3)

    if result["target_hit"] and first_hit_dt is not None and "High" in intraday.columns:
        after = intraday[intraday.index >= first_hit_dt].copy()
        if not after.empty:
            after_hi = pd.to_numeric(after["High"], errors="coerce").dropna()
            if not after_hi.empty:
                best_val = float(after_hi.max())
                result["best_exit_price_after_target"] = best_val

                try:
                    best_time = after_hi.idxmax()
                    if isinstance(best_time, datetime):
                        best_time = best_time.astimezone(LOCAL_TZ)
                        result["best_exit_time_after_target"] = best_time.strftime("%Y-%m-%d %H:%M:%S")
                        result["best_exit_latency_minutes"] = max(
                            round((best_time - reco_dt).total_seconds() / 60.0, 2),
                            0.0,
                        )
                except Exception:
                    pass

                if entry is not None and entry > 0:
                    result["best_exit_from_entry_pct"] = round(((best_val - entry) / entry) * 100.0, 3)

                if target is not None and target > 0:
                    result["best_exit_from_target_pct"] = round(((best_val - target) / target) * 100.0, 3)

    result["outcome"] = _compute_outcome_from_hits(result["target_hit"], result["stop_hit"])
    return result


def evaluate_midday_next_session(df_in: pd.DataFrame, run_date: str) -> pd.DataFrame:
    """
    Midday next-session evaluation:
    - reco starts at midday recommendation timestamp
    - evaluation window ends next trading day's close
    """
    if df_in is None or df_in.empty:
        return pd.DataFrame()

    df = df_in.copy()
    df = _ensure_cols(
        df,
        ["symbol", "current", "target_price", "stop_loss", "run_ts", "source_mode"],
    )

    df["symbol"] = df["symbol"].astype(str).str.upper().str.strip()
    df["entry_price"] = pd.to_numeric(df["current"], errors="coerce")

    session_start, session_end = _session_bounds(run_date)
    next_day = _next_trading_day(run_date)
    next_session_end = datetime.combine(next_day, MARKET_CLOSE_CT, tzinfo=LOCAL_TZ)
    next_run_date = next_day.strftime("%Y-%m-%d")

    next_rows = []

    for _, r in df.iterrows():
        sym = str(r.get("symbol", "")).upper().strip()
        entry = _safe_float(r.get("entry_price"))
        target = _safe_float(r.get("target_price"))
        stop = _safe_float(r.get("stop_loss"))

        reco_dt = _parse_ts_maybe(r.get("run_ts")) or session_start
        if reco_dt < session_start:
            reco_dt = session_start
        if reco_dt.tzinfo is None:
            reco_dt = reco_dt.replace(tzinfo=LOCAL_TZ)

        res = _evaluate_window(
            symbol=sym,
            entry=entry,
            target=target,
            stop=stop,
            reco_dt=reco_dt,
            window_end=next_session_end,
            close_run_date=next_run_date,
        )

        next_rows.append({
            "symbol": sym,
            "next_session_end": next_session_end.strftime("%Y-%m-%d %H:%M:%S"),
            "next_close_price": res["close_price"],
            "next_actual_change_pct_close": res["actual_change_pct_close"],
            "next_target_hit": res["target_hit"],
            "next_stop_hit": res["stop_hit"],
            "next_first_hit": res["first_hit"],
            "next_first_hit_time": res["first_hit_time"],
            "next_hit_latency_minutes": res["hit_latency_minutes"],
            "next_target_overshoot_pct": res["target_overshoot_pct"],
            "next_best_exit_price_after_target": res["best_exit_price_after_target"],
            "next_best_exit_time_after_target": res["best_exit_time_after_target"],
            "next_best_exit_from_entry_pct": res["best_exit_from_entry_pct"],
            "next_best_exit_from_target_pct": res["best_exit_from_target_pct"],
            "next_best_exit_latency_minutes": res["best_exit_latency_minutes"],
            "next_outcome": res["outcome"],
        })

    return pd.DataFrame(next_rows)

def backfill_prior_midday_next_session(perf_log_csv: str, daily_log_csv: str, run_date: str) -> None:
    """
    On today's postmarket run, backfill yesterday's midday rows with completed next-session results.
    Example:
      - run_date = 2026-03-18
      - backfill midday rows from 2026-03-17
      - next-session window ends today (2026-03-18 close)
    """
    try:
        if (not os.path.exists(perf_log_csv)) or os.path.getsize(perf_log_csv) == 0:
            return
        if (not os.path.exists(daily_log_csv)) or os.path.getsize(daily_log_csv) == 0:
            return

        perf = pd.read_csv(perf_log_csv)
        daily = pd.read_csv(daily_log_csv)

        if perf.empty or daily.empty:
            return

        today = pd.to_datetime(run_date, errors="coerce")
        if pd.isna(today):
            return

        # previous trading day
        prev_day = today.date()
        while True:
            prev_day = prev_day - timedelta(days=1)
            if prev_day.weekday() < 5:
                break
        prev_run_date = prev_day.strftime("%Y-%m-%d")

        # load yesterday midday source rows from DAILY_LOG
        daily_mid = daily.copy()
        daily_mid = daily_mid[daily_mid.get("run_date", "").astype(str) == prev_run_date]
        if "mode" in daily_mid.columns:
            daily_mid = daily_mid[daily_mid["mode"].astype(str).str.lower() == "midday"].copy()

        if daily_mid.empty:
            return

        # latest per symbol
        if "run_ts" in daily_mid.columns:
            daily_mid["run_ts"] = daily_mid["run_ts"].apply(_parse_ts_maybe)
            daily_mid = daily_mid.sort_values("run_ts").drop_duplicates(subset=["symbol"], keep="last")

        daily_mid["source_mode"] = "midday"

        # compute completed next-session using prev_run_date as source date
        next_eval = evaluate_midday_next_session(daily_mid, prev_run_date)
        if next_eval.empty:
            return

        perf = _ensure_cols(perf, PERF_COLS)

        # match ONLY yesterday midday rows in performance_log
        mask = (
            perf["run_date"].astype(str).str[:10].eq(prev_run_date) &
            perf["source_mode"].astype(str).str.lower().eq("midday")
        )

        if not mask.any():
            return

        next_eval["symbol"] = next_eval["symbol"].astype(str).str.upper().str.strip()
        perf["symbol"] = perf["symbol"].astype(str).str.upper().str.strip()

        next_map = next_eval.set_index("symbol").to_dict(orient="index")

        next_cols = [
            "next_session_end",
            "next_close_price",
            "next_actual_change_pct_close",
            "next_target_hit",
            "next_stop_hit",
            "next_first_hit",
            "next_first_hit_time",
            "next_hit_latency_minutes",
            "next_target_overshoot_pct",
            "next_best_exit_price_after_target",
            "next_best_exit_time_after_target",
            "next_best_exit_from_entry_pct",
            "next_best_exit_from_target_pct",
            "next_best_exit_latency_minutes",
            "next_outcome",
        ]

        for c in next_cols:
            if c in perf.columns:
                perf[c] = perf[c].astype("object")

        for i in perf[mask].index:
            sym = perf.at[i, "symbol"]
            row = next_map.get(sym)
            if not row:
                continue
            for c in next_cols:
                perf.at[i, c] = row.get(c, pd.NA)

        perf.to_csv(perf_log_csv, index=False)

    except Exception as e:
        print("⚠️ prior midday next-session backfill failed:", e)

def llm_daily_narrative(run_date: str, prem_s: dict, mid_s: dict, all_s: dict) -> str:
    prompt = f"""
You are a cautious trading performance analyst.

Write a short post-market narrative using ONLY these stats.
No financial advice. No made-up numbers. Keep it concise.

Stats:
run_date: {run_date}
premarket: {prem_s}
midday: {mid_s}
combined: {all_s}

Output format (exact):
- What happened:
- What worked:
- What didn’t:
- Tomorrow tweak:

Return exactly 4 lines, each starting with the bullet label above. No extra lines.
"""
    try:
        return llm_text(prompt, max_output_tokens=220).strip()
    except Exception as e:
        return f"LLM unavailable ({type(e).__name__})"


def _summarize_eval(df: pd.DataFrame, outcome_col: str = "outcome") -> Dict[str, float]:
    if df is None or df.empty or outcome_col not in df.columns:
        return {
            "evaluated": 0,
            "wins": 0,
            "late_wins": 0,
            "losses": 0,
            "not_hit": 0,
            "win_rate": 0.0,
        }

    valid = ["🏆 Target Hit", "🟡 Late Win", "🛑 Stop Hit", "⏳ Not Hit"]
    eval_df = df[df[outcome_col].isin(valid)].copy()

    total = int(len(eval_df))
    wins = int((eval_df[outcome_col] == "🏆 Target Hit").sum()) if total else 0
    late_wins = int((eval_df[outcome_col] == "🟡 Late Win").sum()) if total else 0
    losses = int((eval_df[outcome_col] == "🛑 Stop Hit").sum()) if total else 0
    not_hit = int((eval_df[outcome_col] == "⏳ Not Hit").sum()) if total else 0

    # count late wins as wins for high-level effectiveness
    effective_wins = wins + late_wins
    rate = (effective_wins / total * 100.0) if total else 0.0

    return {
        "evaluated": total,
        "wins": wins,
        "late_wins": late_wins,
        "losses": losses,
        "not_hit": not_hit,
        "win_rate": rate,
    }


def _summarize_eval_by_instrument(df: pd.DataFrame, outcome_col: str = "outcome") -> Dict[str, Dict[str, float]]:
    out = {}
    if df is None or df.empty:
        return {"stock": _summarize_eval(pd.DataFrame(), outcome_col), "options": _summarize_eval(pd.DataFrame(), outcome_col)}

    for k in ["stock", "options"]:
        subset = df[df.get("instrument_type", "") == k].copy()
        out[k] = _summarize_eval(subset, outcome_col=outcome_col)
    return out


# -----------------------------
# NEW: Post-only deterministic Strategy Suggestions (no hallucinations)
# -----------------------------
def _stat_int(d: dict, k: str) -> int:
    try:
        return int(d.get(k, 0) or 0)
    except Exception:
        return 0


def _stat_float(d: dict, k: str) -> float:
    try:
        return float(d.get(k, 0.0) or 0.0)
    except Exception:
        return 0.0


def compute_strategy_suggestions_from_stats(run_date: str, prem_s: dict, mid_s: dict, all_s: dict) -> dict:
    evaluated = _stat_int(all_s, "evaluated")
    wins = _stat_int(all_s, "wins")
    losses = _stat_int(all_s, "losses")
    not_hit = _stat_int(all_s, "not_hit")
    win_rate = _stat_float(all_s, "win_rate")

    stop_rate = (losses / evaluated * 100.0) if evaluated else 0.0
    not_hit_rate = (not_hit / evaluated * 100.0) if evaluated else 0.0

    prem_eval = _stat_int(prem_s, "evaluated")
    mid_eval = _stat_int(mid_s, "evaluated")
    prem_wr = _stat_float(prem_s, "win_rate")
    mid_wr = _stat_float(mid_s, "win_rate")

    suggestions = []

    if evaluated == 0:
        confidence_label = "not provided"
        suggestions.append({
            "observation": "No evaluated trades found for today.",
            "why": "not provided",
            "change_in_pre_mid": "not provided",
            "risk_control": "not provided",
        })
    else:
        if evaluated >= 12:
            confidence_label = "high"
        elif evaluated >= 6:
            confidence_label = "medium"
        else:
            confidence_label = "low"

        if stop_rate >= 40.0:
            suggestions.append({
                "observation": f"Stop-hit rate is elevated ({stop_rate:.2f}%).",
                "why": "Stops may be tight relative to realized volatility, or confidence gates may be too permissive.",
                "change_in_pre_mid": "Consider raising the weaker runner’s confidence gate by +1; filter more noise (e.g., stricter mover threshold).",
                "risk_control": "Reduce position size on lower-confidence picks until stop rate improves.",
            })

        if not_hit_rate >= 55.0:
            suggestions.append({
                "observation": f"Not-hit rate is high ({not_hit_rate:.2f}%).",
                "why": "Targets may be too far for the available time window (especially for midday).",
                "change_in_pre_mid": "Strengthen midday time-scaling or skip very late recommendations.",
                "risk_control": "Cap number of midday picks when time-left is low.",
            })

        if win_rate >= 55.0 and evaluated >= 5:
            suggestions.append({
                "observation": f"Win rate is solid ({win_rate:.2f}%) with evaluated={evaluated}.",
                "why": "Gates/targets appear aligned with today’s price action.",
                "change_in_pre_mid": "Keep parameters stable; only adjust if weekly stats disagree.",
                "risk_control": "Maintain sizing discipline; avoid expanding into lower-confidence names.",
            })

        if prem_eval >= 3 and mid_eval >= 3 and (mid_wr + 10.0 < prem_wr):
            suggestions.append({
                "observation": f"Midday underperforms premarket (midday={mid_wr:.2f}%, premarket={prem_wr:.2f}%).",
                "why": "Late timing reduces chance of reaching target; sudden movers can mean-revert.",
                "change_in_pre_mid": "Raise MIDDAY_MIN_CONFIDENCE by +1 OR increase SUDDEN_MOVER_PCT_THRESHOLD to reduce noisy movers.",
                "risk_control": "Prefer tighter time-scaled target/stop for midday.",
            })

        if not suggestions:
            suggestions.append({
                "observation": "No strong issues detected from today’s summary.",
                "why": "not provided",
                "change_in_pre_mid": "Keep parameters stable.",
                "risk_control": "Maintain existing risk controls.",
            })

    return {
        "run_date": run_date,
        "confidence": confidence_label,
        "kpis": {
            "evaluated": evaluated,
            "wins": wins,
            "losses": losses,
            "not_hit": not_hit,
            "win_rate": round(win_rate, 2) if evaluated else 0.0,
            "stop_rate": round(stop_rate, 2) if evaluated else 0.0,
            "not_hit_rate": round(not_hit_rate, 2) if evaluated else 0.0,
        },
        "inputs": {"premarket": prem_s, "midday": mid_s, "combined": all_s},
        "suggestions": suggestions,
    }


def render_strategy_suggestions_html(obj: dict) -> str:
    conf = _html.escape(str(obj.get("confidence", "not provided")))
    kpis = obj.get("kpis", {}) or {}
    sug = obj.get("suggestions", []) or []

    kpi_html = "<br>".join(
        [f"<b>{_html.escape(str(k))}:</b> {_html.escape(str(v))}" for k, v in kpis.items()]
    )

    rows = ""
    for s in sug:
        rows += f"""
        <tr>
          <td>{_html.escape(str(s.get("observation", "")))}</td>
          <td>{_html.escape(str(s.get("why", "")))}</td>
          <td>{_html.escape(str(s.get("change_in_pre_mid", "")))}</td>
          <td>{_html.escape(str(s.get("risk_control", "")))}</td>
        </tr>
        """

    return f"""
    <h3>🧩 Strategy Suggestions</h3>
    <p><b>Suggestion confidence:</b> {conf}</p>
    <p>{kpi_html}</p>
    <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;font-family:Arial;font-size:13px;">
      <tr style="background:#eee;">
        <th>Observation</th><th>Why</th><th>Change in Pre/Mid</th><th>Risk control</th>
      </tr>
      {rows}
    </table>
    """


def save_strategy_suggestions_json(obj: dict, rf: Path) -> None:
    try:
        p = Path(rf) / f"strategy_suggestions_{obj.get('run_date', '')}.json"
        p.write_text(json.dumps(obj, indent=2), encoding="utf-8")
    except Exception:
        pass


# -----------------------------
# Load sources
# -----------------------------
def load_premarket_today(daily_log_csv: str, run_date: str) -> pd.DataFrame:
    """
    Source of truth: DAILY_LOG
    - premarket rows are mode == 'premarket'
    """
    if (not os.path.exists(daily_log_csv)) or os.path.getsize(daily_log_csv) == 0:
        return pd.DataFrame()

    df = pd.read_csv(daily_log_csv)
    if df.empty or "run_date" not in df.columns:
        return pd.DataFrame()

    df = df[df["run_date"].astype(str) == str(run_date)].copy()
    if df.empty:
        return pd.DataFrame()

    if "mode" in df.columns:
        df = df[df["mode"].astype(str).str.lower() == "premarket"].copy()

    df["source_mode"] = "premarket"
    return df


def load_midday_today(daily_log_csv: str, run_date: str) -> pd.DataFrame:
    """
    Source of truth: DAILY_LOG
    - midday rows are mode == 'midday'
    This matches the new locked midday_runner which writes FINAL_VIEW into daily log.
    """
    if (not os.path.exists(daily_log_csv)) or os.path.getsize(daily_log_csv) == 0:
        return pd.DataFrame()

    df = pd.read_csv(daily_log_csv)
    if df.empty or "run_date" not in df.columns:
        return pd.DataFrame()

    df = df[df["run_date"].astype(str) == str(run_date)].copy()
    if df.empty:
        return pd.DataFrame()

    if "mode" in df.columns:
        df = df[df["mode"].astype(str).str.lower() == "midday"].copy()

    # pick latest per symbol for the day (if duplicates exist)
    if "run_ts" in df.columns:
        df["run_ts"] = df["run_ts"].apply(_parse_ts_maybe)
        df = df.sort_values("run_ts").drop_duplicates(subset=["symbol"], keep="last")

    df["source_mode"] = "midday"
    return df


# -----------------------------
# Evaluation (TARGET HIT logic + analytics)
# -----------------------------
def evaluate_rows(df_in: pd.DataFrame, run_date: str) -> pd.DataFrame:
    if df_in is None or df_in.empty:
        return pd.DataFrame()

    df = df_in.copy()
    df = _ensure_cols(
        df,
        ["symbol", "decision", "score", "confidence", "current", "target_price", "stop_loss",
         "source_mode", "instrument_type", "run_ts"],
    )

    df["symbol"] = df["symbol"].astype(str).str.upper().str.strip()
    df["decision"] = df["decision"].astype(str)
    df["source_mode"] = df["source_mode"].astype(str)

    df["instrument_type"] = [
        _classify_instrument(sym, row=r) for sym, (_, r) in zip(df["symbol"].tolist(), df.iterrows())
    ]

    df["entry_price"] = pd.to_numeric(df["current"], errors="coerce")

    session_start, session_end = _session_bounds(run_date)

    # Build reco_ts from run_ts when available, otherwise session_start
    reco_ts_list = []
    for _, r in df.iterrows():
        rt = None
        if "run_ts" in df.columns:
            rt = _parse_ts_maybe(r.get("run_ts"))
        reco_ts_list.append(rt if rt is not None else session_start)

    df["reco_ts"] = reco_ts_list

    # ✅ HARDEN: ensure reco_ts is always tz-aware datetime (never NaT)
    df["reco_ts"] = df["reco_ts"].apply(lambda x: _parse_ts_maybe(x) or session_start)

    df["session_start"] = session_start.strftime("%Y-%m-%d %H:%M:%S")
    df["session_end"] = session_end.strftime("%Y-%m-%d %H:%M:%S")

    day_highs, day_lows = [], []
    closes, pct_close = [], []
    target_hits, stop_hits = [], []
    first_hits, first_hit_times = [], []
    hit_latency_minutes = []
    target_overshoot_pct = []

    best_exit_price_after_target = []
    best_exit_time_after_target = []
    best_exit_from_entry_pct = []
    best_exit_from_target_pct = []
    best_exit_latency_minutes = []

    outcomes = []

    for _, r in df.iterrows():
        sym = str(r.get("symbol", "")).upper().strip()
        entry = _safe_float(r.get("entry_price"))
        target = _safe_float(r.get("target_price"))
        stop = _safe_float(r.get("stop_loss"))

        # ✅ Use reco_ts (not reco_dt)
        reco_dt = r.get("reco_ts")
        if isinstance(reco_dt, str):
            reco_dt = _parse_ts_maybe(reco_dt)
        if reco_dt is None:
            reco_dt = session_start
        if reco_dt.tzinfo is None:
            reco_dt = reco_dt.replace(tzinfo=LOCAL_TZ)

        if reco_dt < session_start:
            reco_dt = session_start

        if reco_dt > session_end:
            # Local testing override: treat late-run recommendations as if they were made at session_start
            if IS_LOCAL and EVAL_CLAMP_LATE_RECO:
                reco_dt = session_start
            else:
                outcomes.append("⛔ Skipped (Late Recommendation)")
                day_highs.append(pd.NA)
                day_lows.append(pd.NA)
                closes.append(_get_close_for_date(sym, run_date))
                pct_close.append(pd.NA)
                target_hits.append(False)
                stop_hits.append(False)
                first_hits.append("SKIP")
                first_hit_times.append("reco_after_close")
                hit_latency_minutes.append(pd.NA)
                target_overshoot_pct.append(pd.NA)
                best_exit_price_after_target.append(pd.NA)
                best_exit_time_after_target.append("")
                best_exit_from_entry_pct.append(pd.NA)
                best_exit_from_target_pct.append(pd.NA)
                best_exit_latency_minutes.append(pd.NA)
                continue

        close_price = _get_close_for_date(sym, run_date)
        closes.append(close_price)

        if close_price is None or entry is None or entry <= 0:
            pct_close.append(pd.NA)
        else:
            pct_close.append(((close_price - entry) / entry) * 100.0)

        intraday = _intraday_after_reco(sym, reco_dt, session_end)

        th = False
        sh = False
        fh = ""
        fht_str = ""
        latency = pd.NA
        overshoot = pd.NA

        best_px = pd.NA
        best_ts_str = ""
        best_from_entry = pd.NA
        best_from_target = pd.NA
        best_latency = pd.NA

        if intraday is None or intraday.empty:
            day_highs.append(pd.NA)
            day_lows.append(pd.NA)

            if close_price is None:
                target_hits.append(False)
                stop_hits.append(False)
                first_hits.append("NO_DATA")
                first_hit_times.append("no_intraday_no_close")
                hit_latency_minutes.append(pd.NA)
                target_overshoot_pct.append(pd.NA)
                best_exit_price_after_target.append(pd.NA)
                best_exit_time_after_target.append("")
                best_exit_from_entry_pct.append(pd.NA)
                best_exit_from_target_pct.append(pd.NA)
                best_exit_latency_minutes.append(pd.NA)
                outcomes.append("⛔ No Data (Market Closed / Holiday)")
                continue

            if not EVAL_FALLBACK_STRICT:
                if close_price is not None and target is not None and target > 0 and close_price >= target:
                    th = True
                    fh = "TARGET"
                    fht_str = "close_fallback"
                elif close_price is not None and stop is not None and stop > 0 and close_price <= stop:
                    sh = True
                    fh = "STOP"
                    fht_str = "close_fallback"

            target_hits.append(th)
            stop_hits.append(sh)
            first_hits.append(fh)
            first_hit_times.append(fht_str)
            hit_latency_minutes.append(latency)
            target_overshoot_pct.append(overshoot)
            best_exit_price_after_target.append(best_px)
            best_exit_time_after_target.append(best_ts_str)
            best_exit_from_entry_pct.append(best_from_entry)
            best_exit_from_target_pct.append(best_from_target)
            best_exit_latency_minutes.append(best_latency)
            outcomes.append(_compute_outcome_from_hits(th, sh))
            continue

        hi = pd.to_numeric(intraday.get("High"), errors="coerce").dropna() if "High" in intraday.columns else pd.Series(dtype=float)
        lo = pd.to_numeric(intraday.get("Low"), errors="coerce").dropna() if "Low" in intraday.columns else pd.Series(dtype=float)

        max_high = float(hi.max()) if not hi.empty else None
        min_low = float(lo.min()) if not lo.empty else None

        day_highs.append(max_high if max_high is not None else pd.NA)
        day_lows.append(min_low if min_low is not None else pd.NA)

        has_target = target is not None and target > 0
        has_stop = stop is not None and stop > 0

        if not has_target and not has_stop:
            target_hits.append(False)
            stop_hits.append(False)
            first_hits.append("")
            first_hit_times.append("")
            hit_latency_minutes.append(pd.NA)
            target_overshoot_pct.append(pd.NA)
            best_exit_price_after_target.append(pd.NA)
            best_exit_time_after_target.append("")
            best_exit_from_entry_pct.append(pd.NA)
            best_exit_from_target_pct.append(pd.NA)
            best_exit_latency_minutes.append(pd.NA)
            outcomes.append("⏳ Not Hit")
            continue

        first_hit_dt: Optional[datetime] = None

        for ts, bar in intraday.iterrows():
            bar_high = _safe_float(bar.get("High"))
            bar_low = _safe_float(bar.get("Low"))

            hit_target_now = has_target and (bar_high is not None) and (bar_high >= target)
            hit_stop_now = has_stop and (bar_low is not None) and (bar_low <= stop)

            if hit_target_now and hit_stop_now:
                if str(CONSERVATIVE_SAME_BAR_POLICY).strip().lower() == "target_first":
                    th = True
                    fh = "TARGET"
                else:
                    sh = True
                    fh = "STOP"
                first_hit_dt = ts if isinstance(ts, datetime) else None
                break

            if hit_target_now:
                th = True
                fh = "TARGET"
                first_hit_dt = ts if isinstance(ts, datetime) else None
                break

            if hit_stop_now:
                sh = True
                fh = "STOP"
                first_hit_dt = ts if isinstance(ts, datetime) else None
                break

        if first_hit_dt is not None:
            try:
                first_hit_dt = first_hit_dt.astimezone(LOCAL_TZ)
                fht_str = first_hit_dt.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                fht_str = str(first_hit_dt)

        if first_hit_dt is not None and isinstance(reco_dt, datetime):
            try:
                latency = round((first_hit_dt - reco_dt).total_seconds() / 60.0, 2)
                if latency < 0:
                    latency = 0.0
            except Exception:
                latency = pd.NA

        if th and has_target and max_high is not None and target and target > 0:
            overshoot = round(((max_high - target) / target) * 100.0, 3)

        if th and first_hit_dt is not None and "High" in intraday.columns:
            after = intraday[intraday.index >= first_hit_dt].copy()
            if not after.empty:
                after_hi = pd.to_numeric(after["High"], errors="coerce").dropna()
                if not after_hi.empty:
                    best_val = float(after_hi.max())
                    best_px = best_val
                    try:
                        best_time = after_hi.idxmax()
                        if isinstance(best_time, datetime):
                            best_time = best_time.astimezone(LOCAL_TZ)
                            best_ts_str = best_time.strftime("%Y-%m-%d %H:%M:%S")
                            if isinstance(reco_dt, datetime):
                                best_latency = round((best_time - reco_dt).total_seconds() / 60.0, 2)
                                if best_latency < 0:
                                    best_latency = 0.0
                    except Exception:
                        best_ts_str = ""

                    if entry is not None and entry > 0:
                        best_from_entry = round(((best_val - entry) / entry) * 100.0, 3)
                    if target is not None and target > 0:
                        best_from_target = round(((best_val - target) / target) * 100.0, 3)

        target_hits.append(th)
        stop_hits.append(sh)
        first_hits.append(fh)
        first_hit_times.append(fht_str)
        hit_latency_minutes.append(latency)
        target_overshoot_pct.append(overshoot)
        best_exit_price_after_target.append(best_px)
        best_exit_time_after_target.append(best_ts_str)
        best_exit_from_entry_pct.append(best_from_entry)
        best_exit_from_target_pct.append(best_from_target)
        best_exit_latency_minutes.append(best_latency)
        outcomes.append(_compute_outcome_from_hits(th, sh))

    df["target_price"] = pd.to_numeric(df["target_price"], errors="coerce")
    df["stop_loss"] = pd.to_numeric(df["stop_loss"], errors="coerce")

    df["day_high_after_reco"] = day_highs
    df["day_low_after_reco"] = day_lows
    df["close_price"] = closes
    df["actual_change_pct_close"] = pct_close

    df["target_hit"] = target_hits
    df["stop_hit"] = stop_hits
    df["first_hit"] = first_hits
    df["first_hit_time"] = first_hit_times
    df["hit_latency_minutes"] = hit_latency_minutes
    df["target_overshoot_pct"] = target_overshoot_pct
    df["best_exit_price_after_target"] = best_exit_price_after_target
    df["best_exit_time_after_target"] = best_exit_time_after_target
    df["best_exit_from_entry_pct"] = best_exit_from_entry_pct
    df["best_exit_from_target_pct"] = best_exit_from_target_pct
    df["best_exit_latency_minutes"] = best_exit_latency_minutes

    df["outcome"] = outcomes
    df["run_date"] = run_date

    cols_order = PERF_COLS[:]
    df = _ensure_cols(df, cols_order)
    return df[cols_order]


def append_perf_log(out_df: pd.DataFrame, now: Optional[datetime] = None) -> None:
    """
    Bulletproof perf append:
    - backfills entry_price from current if missing
    - writes optional debug artifacts before/after dropna
    """
    try:
        if out_df is None or out_df.empty:
            return

        now = now or datetime.now(LOCAL_TZ)
        rf = run_dir(now, "postmarket")

        out = out_df.copy()
        out = _ensure_cols(out, PERF_COLS + ["current"])

        out["entry_price"] = pd.to_numeric(out.get("entry_price"), errors="coerce")
        cur = pd.to_numeric(out.get("current"), errors="coerce")
        out["entry_price"] = out["entry_price"].fillna(cur)

        out = _ensure_cols(out, PERF_COLS)[PERF_COLS]

        required = ["run_date", "source_mode", "symbol", "decision", "score", "confidence", "entry_price"]
        for c in required:
            if c in out.columns:
                out[c] = out[c].replace("", pd.NA)

        if POSTMARKET_DEBUG:
            out.to_csv(Path(rf) / "perf_debug_before_dropna.csv", index=False)

        out2 = out.dropna(subset=[c for c in required if c in out.columns])

        if POSTMARKET_DEBUG:
            out2.to_csv(Path(rf) / "perf_debug_after_dropna.csv", index=False)

        if out2.empty:
            (Path(rf) / "perf_append_skipped.txt").write_text(
                "append_perf_log skipped: all rows dropped by required fields.\n"
                f"required={required}\n",
                encoding="utf-8",
            )
            return

        file_exists = os.path.exists(PERF_LOG_CSV) and os.path.getsize(PERF_LOG_CSV) > 0
        if file_exists:
            prev = pd.read_csv(PERF_LOG_CSV)
            prev = _ensure_cols(prev, PERF_COLS)[PERF_COLS]
            frames = [df for df in (prev, out2) if df is not None and not df.empty]
            merged = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        else:
            merged = out2.copy()

        merged = merged.drop_duplicates(subset=["run_date", "source_mode", "symbol"], keep="last")
        merged.to_csv(PERF_LOG_CSV, index=False)

    except Exception as e:
        print("⚠️ PERF log append failed:", e)


def build_weekly_dashboard_html(perf_log_csv: str, now: datetime) -> str:
    try:
        if not os.path.exists(perf_log_csv) or os.path.getsize(perf_log_csv) == 0:
            return "<h2>📅 Weekly Dashboard</h2><p>No performance log yet.</p>"

        df = pd.read_csv(perf_log_csv)
        if df.empty or "run_date" not in df.columns:
            return "<h2>📅 Weekly Dashboard</h2><p>Performance log empty/invalid.</p>"

        df = normalize_perf_df(df)

        if "source_mode" in df.columns:
            df["source_mode"] = df["source_mode"].astype(str).str.strip().str.lower()
            df["source_mode"] = df["source_mode"].replace({"nan": "", "none": ""})
            df["source_mode"] = df["source_mode"].replace("", pd.NA).fillna("unknown")

        week_ago = (now - timedelta(days=7)).replace(tzinfo=None)

        if "outcome" in df.columns:
            if "target_hit" in df.columns:
                df.loc[df["outcome"] == "🏆 Target Hit", "target_hit"] = True
            if "stop_hit" in df.columns:
                df.loc[df["outcome"] == "🛑 Stop Hit", "stop_hit"] = True

        if "next_outcome" in df.columns:
            if "next_target_hit" in df.columns:
                df.loc[df["next_outcome"] == "🏆 Target Hit", "next_target_hit"] = True
            if "next_stop_hit" in df.columns:
                df.loc[df["next_outcome"] == "🛑 Stop Hit", "next_stop_hit"] = True

        df["run_date"] = pd.to_datetime(df["run_date"], errors="coerce")

        d7 = df[df["run_date"].notna() & (df["run_date"] >= week_ago)].copy()

        if d7.empty:
            return "<h2>📅 Weekly Dashboard</h2><p>No rows in last 7 days.</p>"

        all_s = _summarize_eval(d7)
        prem_s = _summarize_eval(d7[d7["source_mode"] == "premarket"].copy())
        mid_s = _summarize_eval(d7[d7["source_mode"] == "midday"].copy())
        mid_next_s = _summarize_eval(d7[d7["source_mode"] == "midday"].copy(), outcome_col="next_outcome")

        all_by = _summarize_eval_by_instrument(d7)
        prem_by = _summarize_eval_by_instrument(d7[d7["source_mode"] == "premarket"].copy())
        mid_by = _summarize_eval_by_instrument(d7[d7["source_mode"] == "midday"].copy())
        mid_next_by = _summarize_eval_by_instrument(d7[d7["source_mode"] == "midday"].copy(), outcome_col="next_outcome")

        weekly_narrative = ""
        weekly_coach = ""
        LLM_ENABLED = os.getenv("LLM_ENABLED", "1") == "1"
        if LLM_ENABLED:
            weekly_narrative = safe_llm_weekly_narrative(now, prem_s, mid_s, mid_next_s, all_s)
            weekly_coach = safe_llm_weekly_coach(now, prem_s, mid_s, mid_next_s, all_s)

        html = f"""
        <h2>📅 Weekly Trading Dashboard ({now.strftime('%Y-%m-%d')})</h2>

        <p><b>ALL same-day (7d):</b> evaluated={all_s["evaluated"]}, wins={all_s["wins"]}, losses={all_s["losses"]}, not_hit={all_s["not_hit"]}, win_rate={all_s["win_rate"]:.2f}%</p>
        <p><b>Premarket same-day (7d):</b> evaluated={prem_s["evaluated"]}, win_rate={prem_s["win_rate"]:.2f}%<br>
           <b>Midday same-day (7d):</b> evaluated={mid_s["evaluated"]}, win_rate={mid_s["win_rate"]:.2f}%<br>
           <b>Midday next-session (7d):</b> evaluated={mid_next_s["evaluated"]}, win_rate={mid_next_s["win_rate"]:.2f}%</p>

        <h3>📦 Instrument Buckets (7d)</h3>
        <p><b>ALL same-day - Stock:</b> eval={all_by["stock"]["evaluated"]}, win_rate={all_by["stock"]["win_rate"]:.2f}% |
           <b>Options:</b> eval={all_by["options"]["evaluated"]}, win_rate={all_by["options"]["win_rate"]:.2f}%</p>
        <p><b>Premarket same-day - Stock:</b> eval={prem_by["stock"]["evaluated"]}, win_rate={prem_by["stock"]["win_rate"]:.2f}% |
           <b>Options:</b> eval={prem_by["options"]["evaluated"]}, win_rate={prem_by["options"]["win_rate"]:.2f}%</p>
        <p><b>Midday same-day - Stock:</b> eval={mid_by["stock"]["evaluated"]}, win_rate={mid_by["stock"]["win_rate"]:.2f}% |
           <b>Options:</b> eval={mid_by["options"]["evaluated"]}, win_rate={mid_by["options"]["win_rate"]:.2f}%</p>
        <p><b>Midday next-session - Stock:</b> eval={mid_next_by["stock"]["evaluated"]}, win_rate={mid_next_by["stock"]["win_rate"]:.2f}% |
           <b>Options:</b> eval={mid_next_by["options"]["evaluated"]}, win_rate={mid_next_by["options"]["win_rate"]:.2f}%</p>

        <p>Win definition: target price hit at any time after recommendation (intraday high ≥ target).</p>
        """

        if weekly_narrative:
            html += f"<h3>🧠 Weekly LLM Narrative</h3><pre>{_html.escape(weekly_narrative)}</pre>"

        if weekly_coach:
            html += f"<h3>🧠 Weekly LLM Coach</h3><pre>{_html.escape(weekly_coach)}</pre>"

        return html
    except Exception:
        return "<h2>📅 Weekly Dashboard</h2><p>Dashboard generation failed.</p>"


def _excel_safe_df(df: pd.DataFrame, local_tz: str = "America/Chicago") -> pd.DataFrame:
    if df is None or df.empty:
        return df

    out = df.copy()

    for col in out.columns:
        if isinstance(out[col].dtype, DatetimeTZDtype):
            out[col] = out[col].dt.tz_convert(local_tz).dt.tz_localize(None)

    def _strip_tz(v):
        if isinstance(v, pd.Timestamp) and v.tz is not None:
            return v.tz_convert(local_tz).tz_localize(None)
        return v

    obj_cols = out.select_dtypes(include=["object"]).columns
    for col in obj_cols:
        sample = out[col].dropna().head(50).tolist()
        if any(isinstance(v, pd.Timestamp) and v.tz is not None for v in sample):
            out[col] = out[col].apply(_strip_tz)

    return out


# -----------------------------
# Portfolio wrapper (unchanged)
# -----------------------------
def portfolio_update_and_close(run_dt: datetime) -> Tuple[str, pd.DataFrame]:
    cfg = PortfolioConfig()
    open_df = load_open_portfolio(cfg)
    remaining, closed_df = update_and_close_positions(cfg, open_df, run_dt)
    save_open_portfolio(cfg, remaining)

    if closed_df is not None and not closed_df.empty:
        closed_hist = closed_df.copy()
        closed_hist["action"] = "CLOSE"
        append_trade_history(cfg, closed_hist)

    summary = portfolio_summary(remaining, closed_df)
    return summary, closed_df


def normalize_perf_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    df = df.copy()
    df = _ensure_cols(df, PERF_COLS)

    df["run_date"] = pd.to_datetime(df["run_date"], errors="coerce").dt.normalize()

    if "source_mode" in df.columns:
        df["source_mode"] = df["source_mode"].astype(str).str.strip().str.lower()
        df["source_mode"] = df["source_mode"].replace({"nan": "", "none": ""})
        df["source_mode"] = df["source_mode"].replace("", pd.NA).fillna("unknown")

    if "instrument_type" in df.columns and "symbol" in df.columns:
        df["instrument_type"] = df["instrument_type"].replace("", pd.NA)
        mask = df["instrument_type"].isna()
        df.loc[mask, "instrument_type"] = [
            _classify_instrument(sym) for sym in df.loc[mask, "symbol"].astype(str).tolist()
        ]

    for i, r in df.iterrows():
        if pd.isna(r.get("run_date")):
            continue
        rd = r["run_date"].date().strftime("%Y-%m-%d")
        ss, se = _session_bounds(rd)

        if not str(r.get("session_start") or "").strip():
            df.at[i, "session_start"] = ss.strftime("%Y-%m-%d %H:%M:%S")
        if not str(r.get("session_end") or "").strip():
            df.at[i, "session_end"] = se.strftime("%Y-%m-%d %H:%M:%S")
        if not str(r.get("reco_ts") or "").strip():
            df.at[i, "reco_ts"] = ss.strftime("%Y-%m-%d %H:%M:%S")

    legacy_map = {"✅ Correct": "🏆 Target Hit", "❌ Incorrect": "🛑 Stop Hit"}
    df["outcome"] = df["outcome"].replace(legacy_map)

    for b in ["target_hit", "stop_hit", "next_target_hit", "next_stop_hit"]:
        if b in df.columns:
            s = df[b].astype(str).str.strip().str.upper()
            df[b] = s.map({"TRUE": True, "FALSE": False})
            df[b] = df[b].where(pd.notna(df[b]), False).astype(bool)

    if "outcome" in df.columns:
        if "target_hit" in df.columns:
            df.loc[df["outcome"] == "🏆 Target Hit", "target_hit"] = True
        if "stop_hit" in df.columns:
            df.loc[df["outcome"] == "🛑 Stop Hit", "stop_hit"] = True

    if "next_outcome" in df.columns:
        if "next_target_hit" in df.columns:
            df.loc[df["next_outcome"] == "🏆 Target Hit", "next_target_hit"] = True
        if "next_stop_hit" in df.columns:
            df.loc[df["next_outcome"] == "🛑 Stop Hit", "next_stop_hit"] = True

    if "combined_outcome" in df.columns:
        df["combined_outcome"] = df["combined_outcome"].replace({"": pd.NA}).fillna(
            df.get("outcome", pd.Series(dtype="object"))
        )

    if "recovered_after_stop" in df.columns:
        s = df["recovered_after_stop"].astype(str).str.strip().str.upper()
        df["recovered_after_stop"] = s.map({"TRUE": True, "FALSE": False})
        df["recovered_after_stop"] = df["recovered_after_stop"].where(
            pd.notna(df["recovered_after_stop"]), False
        ).astype(bool)

    return df


# -----------------------------
# Weekly Excel builder (ALWAYS attaches)
# -----------------------------
def build_weekly_excel(perf_log_csv: str, now: datetime) -> Tuple[str, int, int]:
    weekly_excel = out_path(
        f"weekly_dashboard_{now.strftime('%Y%m%d')}_{make_run_id(now)}.xlsx",
        now=now,
        mode="postmarket",
        kind="runs",
    )

    if os.path.exists(perf_log_csv) and os.path.getsize(perf_log_csv) > 0:
        wdf = pd.read_csv(perf_log_csv)
    else:
        wdf = pd.DataFrame(columns=PERF_COLS)

    wdf = normalize_perf_df(wdf)

    if "run_date" in wdf.columns:
        wdf["run_date"] = pd.to_datetime(wdf["run_date"], errors="coerce")
        cutoff = (now - timedelta(days=7)).date()
        wdf7 = wdf[wdf["run_date"].notna() & (wdf["run_date"].dt.date >= cutoff)].copy()
    else:
        wdf7 = pd.DataFrame(columns=PERF_COLS)

    with pd.ExcelWriter(weekly_excel, engine="openpyxl") as writer:
        _excel_safe_df(wdf7).to_excel(writer, sheet_name="LAST_7_DAYS", index=False)
        _excel_safe_df(wdf).to_excel(writer, sheet_name="ALL_HISTORY", index=False)

    wb = load_workbook(weekly_excel)
    for s in wb.sheetnames:
        style_excel_sheet(wb[s])
    wb.save(weekly_excel)

    return weekly_excel, int(len(wdf7)), int(len(wdf))


def _baseline_narrative_from_stats(run_date: str, prem_s: dict, mid_s: dict, all_s: dict) -> str:
    evaluated = int(all_s.get("evaluated", 0) or 0)
    wins = int(all_s.get("wins", 0) or 0)
    losses = int(all_s.get("losses", 0) or 0)
    not_hit = int(all_s.get("not_hit", 0) or 0)
    win_rate = float(all_s.get("win_rate", 0.0) or 0.0)

    if evaluated <= 0:
        return "\n".join([
            "- What happened: No evaluated trades in stats.",
            "- What worked: not provided",
            "- What didn’t: not provided",
            "- Tomorrow tweak: Verify evaluation rows are being written.",
        ])

    return "\n".join([
        f"- What happened: evaluated={evaluated}, wins={wins}, losses={losses}, not_hit={not_hit}, win_rate={win_rate:.2f}%.",
        "- What worked: Wins indicate some follow-through to targets.",
        "- What didn’t: Not-hit/stop outcomes indicate incomplete follow-through or adverse moves.",
        "- Tomorrow tweak: Keep parameters stable unless this repeats across multiple sessions.",
    ])


def safe_llm_daily_narrative(run_date: str, prem_s: dict, mid_s: dict, all_s: dict) -> str:
    """
    Bulletproof narrative:
    - Skip LLM if nothing evaluated (saves tokens + avoids nonsense)
    - Fallback if LLM returns empty / tool-failure text
    - Fallback if output is weak ("not provided" spam)
    - Enforce strict 4-bullet structure (starts-with, in order)
    - Cap length to keep emails clean
    """
    # (2) Skip LLM entirely if nothing was evaluated
    if int(all_s.get("evaluated", 0) or 0) <= 0:
        return _baseline_narrative_from_stats(run_date, prem_s, mid_s, all_s)

    try:
        out = llm_daily_narrative(run_date, prem_s, mid_s, all_s)

        # Empty output -> baseline
        if not str(out or "").strip():
            return _baseline_narrative_from_stats(run_date, prem_s, mid_s, all_s)

        out_str = str(out).strip()

        # Tool-failure text -> baseline
        if out_str.lower().startswith("llm unavailable"):
            return _baseline_narrative_from_stats(run_date, prem_s, mid_s, all_s)

        # Weak output guard: too many "not provided"
        np_count = sum("not provided" in ln.lower() for ln in out_str.splitlines())
        if np_count >= 3:
            return _baseline_narrative_from_stats(run_date, prem_s, mid_s, all_s)

        # (1) Enforce strict structure: 4 bullets, in order, must START WITH the required prefixes
        lines = [ln.strip() for ln in out_str.splitlines() if ln.strip()]
        required = [
            "- What happened:",
            "- What worked:",
            "- What didn’t:",
            "- Tomorrow tweak:",
        ]
        if len(lines) < 4 or any(not lines[i].startswith(required[i]) for i in range(4)):
            return _baseline_narrative_from_stats(run_date, prem_s, mid_s, all_s)

        # Keep only the first 4 bullets (prevents extra rambling)
        out_clean = "\n".join(lines[:4]).strip()

        # Cap length (email safety)
        if len(out_clean) > 1200:
            out_clean = "\n".join(out_clean.splitlines()[:4]).strip()

        return out_clean

    except Exception:
        return _baseline_narrative_from_stats(run_date, prem_s, mid_s, all_s)

def llm_weekly_narrative(now: datetime, prem_s: dict, mid_s: dict, mid_next_s: dict, all_s: dict) -> str:
    prompt = f"""
You are a cautious trading performance analyst.

Write a short WEEKLY narrative using ONLY these stats.
No financial advice. No made-up numbers. Keep it concise.

Stats:
week_end: {now.strftime('%Y-%m-%d')}
premarket_same_day: {prem_s}
midday_same_day: {mid_s}
midday_next_session: {mid_next_s}
combined_same_day: {all_s}

Output format (exact):
- Weekly headline:
- What worked:
- What didn’t:
- Next-week focus:

Return exactly 4 lines, each starting with the bullet label above. No extra lines.
"""
    try:
        return llm_text(prompt, max_output_tokens=260).strip()
    except Exception as e:
        return f"LLM unavailable ({type(e).__name__})"


def safe_llm_weekly_narrative(now: datetime, prem_s: dict, mid_s: dict, mid_next_s: dict, all_s: dict) -> str:
    total = int(all_s.get("evaluated", 0) or 0) + int(mid_next_s.get("evaluated", 0) or 0)
    if total <= 0:
        return "\n".join([
            "- Weekly headline: No evaluated weekly rows found.",
            "- What worked: not provided",
            "- What didn’t: not provided",
            "- Next-week focus: Verify performance_log is writing correctly.",
        ])

    try:
        out = llm_weekly_narrative(now, prem_s, mid_s, mid_next_s, all_s)
        if not str(out or "").strip():
            raise ValueError("empty weekly narrative")

        lines = [ln.strip() for ln in str(out).splitlines() if ln.strip()]
        required = [
            "- Weekly headline:",
            "- What worked:",
            "- What didn’t:",
            "- Next-week focus:",
        ]
        if len(lines) < 4 or any(not lines[i].startswith(required[i]) for i in range(4)):
            raise ValueError("bad weekly narrative structure")

        return "\n".join(lines[:4]).strip()
    except Exception:
        return "\n".join([
            f"- Weekly headline: Premarket same-day win_rate={prem_s.get('win_rate', 0.0):.2f}%, midday same-day win_rate={mid_s.get('win_rate', 0.0):.2f}%, midday next-session win_rate={mid_next_s.get('win_rate', 0.0):.2f}%.",
            "- What worked: Premarket and next-session follow-through should be compared side by side.",
            "- What didn’t: Midday same-day performance appears weaker than desired.",
            "- Next-week focus: Keep monitoring midday next-session truth before changing generation logic.",
        ])


def llm_weekly_coach(now: datetime, prem_s: dict, mid_s: dict, mid_next_s: dict, all_s: dict) -> str:
    prompt = f"""
You are a cautious trading system coach.

Use ONLY the stats below. No made-up facts. No trading advice.
Give a short weekly coaching note for the system owner.

Stats:
week_end: {now.strftime('%Y-%m-%d')}
premarket_same_day: {prem_s}
midday_same_day: {mid_s}
midday_next_session: {mid_next_s}
combined_same_day: {all_s}

Output format (exact):
- Weekly coaching view:
- Key evidence:
- Risk note:
- Next action:

Return exactly 4 lines.
"""
    try:
        return llm_text(prompt, max_output_tokens=260).strip()
    except Exception as e:
        return f"LLM unavailable ({type(e).__name__})"


def safe_llm_weekly_coach(now: datetime, prem_s: dict, mid_s: dict, mid_next_s: dict, all_s: dict) -> str:
    try:
        out = llm_weekly_coach(now, prem_s, mid_s, mid_next_s, all_s)
        if not str(out or "").strip():
            raise ValueError("empty weekly coach")

        lines = [ln.strip() for ln in str(out).splitlines() if ln.strip()]
        required = [
            "- Weekly coaching view:",
            "- Key evidence:",
            "- Risk note:",
            "- Next action:",
        ]
        if len(lines) < 4 or any(not lines[i].startswith(required[i]) for i in range(4)):
            raise ValueError("bad weekly coach structure")

        return "\n".join(lines[:4]).strip()
    except Exception:
        return "\n".join([
            "- Weekly coaching view: Treat midday as discovery mode until next-session results stabilize.",
            "- Key evidence: Compare midday same-day and midday next-session win rates directly.",
            "- Risk note: Small weekly samples can distort signal quality conclusions.",
            "- Next action: Keep premarket unchanged and monitor midday next-session for another week.",
        ])




# -----------------------------
# Postmarket runner
# -----------------------------
def run_postmarket(now: datetime | None = None) -> None:
    now = now or datetime.now(LOCAL_TZ)

    if now.time() < POST_MARKET_START:
        print("⏳ Post-market skipped (too early).")
        if IS_LOCAL:
            send_email("🧪 Postmarket Test (too early)", "<p>Skipped because too early.</p>")
        return

    run_date = now.strftime("%Y-%m-%d")
    rf = run_dir(now, "postmarket")

    psummary_text = ""
    try:
        psummary_text, _closed = portfolio_update_and_close(now)
    except Exception as e:
        psummary_text = f"Portfolio update failed: {e}"

    # ✅ Source-of-truth loads
    prem = load_premarket_today(DAILY_LOG_CSV, run_date)
    mid = load_midday_today(DAILY_LOG_CSV, run_date)

    prem_eval = evaluate_rows(prem, run_date) if not prem.empty else pd.DataFrame()
    mid_eval = evaluate_rows(mid, run_date) if not mid.empty else pd.DataFrame()

    mid_next_eval = evaluate_midday_next_session(mid, run_date) if not mid.empty else pd.DataFrame()

    if not mid_eval.empty and not mid_next_eval.empty:
        mid_eval = mid_eval.merge(mid_next_eval, on="symbol", how="left")

        mid_eval["combined_outcome"] = mid_eval.apply(
            lambda r: _compute_combined_midday_outcome(
                bool(r.get("target_hit", False)),
                bool(r.get("stop_hit", False)),
                bool(r.get("next_target_hit", False)),
            ),
            axis=1,
        )

        mid_eval["recovered_after_stop"] = mid_eval.apply(
            lambda r: bool(r.get("stop_hit", False)) and bool(r.get("next_target_hit", False)),
            axis=1,
        )

        mid_eval = _ensure_cols(mid_eval, PERF_COLS)[PERF_COLS]


    elif not mid_eval.empty:

        mid_eval["combined_outcome"] = mid_eval["outcome"]
        mid_eval["recovered_after_stop"] = False
        mid_eval = _ensure_cols(mid_eval, PERF_COLS)[PERF_COLS]

    all_eval = (
        pd.concat([prem_eval, mid_eval], ignore_index=True)
        if (not prem_eval.empty or not mid_eval.empty)
        else pd.DataFrame()
    )

    if POSTMARKET_DEBUG:
        if prem is not None:
            prem.to_csv(Path(rf) / "debug_premarket_source.csv", index=False)
        if mid is not None:
            mid.to_csv(Path(rf) / "debug_midday_source.csv", index=False)
        if prem_eval is not None and not prem_eval.empty:
            prem_eval.to_csv(Path(rf) / "debug_premarket_eval.csv", index=False)
        if mid_eval is not None and not mid_eval.empty:
            mid_eval.to_csv(Path(rf) / "debug_midday_eval.csv", index=False)

    force_weekly_email = os.getenv("FORCE_WEEKLY_EMAIL", "0") == "1"

    if all_eval.empty:
        html = f"""
            <h2>📊 Post-Market Summary ({run_date})</h2>
            <p>No premarket picks (daily log) AND no midday recommendations (daily log) found for today.</p>
        """
        if psummary_text:
            html += f"<h3>📁 Portfolio Summary</h3><pre>{_html.escape(psummary_text)}</pre>"

        html += f"""
        <hr>
        <p><b>Debug:</b><br>
        daily_log_exists={os.path.exists(DAILY_LOG_CSV)} size={os.path.getsize(DAILY_LOG_CSV) if os.path.exists(DAILY_LOG_CSV) else 0}<br>
        prem_rows_today={len(prem)} | mid_rows_today={len(mid)}
        </p>
        """

        send_postmarket_email_once(now, f"📊 Post-Market Summary ({run_date})", html)

        if force_weekly_email or now.weekday() == 4:
            dash = build_weekly_dashboard_html(PERF_LOG_CSV, now)
            weekly_excel, rows_7d, rows_total = build_weekly_excel(PERF_LOG_CSV, now)

            dash += f"""
            <hr>
            <p><b>Debug:</b><br>
            perf_log_path: {PERF_LOG_CSV}<br>
            perf_log_exists: {os.path.exists(PERF_LOG_CSV)}<br>
            perf_log_size_bytes: {os.path.getsize(PERF_LOG_CSV) if os.path.exists(PERF_LOG_CSV) else 0}<br>
            perf_rows_total: {rows_total}<br>
            perf_rows_last_7d: {rows_7d}<br>
            weekly_excel: {weekly_excel}
            </p>
            """

            if psummary_text:
                dash += f"<h3>📁 Current Portfolio</h3><pre>{_html.escape(psummary_text)}</pre>"

            send_email(f"📅 Weekly Trading Dashboard ({run_date})", dash, attachment_path=weekly_excel)

        return

    prem_s = _summarize_eval(prem_eval)
    mid_s = _summarize_eval(mid_eval)
    all_s = _summarize_eval(all_eval)

    prem_by = _summarize_eval_by_instrument(prem_eval)
    mid_by = _summarize_eval_by_instrument(mid_eval)
    all_by = _summarize_eval_by_instrument(all_eval)

    mid_next_s = _summarize_eval(mid_eval, outcome_col="next_outcome")
    mid_next_by = _summarize_eval_by_instrument(mid_eval, outcome_col="next_outcome")
    mid_combined_s = _summarize_eval(mid_eval, outcome_col="combined_outcome")

    suggestions_obj = compute_strategy_suggestions_from_stats(run_date, prem_s, mid_s, all_s)
    save_strategy_suggestions_json(suggestions_obj, Path(rf))

    summary_html = f"""
        <h2>📊 Post-Market Summary ({run_date})</h2>
        <p><b>Win definition:</b> target hit anytime after recommendation (intraday high ≥ target). Stop is a loss only if it occurs before target.</p>

        <h3>✅ Premarket</h3>
        <p>
          evaluated: {prem_s["evaluated"]}<br>
          wins (target hit): {prem_s["wins"]}<br>
          losses (stop hit): {prem_s["losses"]}<br>
          not hit: {prem_s["not_hit"]}<br>
          win rate: {prem_s["win_rate"]:.2f}%
        </p>
        <p><b>Premarket buckets:</b>
          Stock win_rate={prem_by["stock"]["win_rate"]:.2f}% (n={prem_by["stock"]["evaluated"]}) |
          Options win_rate={prem_by["options"]["win_rate"]:.2f}% (n={prem_by["options"]["evaluated"]})
        </p>

         <h3>⚡ Midday</h3>
        <p>
          <b>Same-day:</b><br>
          evaluated: {mid_s["evaluated"]}<br>
          wins (target hit): {mid_s["wins"]}<br>
          late wins: {mid_s.get("late_wins", 0)}<br>
          losses (stop hit): {mid_s["losses"]}<br>
          not hit: {mid_s["not_hit"]}<br>
          win rate: {mid_s["win_rate"]:.2f}%
        </p>
        <p><b>Midday same-day buckets:</b>
          Stock win_rate={mid_by["stock"]["win_rate"]:.2f}% (n={mid_by["stock"]["evaluated"]}) |
          Options win_rate={mid_by["options"]["win_rate"]:.2f}% (n={mid_by["options"]["evaluated"]})
        </p>

        <p>
          <b>Next-session:</b><br>
          evaluated: {mid_next_s["evaluated"]}<br>
          wins (target hit): {mid_next_s["wins"]}<br>
          late wins: {mid_next_s.get("late_wins", 0)}<br>
          losses (stop hit): {mid_next_s["losses"]}<br>
          not hit: {mid_next_s["not_hit"]}<br>
          win rate: {mid_next_s["win_rate"]:.2f}%
        </p>
        <p><b>Midday next-session buckets:</b>
          Stock win_rate={mid_next_by["stock"]["win_rate"]:.2f}% (n={mid_next_by["stock"]["evaluated"]}) |
          Options win_rate={mid_next_by["options"]["win_rate"]:.2f}% (n={mid_next_by["options"]["evaluated"]})
        </p>

        <p>
          <b>Combined interpretation:</b><br>
          evaluated: {mid_combined_s["evaluated"]}<br>
          wins (same-day target): {mid_combined_s["wins"]}<br>
          late wins (stop first, target later): {mid_combined_s.get("late_wins", 0)}<br>
          losses: {mid_combined_s["losses"]}<br>
          not hit: {mid_combined_s["not_hit"]}<br>
          effective win rate: {mid_combined_s["win_rate"]:.2f}%
        </p>
    """

    summary_html += render_strategy_suggestions_html(suggestions_obj)

    if psummary_text:
        summary_html += f"<h3>📁 Portfolio Summary</h3><pre>{_html.escape(psummary_text)}</pre>"

    narrative = ""
    LLM_ENABLED = os.getenv("LLM_ENABLED", "1") == "1"
    if LLM_ENABLED:
        narrative = safe_llm_daily_narrative(run_date, prem_s, mid_s, all_s)

    if narrative:
        summary_html += f"<h3>🧠 LLM Narrative</h3><pre>{_html.escape(narrative)}</pre>"

    coach = ""
    if LLM_ENABLED:
        coach = safe_postmarket_coach(run_date, prem_s, mid_s, all_s)

    if coach:
        summary_html += f"<h3>🧠 LLM Coach</h3><pre>{_html.escape(coach)}</pre>"

    post_excel = out_path(
        f"post_market_report_{now.strftime('%Y%m%d')}_{make_run_id(now)}.xlsx",
        now=now,
        mode="postmarket",
        kind="runs",
    )

    summary_df = pd.DataFrame([
        {"bucket": "premarket_same_day", "evaluated": prem_s["evaluated"], "wins": prem_s["wins"],
         "losses": prem_s["losses"], "not_hit": prem_s["not_hit"], "win_rate_pct": round(prem_s["win_rate"], 2)},
        {"bucket": "midday_same_day", "evaluated": mid_s["evaluated"], "wins": mid_s["wins"], "losses": mid_s["losses"],
         "not_hit": mid_s["not_hit"], "win_rate_pct": round(mid_s["win_rate"], 2)},
        {"bucket": "midday_next_session", "evaluated": mid_next_s["evaluated"], "wins": mid_next_s["wins"],
         "losses": mid_next_s["losses"], "not_hit": mid_next_s["not_hit"],
         "win_rate_pct": round(mid_next_s["win_rate"], 2)},
        {"bucket": "combined_same_day", "evaluated": all_s["evaluated"], "wins": all_s["wins"],
         "losses": all_s["losses"], "not_hit": all_s["not_hit"], "win_rate_pct": round(all_s["win_rate"], 2)},
        {"bucket": "combined_stock_same_day", "evaluated": all_by["stock"]["evaluated"],
         "wins": all_by["stock"]["wins"], "losses": all_by["stock"]["losses"], "not_hit": all_by["stock"]["not_hit"],
         "win_rate_pct": round(all_by["stock"]["win_rate"], 2)},
        {"bucket": "combined_options_same_day", "evaluated": all_by["options"]["evaluated"],
         "wins": all_by["options"]["wins"], "losses": all_by["options"]["losses"],
         "not_hit": all_by["options"]["not_hit"], "win_rate_pct": round(all_by["options"]["win_rate"], 2)},
        {"bucket": "midday_combined", "evaluated": mid_combined_s["evaluated"], "wins": mid_combined_s["wins"],
         "losses": mid_combined_s["losses"], "not_hit": mid_combined_s["not_hit"],
         "win_rate_pct": round(mid_combined_s["win_rate"], 2)},
    ])

    with pd.ExcelWriter(post_excel, engine="openpyxl") as writer:
        _excel_safe_df(summary_df).to_excel(writer, sheet_name="SUMMARY", index=False)

        if not prem_eval.empty:
            _excel_safe_df(prem_eval).to_excel(writer, sheet_name="PREMARKET_EVAL", index=False)

        if not mid_eval.empty:
            # Same-day midday evaluation
            _excel_safe_df(mid_eval).to_excel(writer, sheet_name="MIDDAY_EVAL", index=False)

            # Next-session midday view
            mid_next_view = mid_eval[[
                "run_date", "symbol", "decision", "score", "confidence",
                "entry_price", "target_price", "stop_loss",
                "reco_ts", "next_session_end",
                "next_close_price", "next_actual_change_pct_close",
                "next_target_hit", "next_stop_hit",
                "next_first_hit", "next_first_hit_time",
                "next_hit_latency_minutes", "next_target_overshoot_pct",
                "next_best_exit_price_after_target", "next_best_exit_time_after_target",
                "next_best_exit_from_entry_pct", "next_best_exit_from_target_pct",
                "next_best_exit_latency_minutes", "next_outcome",
                "combined_outcome", "recovered_after_stop",
            ]].copy()

            _excel_safe_df(mid_next_view).to_excel(writer, sheet_name="MIDDAY_NEXT_SESSION", index=False)

        _excel_safe_df(all_eval).to_excel(writer, sheet_name="ALL_EVAL", index=False)

    wb = load_workbook(post_excel)
    for s in wb.sheetnames:
        style_excel_sheet(wb[s])
    wb.save(post_excel)

    send_postmarket_email_once(
        now,
        f"📊 Post-Market Report (Target-Hit + Analytics) ({run_date})",
        summary_html,
        attachment_path=post_excel,
    )

    append_perf_log(all_eval, now=now)

    # Backfill prior day's midday next-session now that today's close is known
    backfill_prior_midday_next_session(PERF_LOG_CSV, DAILY_LOG_CSV, run_date)

    # Friday = weekly dashboard
    force_weekly_email = os.getenv("FORCE_WEEKLY_EMAIL", "0") == "1"
    if force_weekly_email or now.weekday() == 4:
        dash = build_weekly_dashboard_html(PERF_LOG_CSV, now)
        weekly_excel, rows_7d, rows_total = build_weekly_excel(PERF_LOG_CSV, now)

        dash += f"""
        <hr>
        <p><b>Debug:</b><br>
        perf_log_path: {PERF_LOG_CSV}<br>
        perf_log_exists: {os.path.exists(PERF_LOG_CSV)}<br>
        perf_log_size_bytes: {os.path.getsize(PERF_LOG_CSV) if os.path.exists(PERF_LOG_CSV) else 0}<br>
        perf_rows_total: {rows_total}<br>
        perf_rows_last_7d: {rows_7d}<br>
        weekly_excel: {weekly_excel}
        </p>
        """

        if psummary_text:
            dash += f"<h3>📁 Current Portfolio</h3><pre>{_html.escape(psummary_text)}</pre>"

        send_email(f"📅 Weekly Trading Dashboard ({run_date})", dash, attachment_path=weekly_excel)

    print(
        f"✅ Postmarket complete | "
        f"premarket_eval={prem_s['evaluated']} mid_eval={mid_s['evaluated']} "
        f"combined_eval={all_s['evaluated']} win_rate={all_s['win_rate']:.2f}% | "
        f"excel={post_excel}"
    )


if __name__ == "__main__":
    run_postmarket()