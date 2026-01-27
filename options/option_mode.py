# options/option_mode.py
from __future__ import annotations

import json
from pathlib import Path
from datetime import datetime, date
from zoneinfo import ZoneInfo
from typing import Optional, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from config import (
    APP_ENV,
    IS_LOCAL,
    SENDER_EMAIL,
    RECEIVER_EMAIL,
    LOCAL_RECEIVER_EMAIL,
    EMAIL_SUBJECT_PREFIX_LOCAL,
    EMAIL_SUBJECT_PREFIX_PROD,
)

from email_sender import send_email as _send_email
from options.option_config import load_occ_list
from options.option_runner import run_options
from options.option_universe import build_occ_list_from_underlyings, UniverseConfig

LOCAL_TZ = ZoneInfo("America/Chicago")

EMAIL_SUBJECT_PREFIX = EMAIL_SUBJECT_PREFIX_LOCAL if IS_LOCAL else EMAIL_SUBJECT_PREFIX_PROD
EFFECTIVE_RECEIVER_EMAIL = (LOCAL_RECEIVER_EMAIL or RECEIVER_EMAIL) if IS_LOCAL else RECEIVER_EMAIL


# -----------------------------
# Output helpers
# -----------------------------
def env_base_dir() -> Path:
    base = Path(__file__).resolve().parents[1] / "outputs" / APP_ENV
    base.mkdir(parents=True, exist_ok=True)
    return base


def run_dir(now: datetime, mode: str) -> Path:
    day = now.strftime("%Y%m%d")
    p = env_base_dir() / "runs" / day / mode
    p.mkdir(parents=True, exist_ok=True)
    return p


def make_run_id(now: datetime) -> str:
    return now.strftime("%Y%m%d_%H%M%S")


# -----------------------------
# Excel styling
# -----------------------------
def _normalize_color(color: str) -> str:
    if not color:
        color = "#FFFFFF"
    color = color.lstrip("#")
    if len(color) == 6:
        color = "FF" + color
    return color.upper()


def _fill(hex_color: str) -> PatternFill:
    c = _normalize_color(hex_color)
    return PatternFill(start_color=c, end_color=c, fill_type="solid")


def style_excel_sheet(sheet) -> None:
    """
    Styles header + auto width + highlights key rows:
      - TARGET rows: green
      - ready_to_trade=YES: light green
      - spread_flag=WIDE: light red
      - decision=AVOID: gray
    """
    if sheet.max_row < 1 or sheet.max_column < 1:
        return

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color=_normalize_color("#2F5597"),
        end_color=_normalize_color("#2F5597"),
        fill_type="solid",
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Header
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    sheet.freeze_panes = "A2"

    # Auto-width
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = min(max_length + 2, 60)

    # Row highlighting based on columns if present
    # Build column name -> index map
    header_vals = [c.value for c in sheet[1]]
    col_idx = {str(v).strip(): i + 1 for i, v in enumerate(header_vals) if v}

    def _cell_val(row_idx: int, col_name: str) -> str:
        i = col_idx.get(col_name)
        if not i:
            return ""
        v = sheet.cell(row=row_idx, column=i).value
        return "" if v is None else str(v)

    fill_target = _fill("#C6EFCE")   # green
    fill_ready = _fill("#E2F0D9")    # light green
    fill_wide = _fill("#F8CBAD")     # light red
    fill_avoid = _fill("#D9D9D9")    # gray

    has_target = "target_flag" in col_idx
    has_ready = "ready_to_trade" in col_idx
    has_spread = "spread_flag" in col_idx
    has_decision = "decision" in col_idx

    for r in range(2, sheet.max_row + 1):
        target = _cell_val(r, "target_flag").upper() if has_target else ""
        ready = _cell_val(r, "ready_to_trade").upper() if has_ready else ""
        spread = _cell_val(r, "spread_flag").upper() if has_spread else ""
        decision = _cell_val(r, "decision").upper() if has_decision else ""

        row_fill = None
        if decision == "AVOID":
            row_fill = fill_avoid
        elif target == "TARGET":
            row_fill = fill_target
        elif ready == "YES":
            row_fill = fill_ready
        elif spread == "WIDE":
            row_fill = fill_wide

        if row_fill is not None:
            for c in range(1, sheet.max_column + 1):
                sheet.cell(row=r, column=c).fill = row_fill


# -----------------------------
# Email wrapper (avoid name collision)
# -----------------------------
def send_options_email(subject: str, html_body: str, attachment_path: Optional[str] = None) -> bool:
    final_subject = f"{EMAIL_SUBJECT_PREFIX} {subject}"
    return _send_email(
        subject=final_subject,
        html_body=html_body,
        to_email=EFFECTIVE_RECEIVER_EMAIL,
        from_email=SENDER_EMAIL,
        attachment_path=attachment_path,
    )


# -----------------------------
# Helpers
# -----------------------------
def _df_to_html_table(df: pd.DataFrame, max_rows: int = 25) -> str:
    if df is None or df.empty:
        return "<p>(empty)</p>"

    view_cols = [
        "target_flag",
        "ready_to_trade",
        "entry_order",
        "entry_price",
        "trade_summary",
        "occ",
        "underlying",
        "expiry",
        "type",
        "strike",
        "decision",
        "score",
        "confidence",
        "mid",
        "bid",
        "ask",
        "last",
        "dte",
        "spread_pct",
        "spread_flag",
        "rec_hold_days",
        "target_profit_pct",
        "stop_loss_pct",
        "profit_target_price",
        "stop_price",
        "volume",
        "open_interest",
        "iv",
        "fallback_used",
        "fallback_reason",
        "reasons",
    ]
    cols = [c for c in view_cols if c in df.columns]
    d2 = df[cols].head(max_rows).copy()
    d2 = d2.fillna("")
    return d2.to_html(index=False, escape=True)


def _summary_counts(df: pd.DataFrame) -> Dict[str, int]:
    if df is None or df.empty or "decision" not in df.columns:
        return {"BUY": 0, "WATCH": 0, "AVOID": 0, "TOTAL": 0}
    s = df["decision"].astype(str).str.upper().value_counts()
    return {
        "BUY": int(s.get("BUY", 0)),
        "WATCH": int(s.get("WATCH", 0)),
        "AVOID": int(s.get("AVOID", 0)),
        "TOTAL": int(len(df)),
    }


def _clean_symbols(raw: List[str]) -> List[str]:
    out: List[str] = []
    for s in raw or []:
        if s is None:
            continue
        t = str(s).strip().upper().replace(" ", "")
        if t:
            out.append(t)
    return out


def _read_symbols_from_log_csv(path: Path, run_date: str) -> List[str]:
    """
    Reads symbols from your log CSVs in outputs/{env}/logs/.
    Filters by run_date if a date column exists (run_date/date/day/asof_date).
    """
    if not path.exists():
        return []
    try:
        df = pd.read_csv(path)
        if df is None or df.empty:
            return []
    except Exception:
        return []

    date_cols = [c for c in ["run_date", "date", "day", "asof_date"] if c in df.columns]
    if date_cols:
        c = date_cols[0]
        df[c] = df[c].astype(str)
        df = df[df[c].str.contains(run_date, na=False)]

    sym_col = None
    for c in ["symbol", "ticker", "underlying"]:
        if c in df.columns:
            sym_col = c
            break
    if not sym_col:
        return []

    return _clean_symbols(df[sym_col].astype(str).tolist())


def _parse_expiry_to_date(x) -> Optional[date]:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    s = str(x).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    try:
        return pd.to_datetime(s, errors="coerce").date()  # type: ignore[attr-defined]
    except Exception:
        return None


def _to_float(x) -> Optional[float]:
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return None
        v = float(x)
        if pd.isna(v):
            return None
        return v
    except Exception:
        return None


def _to_int(x) -> Optional[int]:
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return None
        return int(float(x))
    except Exception:
        return None


def _spread_threshold_for_mid(midv: Optional[float]) -> float:
    """
    Consistent spread tolerance:
      - cheaper options can tolerate a bit wider spread
    """
    if midv is None:
        return 0.12
    return 0.15 if midv < 0.30 else 0.12


def _add_trade_plan_columns(df: pd.DataFrame, now: datetime) -> pd.DataFrame:
    """
    Adds decision-ready trade plan fields:
    - dte, mid, spread_pct, spread_flag
    - rec_hold_days
    - target_profit_pct, stop_loss_pct
    - profit_target_price, stop_price
    - entry_price, entry_order
    - ready_to_trade
    - trade_summary
    - target_flag (TARGET per underlying)
    """
    if df is None or df.empty:
        return df

    out = df.copy()

    # --- Mid price ---
    if "mid" not in out.columns and ("bid" in out.columns and "ask" in out.columns):
        out["mid"] = (pd.to_numeric(out["bid"], errors="coerce") + pd.to_numeric(out["ask"], errors="coerce")) / 2.0

    # Ensure numeric
    for c in ["bid", "ask", "mid", "score", "volume", "open_interest"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")

    # --- DTE ---
    if "expiry" in out.columns:
        exp_dates = out["expiry"].apply(_parse_expiry_to_date)
        today = now.date()
        out["dte"] = exp_dates.apply(lambda d: (d - today).days if d else None)
    else:
        out["dte"] = None

    # --- Spread % ---
    if "bid" in out.columns and "ask" in out.columns:
        spread = out["ask"] - out["bid"]
        mid = out["mid"] if "mid" in out.columns else None
        if mid is not None:
            out["spread_pct"] = spread / mid.replace({0: pd.NA})
        else:
            out["spread_pct"] = None
    else:
        out["spread_pct"] = None

    # Spread flag
    def _spread_flag(row) -> str:
        sp = _to_float(row.get("spread_pct"))
        midv = _to_float(row.get("mid"))
        if sp is None or midv is None:
            return ""
        threshold = _spread_threshold_for_mid(midv)
        return "OK" if sp <= threshold else "WIDE"

    out["spread_flag"] = out.apply(_spread_flag, axis=1)

    # --- Target/Stop by confidence ---
    def _targets_from_conf(conf_val) -> Tuple[float, float]:
        ci = _to_int(conf_val)
        if ci is not None:
            if ci >= 7:
                return 0.25, 0.15
            if ci >= 6:
                return 0.20, 0.12
            if ci >= 5:
                return 0.15, 0.12
            return 0.10, 0.10

        c = (str(conf_val or "")).strip().lower()
        if c in ("high", "h", "strong"):
            return 0.25, 0.15
        if c in ("medium", "med", "m"):
            return 0.15, 0.12
        if c in ("low", "l", "weak"):
            return 0.10, 0.10
        return 0.20, 0.12

    def _hold_days_from_dte(dte_val: Optional[float]) -> Optional[int]:
        if dte_val is None or pd.isna(dte_val):
            return None
        dte_i = int(dte_val)
        if 7 <= dte_i <= 14:
            return 2
        if 15 <= dte_i <= 30:
            return 5
        if 31 <= dte_i <= 60:
            return 10
        if dte_i > 60:
            return 20
        if dte_i >= 1:
            return 1
        return None

    out["target_profit_pct"] = None
    out["stop_loss_pct"] = None
    out["rec_hold_days"] = out["dte"].apply(_hold_days_from_dte)

    def _adjust_hold(row) -> Optional[int]:
        hd = row.get("rec_hold_days")
        if hd is None or (isinstance(hd, float) and pd.isna(hd)):
            return None
        hd_i = int(hd)

        conf_i = _to_int(row.get("confidence"))
        wide = str(row.get("spread_flag") or "").upper() == "WIDE"
        low_conf = (conf_i is not None and conf_i <= 5) or (
            str(row.get("confidence") or "").strip().lower() in ("low", "l", "weak")
        )
        if wide or low_conf:
            return max(1, int(round(hd_i * 0.5)))
        return hd_i

    out["rec_hold_days"] = out.apply(_adjust_hold, axis=1)

    def _fill_trade_plan(row):
        tp, sl = _targets_from_conf(row.get("confidence"))
        midv = _to_float(row.get("mid"))
        if midv is None:
            return pd.Series(
                {
                    "target_profit_pct": tp,
                    "stop_loss_pct": sl,
                    "profit_target_price": None,
                    "stop_price": None,
                }
            )
        return pd.Series(
            {
                "target_profit_pct": tp,
                "stop_loss_pct": sl,
                "profit_target_price": midv * (1.0 + tp),
                "stop_price": midv * (1.0 - sl),
            }
        )

    plan = out.apply(_fill_trade_plan, axis=1)
    out["target_profit_pct"] = plan["target_profit_pct"]
    out["stop_loss_pct"] = plan["stop_loss_pct"]
    out["profit_target_price"] = plan["profit_target_price"]
    out["stop_price"] = plan["stop_price"]

    # -----------------------------
    # Entry price + order type
    # -----------------------------
    def _entry_plan(row) -> pd.Series:
        bid = _to_float(row.get("bid"))
        ask = _to_float(row.get("ask"))
        midv = _to_float(row.get("mid"))
        sp = _to_float(row.get("spread_pct"))

        if bid is None or ask is None or midv is None:
            return pd.Series({"entry_price": None, "entry_order": ""})

        threshold = _spread_threshold_for_mid(midv)

        if sp is not None and sp > threshold:
            entry = bid
            order = "LIMIT @ BID"
        else:
            entry = midv
            order = "LIMIT @ MID"

        return pd.Series({"entry_price": entry, "entry_order": order})

    ep = out.apply(_entry_plan, axis=1)
    out["entry_price"] = ep["entry_price"]
    out["entry_order"] = ep["entry_order"]

    # -----------------------------
    # Ready-to-trade gating
    # -----------------------------
    def _ready(row) -> str:
        decision = str(row.get("decision") or "").upper()
        score = _to_float(row.get("score"))
        conf_i = _to_int(row.get("confidence"))
        dte_i = _to_int(row.get("dte"))
        vol = _to_int(row.get("volume"))
        oi = _to_int(row.get("open_interest"))
        spread_ok = str(row.get("spread_flag") or "").upper() == "OK"

        if decision == "AVOID":
            return "NO"
        if score is None or score < 65:
            return "NO"
        if conf_i is None or conf_i < 6:
            return "NO"
        if dte_i is None or not (14 <= dte_i <= 45):
            return "NO"
        if not spread_ok:
            return "NO"
        if vol is None or vol < 50:
            return "NO"
        if oi is None or oi < 100:
            return "NO"
        return "YES"

    out["ready_to_trade"] = out.apply(_ready, axis=1)

    # -----------------------------
    # Trade summary text
    # -----------------------------
    def _trade_summary(row) -> str:
        occ = str(row.get("occ") or "").strip()
        u = str(row.get("underlying") or "").strip()
        exp = str(row.get("expiry") or "").strip()
        typ = str(row.get("type") or "").strip().upper()
        strike = row.get("strike")
        entry_order = str(row.get("entry_order") or "").strip()
        entry_price = _to_float(row.get("entry_price"))
        tgt = _to_float(row.get("profit_target_price"))
        stp = _to_float(row.get("stop_price"))
        dte_i = _to_int(row.get("dte"))

        side = "BUY TO OPEN"
        contract = f"{u} {exp} {strike} {typ}" if u and exp and strike is not None and typ else occ

        parts = [f"{side} {contract}"]
        if entry_order and entry_price is not None:
            parts.append(f"{entry_order} {entry_price:.2f}")
        elif entry_price is not None:
            parts.append(f"LIMIT {entry_price:.2f}")

        if tgt is not None:
            parts.append(f"TARGET {tgt:.2f}")
        if stp is not None:
            parts.append(f"STOP {stp:.2f}")
        if dte_i is not None:
            parts.append(f"DTE {dte_i}")

        return " | ".join(parts)

    out["trade_summary"] = out.apply(_trade_summary, axis=1)

    # -----------------------------
    # Mark TARGET contract per underlying
    # -----------------------------
    out["target_flag"] = ""
    if "underlying" in out.columns:
        d = out.copy()
        d["_is_candidate"] = d["decision"].astype(str).str.upper().isin(["BUY", "WATCH"])
        d["_ok_spread"] = d["spread_flag"].astype(str).str.upper().eq("OK")
        d["_score"] = pd.to_numeric(d["score"], errors="coerce")
        d["_ready"] = d["ready_to_trade"].astype(str).str.upper().eq("YES")

        candidates = d[d["_is_candidate"] & d["_ok_spread"] & d["_score"].notna()]
        if not candidates.empty:
            candidates = candidates.sort_values(
                ["underlying", "_ready", "_score"],
                ascending=[True, False, False],
            )
            idx = candidates.groupby("underlying", as_index=False).head(1).index
            out.loc[idx, "target_flag"] = "TARGET"

        for c in ["_is_candidate", "_ok_spread", "_score", "_ready"]:
            if c in out.columns:
                out.drop(columns=[c], inplace=True)

    # Rounding
    for c in ["spread_pct", "target_profit_pct", "stop_loss_pct"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").round(4)

    for c in ["entry_price", "profit_target_price", "stop_price"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").round(4)

    return out


# -----------------------------
# Options mode
# -----------------------------
def run_options_mode(now: Optional[datetime] = None, dry_run: bool = False) -> pd.DataFrame:
    now = now or datetime.now(LOCAL_TZ)
    run_date = now.strftime("%Y-%m-%d")

    print(f"🧾 OPTIONS MODE @ {now.strftime('%Y-%m-%d %H:%M:%S')} | dry_run={dry_run}")

    occ_list: List[str] = []

    try:
        logs_dir = env_base_dir() / "logs"
        daily_log = logs_dir / "daily_stock_log.csv"
        reco_log = logs_dir / "recommendations_log.csv"

        underlyings: List[str] = []
        underlyings += _read_symbols_from_log_csv(daily_log, run_date)
        underlyings += _read_symbols_from_log_csv(reco_log, run_date)
        underlyings = sorted(set([u for u in underlyings if u]))

        if underlyings:
            cfg = UniverseConfig(
                dte_min=20,
                dte_max=120,
                max_underlyings=15,
                max_contracts_per_underlying=5,
            )
            occ_list = build_occ_list_from_underlyings(underlyings, cp="C", cfg=cfg)
            print(f"✅ Generated {len(occ_list)} options from {len(underlyings)} underlyings.")
        else:
            print("⚠️ No underlyings found in today’s logs for universe generation.")

    except Exception as e:
        print("⚠️ Dynamic universe failed (will fallback to .env OPTIONS_OCC_LIST). Reason:", e)

    if not occ_list:
        occ_list = load_occ_list()
        if occ_list:
            print(f"↩️ Using OPTIONS_OCC_LIST from .env ({len(occ_list)} contracts).")

    if not occ_list:
        print("⚠️ No options to run (no daily symbols + no OPTIONS_OCC_LIST).")
        return pd.DataFrame()

    df = run_options(occ_list)

    if df is None or df.empty:
        print("⚠️ Options runner returned empty DataFrame.")
        return pd.DataFrame()

    df_out = _add_trade_plan_columns(df, now=now)

    print(df_out.to_string(index=False))

    out_xlsx = run_dir(now, "options") / f"options_report_{now.strftime('%Y%m%d')}_{make_run_id(now)}.xlsx"

    if "metrics" in df_out.columns:
        df_out["metrics"] = df_out["metrics"].apply(lambda x: x if isinstance(x, str) else json.dumps(x, default=str))

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="OPTIONS", index=False)

        if "decision" in df_out.columns:
            for bucket in ["BUY", "WATCH", "AVOID"]:
                dsub = df_out[df_out["decision"].astype(str).str.upper() == bucket].copy()
                if not dsub.empty:
                    dsub.to_excel(writer, sheet_name=bucket, index=False)

    wb = load_workbook(out_xlsx)
    for s in wb.sheetnames:
        style_excel_sheet(wb[s])
    wb.save(out_xlsx)

    print(f"✅ Options Excel written: {out_xlsx}")

    counts = _summary_counts(df_out)
    html = f"""
        <h2>🧾 Options Report ({now.strftime('%Y-%m-%d')})</h2>
        <p>
          <b>Total:</b> {counts["TOTAL"]} &nbsp; | &nbsp;
          <b>BUY:</b> {counts["BUY"]} &nbsp; | &nbsp;
          <b>WATCH:</b> {counts["WATCH"]} &nbsp; | &nbsp;
          <b>AVOID:</b> {counts["AVOID"]}
        </p>

        <p>
          <b>How to read this:</b><br/>
          <b>target_flag=TARGET</b> = best contract per underlying (ready first, then score).<br/>
          <b>ready_to_trade=YES</b> = passes liquidity + spread + DTE rules.<br/>
          <b>entry_price</b> = suggested <i>limit</i> buy price (MID if tight spread, BID if wide).<br/>
          <b>trade_summary</b> = exact “call to make” text (BUY TO OPEN + limit/target/stop).
        </p>

        <h3>Top Results</h3>
        {_df_to_html_table(df_out, max_rows=25)}
        <p><i>Attachment contains full detail + bucket tabs.</i></p>
    """

    if dry_run:
        print("🧪 dry-run enabled: not sending email.")
        return df_out

    send_options_email(
        subject=f"🧾 Options Report ({now.strftime('%Y-%m-%d')})",
        html_body=html,
        attachment_path=str(out_xlsx),
    )

    return df_out


if __name__ == "__main__":
    run_options_mode()