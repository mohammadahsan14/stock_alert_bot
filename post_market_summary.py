# post_market_summary.py
import pandas as pd
from config import SENDER_EMAIL, APP_PASSWORD, RECEIVER_EMAIL
import smtplib
from email.message import EmailMessage
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import yfinance as yf

# -----------------------------
# Normalize hex color for Excel
# -----------------------------
def normalize_color(color):
    if not color:
        color = "#FFFFFF"
    color = color.lstrip("#")
    if len(color) == 6:
        color = "FF" + color  # Add alpha channel
    return color.upper()

# -----------------------------
# Style Excel Sheet
# -----------------------------
def style_excel_sheet(ws):
    for row in ws.iter_rows(min_row=2):
        outcome_cell = row[ws.max_column - 1]
        if outcome_cell.value == "✅ Correct":
            fill_color = "#C6EFCE"
            font_color = "#006100"
        elif outcome_cell.value == "❌ Incorrect":
            fill_color = "#FFC7CE"
            font_color = "#9C0006"
        else:
            fill_color = "#FFFFFF"
            font_color = "#000000"

        for cell in row:
            try:
                cell.fill = PatternFill(start_color=normalize_color(fill_color),
                                        end_color=normalize_color(fill_color),
                                        fill_type="solid")
                cell.font = Font(color=font_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            except Exception:
                pass

# -----------------------------
# Evaluate predictions
# -----------------------------
def evaluate_performance(df):
    results = []
    confidences = []
    for _, row in df.iterrows():
        predicted = row.get('decision', 'Not Advisable')
        actual_change = row.get('actual_change_pct', 0)

        # Determine correctness
        if predicted in ["Strong Buy", "Moderate"] and actual_change > 0:
            outcome = "✅ Correct"
        elif predicted == "Not Advisable" and actual_change <= 0:
            outcome = "✅ Correct"
        else:
            outcome = "❌ Incorrect"
        results.append(outcome)

        # Confidence 1–10 (scaled)
        conf = min(max(int(abs(actual_change) / 2 + (1 if outcome == "✅ Correct" else 0)), 1), 10)
        confidences.append(conf)

    df['outcome'] = results
    df['confidence'] = confidences
    return df

# -----------------------------
# Build summary text
# -----------------------------
def build_summary(df):
    total = len(df)
    correct = len(df[df['outcome'] == "✅ Correct"])
    incorrect = len(df[df['outcome'] == "❌ Incorrect"])
    success_rate = (correct / total * 100) if total else 0
    summary = f"""
📊 Post-Market Summary ({datetime.today().strftime('%Y-%m-%d')})
Total Stocks Evaluated: {total}
Predictions Correct: {correct}
Predictions Incorrect: {incorrect}
Success Rate: {success_rate:.2f}%
"""
    return summary

# -----------------------------
# Send Email
# -----------------------------
def send_email(subject, body, attachment_path=None):
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = SENDER_EMAIL
    msg["To"] = RECEIVER_EMAIL
    msg.set_content(body)

    if attachment_path:
        try:
            with open(attachment_path, "rb") as f:
                msg.add_attachment(f.read(), maintype="application",
                                   subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   filename=attachment_path)
        except Exception:
            pass

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(SENDER_EMAIL, APP_PASSWORD)
            smtp.send_message(msg)
            print("✅ Post-market email sent successfully")
    except Exception as e:
        print("❌ Email failed:", e)

# -----------------------------
# MAIN
# -----------------------------
def main():
    log_file = "daily_stock_log.csv"  # Pre-market CSV

    # Load pre-market CSV
    try:
        df = pd.read_csv(log_file)
        if df.empty:
            print("❌ Log file is empty. Exiting.")
            return
    except FileNotFoundError:
        print("❌ Log file not found. Exiting.")
        return

    # Fetch actual closing prices safely
    close_prices = []
    for symbol in df.get('symbol', []):
        try:
            data = yf.Ticker(symbol).history(period="1d")
            close = data['Close'].iloc[-1] if not data.empty else df.loc[df['symbol']==symbol, 'current'].values[0]
        except Exception:
            close = df.loc[df['symbol']==symbol, 'current'].values[0]
        close_prices.append(close)
    df['close_price'] = close_prices
    df['actual_change_pct'] = (df['close_price'] - df['current']) / df['current'] * 100

    # Evaluate performance and compute confidence
    df = evaluate_performance(df)

    # Save detailed Excel
    excel_file = f"post_market_{datetime.today().strftime('%Y%m%d')}.xlsx"
    df.to_excel(excel_file, index=False)

    # Apply Excel styling
    try:
        wb = load_workbook(excel_file)
        for sheet_name in wb.sheetnames:
            style_excel_sheet(wb[sheet_name])
        wb.save(excel_file)
        print(f"✅ Detailed outcomes saved: {excel_file}")
    except Exception as e:
        print(f"❌ Failed to style Excel: {e}")

    # Build and send summary
    summary_text = build_summary(df)
    email_body = f"Hello,\n\nToday’s post-market summary:\n\n{summary_text}\n\nDetailed results are attached."
    send_email("📊 Post-Market Stock Summary", email_body, attachment_path=excel_file)

    print(summary_text)

if __name__ == "__main__":
    main()