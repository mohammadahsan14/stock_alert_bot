# ============================================================
# 🔒 LOCKED FILE — DO NOT EDIT TRADING LOGIC WITHOUT VERSION BUMP
# File: premarket_runner.py
# Version: 2026-03-04_v3
# Notes:
# - CONF_GATE is locked to MIN_CONFIDENCE_TO_TRADE (no env override)
# - FINAL_VIEW uses GO picks else WATCH high-vote fallback
# - Plan card is deterministic (LLM cannot overwrite)
# - Risk-off regime suppresses GO trades
# ============================================================

# premarket_runner.py (FINAL LOCK VERSION + FINAL_VIEW + HIGH-VOTE FALLBACK + ENRICHED RAW + DETERMINISTIC PLAN_CARD + LLM OPTIONAL)
from __future__ import annotations

from llm.explain import safe_explain_pick, safe_explain_insights
from context_insight import build_key_insight

import os
import re
import html as _html
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo
from typing import List, Dict, Any, Optional

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from config import (
    APP_ENV,
    IS_LOCAL,
    SENDER_EMAIL,
    RECEIVER_EMAIL,
    LOCAL_RECEIVER_EMAIL,
    EMAIL_SUBJECT_PREFIX_LOCAL,
    EMAIL_SUBJECT_PREFIX_PROD,
    TOP_N,
    TRADE_MAX_PICKS,
    SCORE_COLORS,
    SCORE_HIGH,
    SCORE_MEDIUM,
    MIN_CONFIDENCE_TO_TRADE,
    MAX_PRICE,
    ELITE_SCORE_OVERRIDE,
    ELITE_CONF_OVERRIDE,
)

from email_sender import send_email as _send_email
from top_movers import fetch_sp500_tickers, calculate_top_movers
from scoring_engine import get_predictive_score_with_reasons
from forecast_engine import forecast_price_levels
from news_fetcher import fetch_news_links
from price_category import get_price_category
from price_action import compute_price_action_summary

from performance_tracker import (
    PortfolioConfig,
    load_open_portfolio,
    save_open_portfolio,
    add_new_positions_from_picks,
    append_open_actions,
)

import warnings
warnings.filterwarnings(
    "ignore",
    category=FutureWarning,
    message="The behavior of DataFrame concatenation with empty or all-NA entries is deprecated"
)

LOCAL_TZ = ZoneInfo("America/Chicago")

POS_WORDS = {"beat", "strong", "growth", "surge", "upgrade", "raises", "record", "profit", "wins", "bull"}
NEG_WORDS = {"miss", "drop", "loss", "cuts", "downgrade", "falls", "weak", "lawsuit", "plunge", "bear"}

# LLM flags
PREMARKET_LLM_ENABLED = os.getenv("LLM_ENABLED", "1") == "1"
PREMARKET_LLM_TOP_N = int(os.getenv("PREMARKET_LLM_TOP_N", "8"))

# FINAL view behavior
FINAL_VIEW_TOP_N = int(os.getenv("FINAL_VIEW_TOP_N", "3"))

# 🔒 LOCKED gate: single source of truth (no env override)
CONF_GATE = int(MIN_CONFIDENCE_TO_TRADE)

##
MIN_PRICE = float(os.getenv("MIN_PRICE", "5"))
MIN_AVG_DOLLAR_VOL = float(os.getenv("MIN_AVG_DOLLAR_VOL", "20000000"))  # $20M
##


# -----------------------------
# Core numeric helper (MUST be global)
# -----------------------------
def _safe_float(x) -> Optional[float]:
    try:
        if x is None:
            return None
        v = pd.to_numeric(x, errors="coerce")
        if pd.isna(v):
            return None
        return float(v)
    except Exception:
        return None


# -----------------------------
# LLM helpers
# -----------------------------
def _llm_num_or_np(x) -> float | str:
    """Return a positive float, else 'not provided' (prevents fake 0.0 values)."""
    try:
        v = float(pd.to_numeric(x, errors="coerce"))
        return v if v > 0 else "not provided"
    except Exception:
        return "not provided"


def _apply_llm_explanations(
    dfx: pd.DataFrame,
    *,
    horizon: str,
    top_n: int | None = None,
    write_plan_card: bool = True,
) -> pd.DataFrame:
    if dfx is None or dfx.empty or (not PREMARKET_LLM_ENABLED):
        return dfx

    if "plan_card" not in dfx.columns:
        dfx["plan_card"] = ""
    if "llm_insights" not in dfx.columns:
        dfx["llm_insights"] = ""

    def _num_or_none(x):
        v = pd.to_numeric(x, errors="coerce")
        return None if pd.isna(v) else float(v)

    def _int_or_zero(x) -> int:
        v = pd.to_numeric(x, errors="coerce")
        return int(v) if pd.notna(v) else 0

    def _float_or_zero(x) -> float:
        v = pd.to_numeric(x, errors="coerce")
        return float(v) if pd.notna(v) else 0.0

    gate = int(CONF_GATE)  # 🔒 locked
    pos_size = _num_or_none(os.getenv("DEFAULT_POSITION_SIZE_USD", "0")) or 0.0

    n = len(dfx) if top_n is None else min(int(top_n), len(dfx))
    explain_df = dfx.head(n).copy()

    plan_cards: list[str] = []
    insights: list[str] = []

    for _, rr in explain_df.iterrows():
        payload = {
            "symbol": str(rr.get("symbol", "")).upper().strip(),
            "decision": str(rr.get("decision", "")).strip(),
            "score": _int_or_zero(rr.get("score")),
            "confidence": _int_or_zero(rr.get("confidence")),
            "pct_change": _float_or_zero(rr.get("pct_change")),

            "current": _num_or_none(rr.get("current")),
            "predicted_price": _num_or_none(rr.get("predicted_price")),
            "target_price": _num_or_none(rr.get("target_price")),
            "stop_loss": _num_or_none(rr.get("stop_loss")),
            "forecast_atr": _num_or_none(rr.get("forecast_atr")),

            "forecast_trend": str(rr.get("forecast_trend") or ""),
            "forecast_reason": str(rr.get("forecast_reason") or ""),
            "news_flag": str(rr.get("news_flag") or ""),
            "main_news_title": str(rr.get("main_news_title") or ""),
            "reasons": str(rr.get("reasons") or ""),

            "horizon": horizon,
            "position_size_usd": pos_size,
            "holding_window": str(rr.get("holding_window") or rr.get("horizon") or "intraday"),
            "conf_gate": gate,
        }

        if write_plan_card:
            plan_cards.append(str(safe_explain_pick(payload) or "").strip())

        insights.append(str(safe_explain_insights(payload) or "").strip())

    if write_plan_card:
        dfx.loc[explain_df.index, "plan_card"] = (
            pd.Series(plan_cards, index=explain_df.index).astype(str)
        )

    dfx.loc[explain_df.index, "llm_insights"] = (
        pd.Series(insights, index=explain_df.index).astype(str)
    )
    return dfx


# -----------------------------
# Output helpers (env-aware)
# -----------------------------
def env_base_dir() -> Path:
    base = Path(__file__).resolve().parent / "outputs" / APP_ENV
    base.mkdir(parents=True, exist_ok=True)
    return base


def logs_dir() -> Path:
    p = env_base_dir() / "logs"
    p.mkdir(parents=True, exist_ok=True)
    return p


def run_dir(now: datetime, mode: str) -> Path:
    day = now.strftime("%Y%m%d")
    p = env_base_dir() / "runs" / day / mode
    p.mkdir(parents=True, exist_ok=True)
    return p


def out_path(filename: str, *, now: datetime | None = None, mode: str | None = None, kind: str = "runs") -> str:
    if kind == "logs":
        return str(logs_dir() / filename)
    if now is None or mode is None:
        return str(env_base_dir() / filename)
    return str(run_dir(now, mode) / filename)


def ensure_csv_exists(path: str, header_cols: list[str]) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    if (not p.exists()) or p.stat().st_size == 0:
        pd.DataFrame(columns=header_cols).to_csv(path, index=False)


def make_run_id(now: datetime) -> str:
    return now.strftime("%Y%m%d_%H%M%S")


def _premarket_email_marker(now: datetime) -> Path:
    d = run_dir(now, "premarket")
    run_date = now.strftime("%Y-%m-%d")
    return d / f"email_sent_{run_date}.txt"


# -----------------------------
# Email routing (env-aware)
# -----------------------------
EMAIL_SUBJECT_PREFIX = EMAIL_SUBJECT_PREFIX_LOCAL if IS_LOCAL else EMAIL_SUBJECT_PREFIX_PROD
EFFECTIVE_RECEIVER_EMAIL = (LOCAL_RECEIVER_EMAIL or RECEIVER_EMAIL) if IS_LOCAL else RECEIVER_EMAIL


def send_email(subject: str, html_body: str, attachment_path: str | None = None) -> bool:
    final_subject = f"{EMAIL_SUBJECT_PREFIX} {subject}"
    return _send_email(
        subject=final_subject,
        html_body=html_body,
        to_email=EFFECTIVE_RECEIVER_EMAIL,
        from_email=SENDER_EMAIL,
        attachment_path=attachment_path,
    )


# -----------------------------
# Styling
# -----------------------------
def normalize_color(color: str) -> str:
    if not color:
        color = "#FFFFFF"
    color = color.lstrip("#")
    if len(color) == 6:
        color = "FF" + color
    return color.upper()


def style_excel_sheet(sheet):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color=normalize_color("#2F5597"),
        end_color=normalize_color("#2F5597"),
        fill_type="solid",
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if sheet.max_row < 1 or sheet.max_column < 1:
        return

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    sheet.freeze_panes = "A2"

    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = min(max_length + 2, 55)

    headers = {str(c.value).strip(): idx + 1 for idx, c in enumerate(sheet[1]) if c.value}
    score_label_col = headers.get("score_label")
    decision_col = headers.get("decision")

    for r in range(2, sheet.max_row + 1):
        try:
            if score_label_col:
                cell = sheet.cell(row=r, column=score_label_col)
                label = str(cell.value or "")
                color = SCORE_COLORS.get(label, "#FFFFFF")
                cell.fill = PatternFill(
                    start_color=normalize_color(color),
                    end_color=normalize_color(color),
                    fill_type="solid",
                )
                cell.alignment = center

            if decision_col:
                dcell = sheet.cell(row=r, column=decision_col)
                dval = str(dcell.value or "")
                if dval == "Strong Buy":
                    c = "#92D050"
                elif dval == "Moderate":
                    c = "#FFF2CC"
                else:
                    c = "#F4CCCC"
                dcell.fill = PatternFill(
                    start_color=normalize_color(c),
                    end_color=normalize_color(c),
                    fill_type="solid",
                )
                dcell.alignment = center
        except Exception:
            pass


# -----------------------------
# Helpers
# -----------------------------
def map_score_to_decision(score: int) -> str:
    if score >= SCORE_HIGH:
        return "Strong Buy"
    if score >= SCORE_MEDIUM:
        return "Moderate"
    return "Not Advisable"


def extract_headline_from_html(news_html: str) -> str:
    if not news_html:
        return ""
    news_html = str(news_html).replace("\uFFFC", "").strip()
    m = re.search(r">(.*?)</a>", news_html)
    return (m.group(1).strip() if m else re.sub(r"<.*?>", "", news_html).strip())


def extract_url_from_html(news_html: str) -> str:
    if not news_html:
        return ""
    m = re.search(r'href="([^"]+)"', str(news_html))
    return m.group(1).strip() if m else ""


def news_flag_from_headlines(headlines: List[str]) -> str:
    if not headlines:
        return "🟡"
    score = 0
    for h in headlines:
        t = (h or "").lower()
        if any(w in t for w in POS_WORDS):
            score += 1
        if any(w in t for w in NEG_WORDS):
            score -= 1
    if score >= 1:
        return "🟢"
    if score <= -1:
        return "🔴"
    return "🟡"


def get_market_snapshot() -> dict:
    out = {
        "trend": "up",
        "spy_gap_pct": 0.0,
        "vix": None,
        "breadth": None,
    }
    try:
        spy = yf.Ticker("SPY").history(period="2d", auto_adjust=False)
        if not spy.empty and len(spy) >= 2:
            prev_close = float(spy["Close"].iloc[-2])
            last_close = float(spy["Close"].iloc[-1])
            out["trend"] = "up" if last_close > prev_close else "down"
            out["spy_gap_pct"] = ((last_close - prev_close) / prev_close) * 100.0
    except Exception:
        pass
    try:
        vix = yf.Ticker("^VIX").history(period="1d", auto_adjust=False)
        if not vix.empty:
            out["vix"] = float(vix["Close"].iloc[-1])
    except Exception:
        pass

    # Market breadth (Advance / Decline)
    try:
        tickers = fetch_sp500_tickers()
        adv = 0
        dec = 0

        for t in tickers[:200]:  # limit for speed
            try:
                h = yf.Ticker(t).history(period="2d", auto_adjust=False)
                if len(h) >= 2:
                    if h["Close"].iloc[-1] > h["Close"].iloc[-2]:
                        adv += 1
                    else:
                        dec += 1
            except Exception:
                pass

        out["breadth"] = adv - dec

    except Exception:
        pass

    return out


# ============================================================
# 🔒 LOCKED: Confidence model (score + market + news + ATR + rel_strength)
# - Do not change without version bump
# ============================================================

def compute_confidence(
    score_val: int,
    pct_change: float,
    market_trend: str,
    news_flag: str,
    atr_pct: float = 0.0,
    rel_strength: float = 0.0,
) -> int:

    score_val = max(0, min(int(score_val), 100))

    # base confidence from score (1–10)
    base = round((score_val / 100.0) * 10.0)

    # move-based volatility penalty
    move = min(abs(float(pct_change or 0.0)), 8.0)
    vol_penalty = 2 if move >= 6 else (1 if move >= 4 else 0)

    # ATR volatility penalty
    atr_penalty = 2 if atr_pct >= 7 else (1 if atr_pct >= 5 else 0)

    # market + news adjustments
    market_bonus = 1 if market_trend == "up" else 0
    news_bonus = 1 if news_flag == "🟢" else (-1 if news_flag == "🔴" else 0)

    # relative strength vs SPY
    strength_bonus = 1 if rel_strength >= 3 else 0

    conf = int(base + market_bonus + news_bonus + strength_bonus - vol_penalty - atr_penalty)

    return max(1, min(conf, 10))


def _volatility_target_stop(row: pd.Series) -> tuple[Optional[float], Optional[float]]:
    """
    Volatility-aware target/stop:
      target = current + 0.80 * ATR
      stop   = current - 0.60 * ATR
    Falls back to existing target/stop/predicted if ATR missing.
    """
    entry = _safe_float(row.get("current"))
    atr = _safe_float(row.get("forecast_atr"))

    tgt = _safe_float(row.get("target_price"))
    stp = _safe_float(row.get("stop_loss"))

    if entry is not None and entry > 0 and atr is not None and atr > 0:
        return (entry + (0.80 * atr), entry - (0.60 * atr))

    if tgt is None:
        tgt = _safe_float(row.get("predicted_price"))

    if stp is None and entry is not None and entry > 0:
        stp = entry * 0.98

    if tgt is None and entry is not None and entry > 0:
        tgt = entry * 1.01

    return tgt, stp


def _infer_intraday_target_stop(row: pd.Series) -> tuple[Optional[float], Optional[float]]:
    entry = _safe_float(row.get("current"))
    tgt = _safe_float(row.get("target_price"))
    stp = _safe_float(row.get("stop_loss"))

    if tgt is None:
        tgt = _safe_float(row.get("predicted_price"))

    if stp is None and entry is not None and entry > 0:
        conf_raw = pd.to_numeric(row.get("confidence"), errors="coerce")
        conf = int(conf_raw) if pd.notna(conf_raw) else 5
        stp_pct = 0.012 if conf >= 7 else (0.015 if conf >= 6 else 0.02)
        stp = entry * (1.0 - stp_pct)

    if tgt is None and entry is not None and entry > 0:
        conf_raw = pd.to_numeric(row.get("confidence"), errors="coerce")
        conf = int(conf_raw) if pd.notna(conf_raw) else 5
        tgt_pct = 0.015 if conf >= 7 else (0.012 if conf >= 6 else 0.01)
        tgt = entry * (1.0 + tgt_pct)

    return tgt, stp


# -----------------------------
# Deterministic Plan Card (LOCKED FORMAT)
# -----------------------------
def _entry_range(cur: Optional[float], atr: Optional[float]) -> tuple[Optional[float], Optional[float]]:
    if cur is None or cur <= 0:
        return None, None
    if atr is None or atr <= 0:
        lo = cur * 0.9925
        hi = cur * 1.0075
        return lo, hi
    lo = cur - 0.35 * atr
    hi = cur + 0.35 * atr
    return lo, hi


def _time_window_bucket(conf: int, trend: str) -> str:
    trend = (trend or "").lower()
    if conf >= 7 and trend == "up":
        return "2–3 days"
    if conf >= 6:
        return "1–2 days"
    return "day"


def _plan_card_row(r: pd.Series) -> str:
    cur = _safe_float(r.get("current"))
    tgt = _safe_float(r.get("target_price"))
    stp = _safe_float(r.get("stop_loss"))
    atr = _safe_float(r.get("forecast_atr"))
    conf = int(pd.to_numeric(r.get("confidence"), errors="coerce") or 0)
    trend = str(r.get("forecast_trend") or "")

    lo, hi = _entry_range(cur, atr)
    tw = _time_window_bucket(conf, trend)

    def m(x):
        return "" if x is None else f"{x:.2f}"

    pl_up = pl_dn = rr = None
    up_pct = dn_pct = None

    if cur and tgt and stp and cur > 0:
        pl_up = tgt - cur
        pl_dn = stp - cur
        up_pct = (pl_up / cur) * 100.0
        dn_pct = (pl_dn / cur) * 100.0
        risk = abs(pl_dn)
        rr = (pl_up / risk) if risk > 0 else None

    return "\n".join([
        f"• Entry range: ${m(lo)}–${m(hi)} (cur ${m(cur)})" if (lo is not None and hi is not None) else f"• Entry range: not provided (cur ${m(cur)})",
        f"• Time window: {tw}",
        f"• Levels: target ${m(tgt)} | stop ${m(stp)}",
        f"• P/L: +${m(pl_up)} ({m(up_pct)}%) | ${m(pl_dn)} ({m(dn_pct)}%) | R:R {m(rr)}" if rr is not None else "• P/L: not provided",
        f"• Stance reason: conf {r.get('confidence')} vs gate {CONF_GATE} | news {r.get('news_flag')}",
    ])


# -----------------------------
# NEW: RAW enrichment + High-vote + FINAL_VIEW
# -----------------------------
def _enrich_raw_movers(raw_df: pd.DataFrame, all_scored_df: pd.DataFrame) -> pd.DataFrame:
    """RAW_MOVERS typically lacks target/stop/atr/trend. Enrich by joining ALL_SCORED on symbol."""
    raw_df = raw_df.copy() if raw_df is not None else pd.DataFrame()
    if raw_df.empty:
        return raw_df

    if "symbol" not in raw_df.columns or all_scored_df is None or all_scored_df.empty:
        return raw_df

    keep_cols = [
        "symbol",
        "predicted_price", "target_price", "stop_loss",
        "forecast_trend", "forecast_atr", "forecast_reason",
        "score", "score_label", "confidence",
        "news_flag", "main_news_title", "main_news_link",
        "reasons", "decision",
    ]
    enrich = all_scored_df.copy()
    for c in keep_cols:
        if c not in enrich.columns:
            enrich[c] = pd.NA
    enrich = enrich[keep_cols].drop_duplicates(subset=["symbol"])

    out = raw_df.merge(enrich, on="symbol", how="left", suffixes=("", "_y"))
    return out


def _vote_score(rr: pd.Series) -> int:
    """Deterministic 'high-vote' score. No LLM involved."""
    score = 0
    conf = pd.to_numeric(rr.get("confidence"), errors="coerce")
    scr = pd.to_numeric(rr.get("score"), errors="coerce")
    pct = pd.to_numeric(rr.get("pct_change"), errors="coerce")
    trend = str(rr.get("forecast_trend") or "").lower()
    news = str(rr.get("news_flag") or "")

    if pd.notna(conf) and int(conf) >= 6:
        score += 3
    if pd.notna(scr) and float(scr) >= 60:
        score += 2
    if pd.notna(pct) and abs(float(pct)) >= 5.0:
        score += 1
    if trend == "up":
        score += 1

    mover_signal = str(rr.get("mover_signal") or "")
    if "❌" in mover_signal:
        score -= 2

    if news == "🔴":
        score -= 2
    elif news == "🟡":
        score -= 1

    return int(score)


def _build_final_view(
    *,
    picks_df: pd.DataFrame,
    monitor_df: pd.DataFrame,
    all_scored_df: pd.DataFrame,
    raw_enriched_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    FINAL_VIEW:
      - If PICKS exist -> GO list (top picks)
      - Else -> WATCH list (top 3 high-vote from combined sources)
    """
    cols_min = [
        "symbol", "current", "pct_change",
        "target_price", "stop_loss", "predicted_price",
        "forecast_trend", "forecast_atr", "forecast_reason",
        "score", "score_label", "confidence",
        "news_flag", "main_news_title", "main_news_link",
        "reasons", "decision",
        "plan_card", "llm_insights",
    ]

    def _ensure_cols(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy() if df is not None else pd.DataFrame()
        for c in cols_min:
            if c not in df.columns:
                df[c] = pd.NA
        return df

    picks_df = _ensure_cols(picks_df)
    monitor_df = _ensure_cols(monitor_df)
    all_scored_df = _ensure_cols(all_scored_df)
    raw_enriched_df = _ensure_cols(raw_enriched_df)

    if picks_df is not None and not picks_df.empty:
        final = picks_df.copy()
        final["stance"] = "GO"
        final["decision"] = "GO"  # ✅ NEW (presentation)
        final["vote_score"] = 0
        final["stance_reason"] = final.apply(
            lambda r: f"conf {r.get('confidence')} vs gate {CONF_GATE} | news {r.get('news_flag')}",
            axis=1
        )
        return final.reset_index(drop=True)

    combined = pd.concat([monitor_df, all_scored_df, raw_enriched_df], ignore_index=True)
    combined = combined.dropna(subset=["symbol"]).copy()
    combined["symbol"] = combined["symbol"].astype(str).str.upper().str.strip()
    combined = combined.drop_duplicates(subset=["symbol"])

    combined["vote_score"] = combined.apply(_vote_score, axis=1)
    combined = combined.sort_values(by=["vote_score", "confidence", "score"], ascending=False)

    final = combined.head(FINAL_VIEW_TOP_N).copy()
    final["stance"] = "WATCH"
    final["decision"] = "WATCH"  # ✅ NEW (presentation)
    final["stance_reason"] = final.apply(
        lambda
            r: f"conf {r.get('confidence')} vs gate {CONF_GATE} | vote {r.get('vote_score')} | news {r.get('news_flag')}",
        axis=1
    )
    return final.reset_index(drop=True)


# -----------------------------
# Logs
# -----------------------------
DAILY_LOG_CSV = out_path("daily_stock_log.csv", kind="logs")
RECO_LOG_CSV = out_path("recommendations_log.csv", kind="logs")

ensure_csv_exists(RECO_LOG_CSV, [
    "run_ts", "run_date", "mode", "app_env",
    "symbol", "current", "pct_change",
    "decision", "score", "score_label", "confidence",
    "predicted_price", "target_price", "stop_loss",
    "forecast_trend", "forecast_atr",
    "news_flag", "main_news_title", "main_news_link",
    "reasons",
])

ensure_csv_exists(DAILY_LOG_CSV, [
    "run_ts",
    "run_date", "mode", "symbol", "price_category",
    "current", "predicted_price", "target_price", "stop_loss",
    "forecast_trend", "forecast_atr", "forecast_reason",
    "trade_plan", "earnings_risk",
    "decision", "score", "score_label", "confidence",
    "reasons", "news_flag", "main_news_title", "main_news_link",
    "plan_card",
    "llm_insights",
    "key_insight",
    "expected_rr",
    "win_prob",
    "rel_strength",
    "candle_bias",
    "body_pct",
    "upper_wick_pct",
    "lower_wick_pct",
    "range_low",
    "range_high",
    "range_position_pct",
    "range_zone",
    "near_support",
    "near_resistance",
])


def append_recommendations_log(df_reco: pd.DataFrame, now: datetime, mode: str) -> None:
    if df_reco is None or df_reco.empty:
        return

    out = df_reco.copy()
    out.insert(0, "run_ts", now.strftime("%Y-%m-%d %H:%M:%S"))
    out.insert(1, "run_date", now.strftime("%Y-%m-%d"))
    out.insert(2, "mode", mode)
    out.insert(3, "app_env", APP_ENV)

    cols_keep = [
        "run_ts", "run_date", "mode", "app_env",
        "symbol", "current", "pct_change",
        "decision", "score", "score_label", "confidence",
        "predicted_price", "target_price", "stop_loss",
        "forecast_trend", "forecast_atr",
        "news_flag", "main_news_title", "main_news_link",
        "reasons",
    ]
    for c in cols_keep:
        if c not in out.columns:
            out[c] = pd.NA

    for c in ["current", "pct_change", "predicted_price", "target_price", "stop_loss", "forecast_atr"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")

    out = out[cols_keep]
    file_exists = os.path.exists(RECO_LOG_CSV) and os.path.getsize(RECO_LOG_CSV) > 0
    out.to_csv(RECO_LOG_CSV, mode="a", header=not file_exists, index=False)


def _clean_text_cell(x) -> str:
    s = "" if x is None else str(x)
    s = s.replace("\uFFFC", "").strip()
    s = re.sub(r"\bnan\b", "", s, flags=re.IGNORECASE).strip()
    s = re.sub(r"\s*'[^']+'!\d+:\d+\s*$", "", s).strip()
    s = re.sub(r"\s*\b[A-Z]{1,3}\d+\b\s*$", "", s).strip()
    s = re.sub(r"\s*\b\d+:\d+\b\s*$", "", s).strip()
    s = re.sub(r"\s{2,}", " ", s).strip()

    # ✅ remove ".V11" style suffix coming from LLM
    s = re.sub(r"\.V\d+\b\s*$", "", s).strip()

    # existing cleanup
    s = re.sub(r"\bV\d+\b\s*$", "", s).strip()

    # ✅ NEW: trim trailing punctuation artifacts like "flags=...,"
    s = re.sub(r"[,\s]+$", "", s).strip()

    return s


def append_daily_log(df_rows: pd.DataFrame, now: datetime, run_date: str, mode: str) -> None:
    if df_rows is None or df_rows.empty:
        return

    cols = [
        "run_ts",
        "run_date", "mode", "symbol", "price_category",
        "current", "predicted_price", "target_price", "stop_loss",
        "forecast_trend", "forecast_atr", "forecast_reason",
        "trade_plan", "earnings_risk",
        "decision", "score", "score_label", "confidence",
        "reasons", "news_flag", "main_news_title", "main_news_link",
        "plan_card",
        "llm_insights",
        "key_insight",
        "expected_rr", "win_prob", "rel_strength",
        "candle_bias",
        "body_pct",
        "upper_wick_pct",
        "lower_wick_pct",
        "range_low",
        "range_high",
        "range_position_pct",
        "range_zone",
        "near_support",
        "near_resistance",
    ]

    out = df_rows.copy()
    out["run_ts"] = now.strftime("%Y-%m-%d %H:%M:%S")
    out["run_date"] = run_date
    out["mode"] = mode

    for c in cols:
        if c not in out.columns:
            out[c] = pd.NA
    out = out[cols]

    for c in [
        "current", "predicted_price", "target_price", "stop_loss",
        "forecast_atr", "expected_rr", "win_prob", "rel_strength",
        "body_pct", "upper_wick_pct", "lower_wick_pct",
        "range_low", "range_high", "range_position_pct",
    ]:
        out[c] = pd.to_numeric(out[c], errors="coerce")

    existing = (
        pd.read_csv(DAILY_LOG_CSV)
        if (os.path.exists(DAILY_LOG_CSV) and os.path.getsize(DAILY_LOG_CSV) > 0)
        else pd.DataFrame(columns=cols)
    )
    required_keys = {"run_date", "mode", "symbol"}
    missing = required_keys - set(existing.columns)
    if missing:
        existing = pd.DataFrame(columns=cols)
        existing = existing.reindex(columns=cols)

    merged = pd.concat([existing, out], ignore_index=True)
    merged["symbol"] = merged["symbol"].astype(str).str.upper().str.strip()
    merged = merged.drop_duplicates(subset=["run_date", "mode", "symbol"], keep="last")
    merged.to_csv(DAILY_LOG_CSV, index=False)


# -----------------------------
# Excel writer
# -----------------------------
def write_premarket_excel(
    excel_path: str,
    final_view_df: pd.DataFrame,
    picks_df: pd.DataFrame,
    candidates_df: pd.DataFrame,
    monitor_df: pd.DataFrame,
    all_scored_df: pd.DataFrame,
    raw_movers_df: pd.DataFrame,
    rf: Path,
) -> None:
    try:
        excel_path_p = Path(excel_path)
        excel_path_p.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(excel_path, engine="openpyxl") as xw:
            (final_view_df if final_view_df is not None else pd.DataFrame()).to_excel(xw, sheet_name="FINAL_VIEW", index=False)
            (picks_df if picks_df is not None else pd.DataFrame()).to_excel(xw, sheet_name="PICKS", index=False)
            (candidates_df if candidates_df is not None else pd.DataFrame()).to_excel(xw, sheet_name="CANDIDATES", index=False)
            (monitor_df if monitor_df is not None else pd.DataFrame()).to_excel(xw, sheet_name="MONITOR_TOP20", index=False)
            (all_scored_df if all_scored_df is not None else pd.DataFrame()).to_excel(xw, sheet_name="ALL_SCORED", index=False)
            (raw_movers_df if raw_movers_df is not None else pd.DataFrame()).to_excel(xw, sheet_name="RAW_MOVERS", index=False)

        wb = load_workbook(excel_path)
        for s in wb.sheetnames:
            style_excel_sheet(wb[s])
        wb.save(excel_path)

    except Exception as e:
        (Path(rf) / "excel_write_error.txt").write_text(repr(e), encoding="utf-8")


def _expected_rr(entry: float | None, target: float | None, stop: float | None) -> float | None:
    try:
        if entry is None or target is None or stop is None:
            return None
        entry = float(entry); target = float(target); stop = float(stop)
        if entry <= 0:
            return None
        reward = target - entry
        risk = entry - stop
        if risk <= 0:
            return None
        return reward / risk
    except Exception:
        return None


def _win_probability(
    *,
    score: int,
    conf: int,
    rr: float | None,
    trend: str,
    news_flag: str,
    market_regime: str,
) -> int:
    p = 35 + (int(conf) * 4)
    p += int((int(score) - 60) * 0.6)

    t = (trend or "").lower()
    if t == "up":
        p += 3
    elif t == "down":
        p -= 4

    if news_flag == "🟢":
        p += 2
    elif news_flag == "🔴":
        p -= 5

    if market_regime == "risk_on":
        p += 2
    elif market_regime == "risk_off":
        p -= 6

    if rr is not None:
        if rr >= 1.8:
            p += 4
        elif rr >= 1.4:
            p += 2
        elif rr < 1.0:
            p -= 4

    return int(max(1, min(99, round(p))))


# -----------------------------
# Premarket runner
# -----------------------------
def get_avg_dollar_volume(symbol: str, lookback_days: int = 20) -> Optional[float]:
    try:
        hist = yf.Ticker(symbol).history(period=f"{max(lookback_days, 10) + 10}d", auto_adjust=False)
        if hist is None or hist.empty:
            return None
        hist = hist.dropna(subset=["Close", "Volume"]).tail(lookback_days)
        if hist.empty:
            return None
        dv = (hist["Close"] * hist["Volume"]).mean()
        return float(dv) if dv and dv > 0 else None
    except Exception:
        return None


def run_premarket(now: datetime | None = None) -> None:
    now = now or datetime.now(LOCAL_TZ)
    mode = "premarket"
    rf = run_dir(now, mode)
    run_date = now.strftime("%Y-%m-%d")

    marker = _premarket_email_marker(now)
    if marker.exists():
        print("📩 Premarket email already sent for this run_date — skipping resend.")
        return

    tickers = fetch_sp500_tickers()
    movers = calculate_top_movers(tickers, top_n=TOP_N)
    df_raw = pd.DataFrame(movers)
    if not df_raw.empty and "symbol" in df_raw.columns:
        df_raw = df_raw.dropna(subset=["symbol"]).copy()
        df_raw["symbol"] = df_raw["symbol"].astype(str).str.upper().str.strip()
        df_raw = df_raw.drop_duplicates(subset=["symbol"], keep="first")

    excel_path = out_path(
        f"premarket_{now.strftime('%Y%m%d')}_{make_run_id(now)}.xlsx",
        now=now, mode=mode, kind="runs"
    )

    if df_raw.empty:
        html = f"""
        <h2>🌅 Premarket ({run_date})</h2>
        <p>No movers returned.</p>
        """
        send_email(f"🌅 Premarket ({run_date})", html, attachment_path=excel_path)
        return

    df_raw["pct_change"] = pd.to_numeric(df_raw.get("pct_change"), errors="coerce").fillna(0.0)
    df_raw["current"] = pd.to_numeric(df_raw.get("current"), errors="coerce").fillna(0.0)

    # Liquidity filter (BEFORE scoring loop)
    if not df_raw.empty:
        liq_cache: dict[str, Optional[float]] = {}
        keep_rows = []

        for _, r in df_raw.iterrows():
            sym = str(r.get("symbol", "")).upper().strip()
            price = float(r.get("current", 0.0) or 0.0)

            if price < MIN_PRICE:
                continue

            if sym not in liq_cache:
                liq_cache[sym] = get_avg_dollar_volume(sym, lookback_days=20)

            adv = liq_cache[sym]
            if adv is None or adv < MIN_AVG_DOLLAR_VOL:
                continue

            keep_rows.append(r)

        df_raw = pd.DataFrame(keep_rows)

        if not df_raw.empty and "symbol" in df_raw.columns:
            df_raw = df_raw.dropna(subset=["symbol"]).copy()
            df_raw["symbol"] = df_raw["symbol"].astype(str).str.upper().str.strip()
            df_raw = df_raw.drop_duplicates(subset=["symbol"], keep="first")

        if df_raw.empty:
            empty = pd.DataFrame()
            write_premarket_excel(
                excel_path,
                final_view_df=empty,
                picks_df=empty,
                candidates_df=empty,
                monitor_df=empty,
                all_scored_df=empty,
                raw_movers_df=empty,
                rf=Path(rf),
            )
            html = f"""
            <h2>🌅 Premarket ({run_date})</h2>
            <p>No movers passed liquidity filters (price≥${MIN_PRICE}, avg $vol≥${int(MIN_AVG_DOLLAR_VOL):,}).</p>
            """
            if send_email(f"🌅 Premarket ({run_date})", html, attachment_path=excel_path):
                marker.write_text("sent\n", encoding="utf-8")
            return

    snapshot = get_market_snapshot()
    market_trend = snapshot.get("trend", "up")
    spy_pct = float(snapshot.get("spy_gap_pct") or 0.0)

    rows: List[Dict[str, Any]] = []

    candle_biases = []
    body_pcts = []
    upper_wick_pcts = []
    lower_wick_pcts = []
    range_lows = []
    range_highs = []
    range_pos_pcts = []
    range_zones = []
    near_supports = []
    near_resistances = []

    for _, r in df_raw.iterrows():

        sym = str(r.get("symbol", "")).upper().strip()
        current = float(r.get("current", 0.0) or 0.0)
        pct_change = float(r.get("pct_change", 0.0) or 0.0)

        pa = compute_price_action_summary(sym, current)

        candle_biases.append(pa.get("candle_bias"))
        body_pcts.append(pa.get("body_pct"))
        upper_wick_pcts.append(pa.get("upper_wick_pct"))
        lower_wick_pcts.append(pa.get("lower_wick_pct"))
        range_lows.append(pa.get("range_low"))
        range_highs.append(pa.get("range_high"))
        range_pos_pcts.append(pa.get("range_position_pct"))
        range_zones.append(pa.get("range_zone"))
        near_supports.append(pa.get("near_support"))
        near_resistances.append(pa.get("near_resistance"))

        # ✅ NEW: rel strength vs SPY
        rel_strength = pct_change - spy_pct

        # score first
        score_val, score_label, reasons = get_predictive_score_with_reasons(sym)
        score_val = int(score_val)
        decision = map_score_to_decision(score_val)

        # news flag
        news_items = fetch_news_links(sym, max_articles=1)
        main_item = news_items[0] if news_items else ""
        title = extract_headline_from_html(main_item)
        link = extract_url_from_html(main_item)
        flag = news_flag_from_headlines([title])

        # forecast for ATR
        f = forecast_price_levels(sym, current=current, score=score_val, horizon="intraday")

        atr_val = getattr(f, "atr", None)
        try:
            atr_val = float(atr_val) if atr_val is not None else None
        except Exception:
            atr_val = None

        # NEW: calculate ATR volatility %
        atr_pct = (atr_val / current) * 100.0 if (atr_val and current > 0) else 0.0

        # ✅ UPDATED confidence calculation (ONLY ONCE, AFTER vars exist)
        conf = compute_confidence(
            score_val,
            pct_change,
            market_trend,
            flag,
            atr_pct=atr_pct,
            rel_strength=rel_strength,
        )

        if atr_val is not None and atr_val > 0 and current > 0:
            tgt_calc = current + (0.80 * atr_val)
            stp_calc = current - (0.60 * atr_val)
        else:
            tgt_calc = None
            stp_calc = None

        price_cat = get_price_category(current)

        rows.append({
            "symbol": sym,
            "price_category": price_cat,
            "current": current,
            "pct_change": pct_change,
            "predicted_price": getattr(f, "predicted_price", pd.NA),
            "target_price": (tgt_calc if tgt_calc is not None else getattr(f, "target_price", pd.NA)),
            "stop_loss": (stp_calc if stp_calc is not None else getattr(f, "stop_loss", pd.NA)),
            "forecast_trend": getattr(f, "trend", ""),
            "forecast_atr": getattr(f, "atr", pd.NA),
            "forecast_reason": getattr(f, "reason", ""),
            "trade_plan": "Enter near current; target-hit win tracking uses intraday high vs target.",
            "earnings_risk": "",
            "decision": decision,
            "score": score_val,
            "score_label": score_label,
            "confidence": int(conf),
            "reasons": reasons,
            "news_flag": flag,
            "main_news_title": title,
            "main_news_link": link,
            "candle_bias": pa.get("candle_bias"),
            "body_pct": pa.get("body_pct"),
            "upper_wick_pct": pa.get("upper_wick_pct"),
            "lower_wick_pct": pa.get("lower_wick_pct"),
            "range_low": pa.get("range_low"),
            "range_high": pa.get("range_high"),
            "range_position_pct": pa.get("range_position_pct"),
            "range_zone": pa.get("range_zone"),
            "near_support": pa.get("near_support"),
            "near_resistance": pa.get("near_resistance"),
        })

    out_df = pd.DataFrame(rows)

    for c in ["current", "pct_change", "predicted_price", "target_price", "stop_loss", "forecast_atr"]:
        if c in out_df.columns:
            out_df[c] = pd.to_numeric(out_df[c], errors="coerce")

    all_scored_df = out_df.sort_values(by=["confidence", "score"], ascending=False).copy()
    monitor_df = all_scored_df.head(20).copy()

    candidates_df = out_df.copy()
    conf_gate = int(CONF_GATE)  # 🔒 lock per run
    candidates_df = candidates_df[pd.to_numeric(candidates_df["confidence"], errors="coerce").fillna(0).astype(int) >= conf_gate].copy()

    candidates_df = candidates_df[
        (candidates_df["current"] <= MAX_PRICE) |
        ((candidates_df["score"] >= ELITE_SCORE_OVERRIDE) & (candidates_df["confidence"] >= ELITE_CONF_OVERRIDE))
    ].copy()
    candidates_df = candidates_df.sort_values(by=["confidence", "score"], ascending=False)

    # -----------------------------
    # Market regime detection (VIX + SPY gap) ✅ MUST happen BEFORE picks_df
    # -----------------------------
    vix_raw = snapshot.get("vix")
    try:
        vix = float(vix_raw) if vix_raw is not None else None
    except Exception:
        vix = None

    gap = float(snapshot.get("spy_gap_pct") or 0.0)
    breadth = snapshot.get("breadth")

    market_regime = "unknown"

    if vix is not None:
        if vix >= 22 or gap <= -0.8:
            market_regime = "risk_off"
        elif vix <= 16 and gap >= 0.2:
            market_regime = "risk_on"

    if breadth is not None:
        if breadth <= -80:
            market_regime = "risk_off"
        elif breadth >= 80:
            market_regime = "risk_on"

    # Fallback regime when VIX/breadth are missing
    if market_regime == "unknown":
        if gap >= 0.2 and market_trend == "up":
            market_regime = "risk_on"
        elif gap <= -0.8 and market_trend == "down":
            market_regime = "risk_off"
        else:
            market_regime = "neutral"

    # 🔒 Risk-off trading guard (single source of truth)
    risk_off_block = (market_regime == "risk_off")

    # -----------------------------
    # GO picks creation ✅ AFTER market_regime is known
    # -----------------------------
    if risk_off_block:
        print("⚠️ Market regime = risk_off — suppressing GO trades.")
        picks_df = pd.DataFrame()
    else:
        picks_df = candidates_df.head(int(TRADE_MAX_PICKS)).copy().reset_index(drop=True)

    if not picks_df.empty:
        tgt_fix, stp_fix = [], []
        for _, r in picks_df.iterrows():
            tgt, stp = _volatility_target_stop(r)
            tgt_fix.append(tgt if tgt is not None else pd.NA)
            stp_fix.append(stp if stp is not None else pd.NA)
        picks_df["target_price"] = pd.to_numeric(pd.Series(tgt_fix, index=picks_df.index), errors="coerce")
        picks_df["stop_loss"] = pd.to_numeric(pd.Series(stp_fix, index=picks_df.index), errors="coerce")

    raw_enriched_df = _enrich_raw_movers(df_raw, all_scored_df)

    final_view_df = _build_final_view(
        picks_df=picks_df,
        monitor_df=monitor_df,
        all_scored_df=all_scored_df,
        raw_enriched_df=raw_enriched_df,
    )

    # FINAL_VIEW: enforce levels + deterministic plan_card
    if not final_view_df.empty:
        tgt_fix, stp_fix = [], []
        for _, r in final_view_df.iterrows():
            tgt, stp = _volatility_target_stop(r)
            tgt_fix.append(tgt if tgt is not None else pd.NA)
            stp_fix.append(stp if stp is not None else pd.NA)

        final_view_df["target_price"] = pd.to_numeric(pd.Series(tgt_fix, index=final_view_df.index), errors="coerce")
        final_view_df["stop_loss"] = pd.to_numeric(pd.Series(stp_fix, index=final_view_df.index), errors="coerce")
        final_view_df["plan_card"] = final_view_df.apply(lambda r: _plan_card_row(r), axis=1)
    else:
        final_view_df = final_view_df.copy()
        final_view_df["plan_card"] = ""

    # ✅ NEW: rel_strength for FINAL_VIEW (and daily log)
    final_view_df["rel_strength"] = (
        pd.to_numeric(final_view_df.get("pct_change"), errors="coerce").fillna(0.0)
        - float(snapshot.get("spy_gap_pct") or 0.0)
    )

    final_view_df["expected_rr"] = final_view_df.apply(
        lambda r: _expected_rr(
            _safe_float(r.get("current")),
            _safe_float(r.get("target_price")),
            _safe_float(r.get("stop_loss")),
        ),
        axis=1,
    )

    final_view_df["win_prob"] = final_view_df.apply(
        lambda r: _win_probability(
            score=int(pd.to_numeric(r.get("score"), errors="coerce") or 0),
            conf=int(pd.to_numeric(r.get("confidence"), errors="coerce") or 0),
            rr=_safe_float(r.get("expected_rr")),
            trend=str(r.get("forecast_trend") or ""),
            news_flag=str(r.get("news_flag") or ""),
            market_regime=str(market_regime),
        ),
        axis=1,
    )

    # Always copy before mutation/scrub
    final_view_df = final_view_df.copy()
    picks_df = picks_df.copy()
    monitor_df = monitor_df.copy()
    candidates_df = candidates_df.copy()
    all_scored_df = all_scored_df.copy()
    raw_enriched_df = raw_enriched_df.copy()

    # LLM explanations (OPTIONAL)
    picks_df = _apply_llm_explanations(picks_df, horizon="premarket_picks", top_n=PREMARKET_LLM_TOP_N)
    monitor_df = _apply_llm_explanations(monitor_df, horizon="premarket_monitor", top_n=PREMARKET_LLM_TOP_N)
    candidates_df = _apply_llm_explanations(candidates_df, horizon="premarket_candidates", top_n=PREMARKET_LLM_TOP_N)
    all_scored_df = _apply_llm_explanations(all_scored_df, horizon="premarket_all_scored", top_n=PREMARKET_LLM_TOP_N)

    final_view_df = _apply_llm_explanations(
        final_view_df,
        horizon="premarket_final_view",
        top_n=None,
        write_plan_card=False,  # ✅ keep deterministic plan_card
    )

    raw_enriched_df = _apply_llm_explanations(
        raw_enriched_df,
        horizon="premarket_raw_enriched",
        top_n=PREMARKET_LLM_TOP_N,
    )

    # Key insight (deterministic) + add market_regime everywhere
    for df_ in [final_view_df, picks_df, monitor_df, candidates_df, all_scored_df, raw_enriched_df]:
        if df_ is None or df_.empty:
            continue

        df_["market_regime"] = market_regime

        if "key_insight" not in df_.columns:
            df_["key_insight"] = ""

        # ✅ FIX: hard-normalize gate + ints so build_key_insight can't mis-flag "below gate"
        def _insight_payload(r: pd.Series) -> dict:
            d = r.to_dict()

            conf = int(pd.to_numeric(d.get("confidence"), errors="coerce") or 0)
            gate = int(CONF_GATE)

            d["confidence"] = conf
            d["conf_gate"] = gate

            d["min_confidence_to_trade"] = gate
            d["is_tradeable"] = conf >= gate

            d["stance"] = str(d.get("stance") or "")
            d["is_watchlist"] = (d["stance"] == "WATCH")

            vs = d.get("vote_score")
            if vs is None or (isinstance(vs, float) and pd.isna(vs)):
                d["vote_score"] = 0
            else:
                d["vote_score"] = int(pd.to_numeric(vs, errors="coerce") or 0)

            return d

        df_["key_insight"] = df_.apply(lambda r: build_key_insight(_insight_payload(r)), axis=1)

    # scrub trailing artifacts AFTER everything is created
    for df_ in [final_view_df, picks_df, monitor_df, candidates_df, all_scored_df, raw_enriched_df]:
        if df_ is None or df_.empty:
            continue
        for c in ["plan_card", "llm_insights", "key_insight"]:
            if c in df_.columns:
                df_[c] = df_[c].apply(_clean_text_cell)

    append_recommendations_log(candidates_df, now, mode="premarket")
    append_daily_log(final_view_df, now, run_date, mode="premarket")

    # Portfolio open add (best-effort) only when real PICKS exist
    if not picks_df.empty:
        try:
            cfg = PortfolioConfig()
            open_df = load_open_portfolio(cfg)
            for c in ["symbol", "current", "target_price", "stop_loss", "score", "confidence", "decision", "forecast_trend"]:
                if c not in picks_df.columns:
                    picks_df[c] = pd.NA
            updated_open, added_df = add_new_positions_from_picks(cfg, open_df, picks_df, now)
            save_open_portfolio(cfg, updated_open)
            append_open_actions(cfg, added_df)
        except Exception as e:
            (Path(rf) / "portfolio_error.txt").write_text(repr(e), encoding="utf-8")

    # Excel
    write_premarket_excel(
        excel_path,
        final_view_df=final_view_df,
        picks_df=picks_df,
        candidates_df=candidates_df,
        monitor_df=monitor_df,
        all_scored_df=all_scored_df,
        raw_movers_df=raw_enriched_df,
        rf=Path(rf),
    )

    # ============================================================
    # 🔒 LOCKED EMAIL OUTPUT SECTION
    # ============================================================
    def _fmt_money(v) -> str:
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ""
            return f"{float(v):.2f}"
        except Exception:
            return ""

    def _fmt_pct(v) -> str:
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ""
            return f"{float(v):+.2f}%"
        except Exception:
            return ""

    def row_html(rr):
        sym = _html.escape(str(rr.get("symbol", "")))
        news_flag = _html.escape(str(rr.get("news_flag") or ""))

        cur = _fmt_money(rr.get("current"))
        pct = _fmt_pct(rr.get("pct_change"))

        tgt = _fmt_money(rr.get("target_price"))
        stp = _fmt_money(rr.get("stop_loss"))

        conf = _html.escape(str(rr.get("confidence", "")))
        score = _html.escape(str(rr.get("score", "")))
        stance = _html.escape(str(rr.get("stance", rr.get("decision", ""))))

        title = _html.escape(str(rr.get("main_news_title") or ""))
        link = str(rr.get("main_news_link") or "").strip() or "#"

        ki = _html.escape(str(rr.get("key_insight") or ""))[:500].replace("\n", "<br>")
        plan = _html.escape(str(rr.get("plan_card") or ""))[:900].replace("\n", "<br>")
        ins = _html.escape(str(rr.get("llm_insights") or ""))[:600].replace("\n", "<br>")

        return f"""
        <tr>
          <td><b>{news_flag} {sym}</b></td>
          <td>{cur}</td>
          <td>{pct}</td>
          <td>{tgt}</td>
          <td>{stp}</td>
          <td>{conf}</td>
          <td>{score}</td>
          <td>{stance}</td>
          <td style="color:#555;white-space:normal;">{ki}</td>
          <td style="color:#333;white-space:normal;">{plan}</td>
          <td style="color:#555;white-space:normal;">{ins}</td>
          <td><a href="{link}" target="_blank">{title}</a></td>
        </tr>
        """

    rows_html = "\n".join([row_html(rr) for _, rr in final_view_df.iterrows()])
    mode_note = "Qualified Picks (GO)" if (picks_df is not None and not picks_df.empty) else "High Vote Watchlist (WATCH)"

    html = f"""
    <h2>🌅 Premarket ({run_date})</h2>
    <p>
      <b>Market trend:</b> {_html.escape(str(snapshot.get("trend")))} |
      <b>SPY gap:</b> {snapshot.get("spy_gap_pct", 0.0):.2f}% |
      <b>VIX:</b> {snapshot.get("vix")} |
      <b>Breadth:</b> {snapshot.get("breadth")} |
      <b>Regime:</b> {_html.escape(str(market_regime))}
    </p>
    <p><b>Mode:</b> {mode_note}</p>
    <p>Filters: confidence ≥ {MIN_CONFIDENCE_TO_TRADE}, price ≤ ${MAX_PRICE} (elite override allowed).</p>

    <table border="1" cellpadding="6" cellspacing="0"
           style="border-collapse:collapse;font-family:Arial;font-size:13px;">
      <tr style="background:#eee;">
        <th>Symbol</th>
        <th>Price</th>
        <th>%Chg</th>
        <th>Target</th>
        <th>Stop</th>
        <th>Conf</th>
        <th>Score</th>
        <th>Stance</th>
        <th>Key Insight</th>
        <th>Plan Card (Locked)</th>
        <th>Insights</th>
        <th>Headline</th>
      </tr>
      {rows_html}
    </table>

    <p><b>Attachment:</b> Excel included with sheets: FINAL_VIEW, PICKS, CANDIDATES, MONITOR_TOP20, ALL_SCORED, RAW_MOVERS.</p>
    """

    if send_email(f"🌅 Premarket ({run_date})", html, attachment_path=excel_path):
        marker.write_text("sent\n", encoding="utf-8")

    print(f"✅ Premarket complete | final_view={len(final_view_df)} | picks={len(picks_df)} | excel={excel_path}")


if __name__ == "__main__":
    run_premarket()