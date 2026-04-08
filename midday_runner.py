# midday_runner.py (FINAL LOCK + FINAL_VIEW + HIGH-VOTE FALLBACK + DETERMINISTIC PLAN_CARD + DAILY_LOG TRUTH)
from __future__ import annotations

from llm.explain import safe_explain_pick
from context_insight import build_key_insight
import os
import re
import html as _html
from datetime import datetime, time
from pathlib import Path
from zoneinfo import ZoneInfo
from typing import List, Optional
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from config import (
    APP_ENV,
    IS_LOCAL,
    SENDER_EMAIL,
    RECEIVER_EMAIL,
    LOCAL_RECEIVER_EMAIL,
    EMAIL_SUBJECT_PREFIX_LOCAL,
    EMAIL_SUBJECT_PREFIX_PROD,
    TOP_N,
    SUDDEN_MOVER_PCT_THRESHOLD,
    SUDDEN_MOVER_MIN_CONFIDENCE,
    SCORE_HIGH,
    SCORE_MEDIUM,
    MAX_PRICE,
    ELITE_SCORE_OVERRIDE,
    ELITE_CONF_OVERRIDE,

)

from email_sender import send_email as _send_email
from top_movers import fetch_sp500_tickers, calculate_top_movers
from scoring_engine import get_predictive_score_with_reasons
from news_fetcher import fetch_news_links
from forecast_engine import forecast_price_levels  # pred/target/stop
from ml_strategy import predict_win_prob
from price_action import compute_price_action_summary

LOCAL_TZ = ZoneInfo("America/Chicago")

# Market session times (Central)
SESSION_START = time(8, 30)   # 8:30am CT
SESSION_END   = time(15, 0)   # 3:00pm CT

# Debug artifacts (set to "1" to enable)
MIDDAY_DEBUG = os.getenv("MIDDAY_DEBUG", "0") == "1"

# Market words
POS_WORDS = {"beat", "strong", "growth", "surge", "upgrade", "raises", "record", "profit", "wins", "bull"}
NEG_WORDS = {"miss", "drop", "loss", "cuts", "downgrade", "falls", "weak", "lawsuit", "plunge", "bear"}

# 🔒 LOCKED (single source of truth from config.py)
MIDDAY_MIN_CONFIDENCE = int(SUDDEN_MOVER_MIN_CONFIDENCE)
CONF_GATE = int(SUDDEN_MOVER_MIN_CONFIDENCE)

# Guardrails for forecast sanity (long-only)
FORECAST_TARGET_MAX_MULT = float(os.getenv("FORECAST_TARGET_MAX_MULT", "1.20"))  # target <= 20% above entry
FORECAST_STOP_MIN_MULT   = float(os.getenv("FORECAST_STOP_MIN_MULT", "0.80"))    # stop  >= 80% of entry

# LLM controls
LLM_ENABLED = os.getenv("LLM_ENABLED", "1") == "1"
MIDDAY_LLM_TOP_N = int(os.getenv("MIDDAY_LLM_TOP_N", "10"))

# FINAL view controls
FINAL_VIEW_TOP_N = int(os.getenv("FINAL_VIEW_TOP_N", "3"))


# Volume confirmation (optional)
USE_VOLUME_FILTER = os.getenv("USE_VOLUME_FILTER", "0") == "1"
VOL_LOOKBACK_DAYS = int(os.getenv("VOL_LOOKBACK_DAYS", "20"))
VOL_RATIO_MIN = float(os.getenv("VOL_RATIO_MIN", "1.20"))   # e.g., 1.2x avg volume
VOL_CONF_BONUS = int(os.getenv("VOL_CONF_BONUS", "1"))      # +1 conf if confirmed
VOL_CONF_PENALTY = int(os.getenv("VOL_CONF_PENALTY", "1"))  # -1 conf if not confirmed

#
MIN_PRICE = float(os.getenv("MIN_PRICE", "5"))
MIN_AVG_DOLLAR_VOL = float(os.getenv("MIN_AVG_DOLLAR_VOL", "20000000"))  # $20M
#

# -----------------------------
# Output helpers (env-aware)
# -----------------------------
def env_base_dir() -> Path:
    base = Path(__file__).resolve().parent / "outputs" / APP_ENV
    base.mkdir(parents=True, exist_ok=True)
    return base


def logs_dir() -> Path:
    p = env_base_dir() / "logs"
    p.mkdir(parents=True, exist_ok=True)
    return p


def run_dir(now: datetime, mode: str) -> Path:
    day = now.strftime("%Y%m%d")
    p = env_base_dir() / "runs" / day / mode
    p.mkdir(parents=True, exist_ok=True)
    return p


def out_path(filename: str, *, now: datetime | None = None, mode: str | None = None, kind: str = "runs") -> str:
    if kind == "logs":
        return str(logs_dir() / filename)
    if now is None or mode is None:
        return str(env_base_dir() / filename)
    return str(run_dir(now, mode) / filename)


def ensure_csv_exists(path: str, header_cols: list[str]) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    if (not p.exists()) or p.stat().st_size == 0:
        pd.DataFrame(columns=header_cols).to_csv(path, index=False)


def make_run_id(now: datetime) -> str:
    return now.strftime("%Y%m%d_%H%M%S")


def _midday_email_marker(now: datetime) -> Path:
    d = run_dir(now, "midday")
    run_date = now.strftime("%Y-%m-%d")
    return d / f"email_sent_{run_date}.txt"
###________

def _baseline_coach_from_stats(run_date: str, prem_s: dict, mid_s: dict, all_s: dict) -> str:
    try:
        evaluated = int(all_s.get("evaluated", 0) or 0)
        wins = int(all_s.get("wins", 0) or 0)
        losses = int(all_s.get("losses", 0) or 0)
        not_hit = int(all_s.get("not_hit", 0) or 0)
        win_rate = float(all_s.get("win_rate", 0.0) or 0.0)
    except Exception:
        evaluated, wins, losses, not_hit, win_rate = 0, 0, 0, 0, 0.0

    if evaluated <= 0:
        return "\n".join([
            "- What happened: No evaluated trades in stats.",
            "- Likely loss drivers: not provided",
            "- What worked: not provided",
            "- Next-session process tweak: Verify that evaluation rows are being written to performance_log.",
            "- Risk control tweak: Keep risk minimal until evaluation data is stable.",
        ])

    stop_rate = (losses / evaluated * 100.0) if evaluated else 0.0
    not_hit_rate = (not_hit / evaluated * 100.0) if evaluated else 0.0

    # Deterministic, neutral language (no “you should”)
    likely_driver = "Many setups did not reach targets within the session window." if not_hit_rate >= 50 else "Mixed follow-through; some setups did not complete."
    worked = "Low stop-hit rate suggests risk containment held." if losses == 0 else "Stops triggered, indicating adverse moves were contained."

    tweak_parts = []
    if not_hit_rate >= 60:
        tweak_parts.append("Tighten time-scaled targets (esp. midday) or skip late recommendations.")
    if stop_rate >= 40:
        tweak_parts.append("Increase selectivity: raise confidence gate for the weaker session or reduce noisy movers.")
    if not tweak_parts:
        tweak_parts.append("Keep parameters stable; collect more samples before changing thresholds.")

    risk_parts = []
    if evaluated < 6:
        risk_parts.append("Small sample size: keep position sizing conservative.")
    risk_parts.append("Cap number of midday trades when time-left is low.")

    return "\n".join([
        f"- What happened: evaluated={evaluated}, wins={wins}, losses={losses}, not_hit={not_hit}, win_rate={win_rate:.2f}%.",
        f"- Likely loss drivers: {likely_driver}",
        f"- What worked: {worked}",
        f"- Next-session process tweak: {' '.join(tweak_parts)}",
        f"- Risk control tweak: {' '.join(risk_parts)}",
    ])

def get_avg_dollar_volume(symbol: str, lookback_days: int = 20) -> Optional[float]:
    """
    avg_dollar_volume = mean(Close * Volume) over last N trading days.
    """
    try:
        hist = yf.Ticker(symbol).history(period=f"{max(lookback_days, 10) + 10}d", auto_adjust=False)
        if hist is None or hist.empty:
            return None
        hist = hist.dropna(subset=["Close", "Volume"]).tail(lookback_days)
        if hist.empty:
            return None
        dv = (hist["Close"] * hist["Volume"]).mean()
        return float(dv) if dv and dv > 0 else None
    except Exception:
        return None

# -----------------------------
# Email routing (env-aware)
# -----------------------------
EMAIL_SUBJECT_PREFIX = EMAIL_SUBJECT_PREFIX_LOCAL if IS_LOCAL else EMAIL_SUBJECT_PREFIX_PROD
EFFECTIVE_RECEIVER_EMAIL = (LOCAL_RECEIVER_EMAIL or RECEIVER_EMAIL) if IS_LOCAL else RECEIVER_EMAIL


def send_email(subject: str, html_body: str, attachment_path: str | None = None) -> bool:
    final_subject = f"{EMAIL_SUBJECT_PREFIX} {subject}"
    return _send_email(
        subject=final_subject,
        html_body=html_body,
        to_email=EFFECTIVE_RECEIVER_EMAIL,
        from_email=SENDER_EMAIL,
        attachment_path=attachment_path,
    )


# -----------------------------
# Excel styling
# -----------------------------
def normalize_color(color: str) -> str:
    if not color:
        color = "#FFFFFF"
    color = color.lstrip("#")
    if len(color) == 6:
        color = "FF" + color
    return color.upper()


def style_excel_sheet(sheet) -> None:
    if sheet.max_row < 1 or sheet.max_column < 1:
        return

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color=normalize_color("#2F5597"),
        end_color=normalize_color("#2F5597"),
        fill_type="solid",
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    sheet.freeze_panes = "A2"

    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = min(max_length + 2, 55)


def write_midday_excel(
    excel_path: str,
    *,
    final_view_df: pd.DataFrame,
    pass_df: pd.DataFrame,
    all_df: pd.DataFrame,
    threshold_df: pd.DataFrame,
    rf: Path,
) -> None:
    try:
        p = Path(excel_path)
        p.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(excel_path, engine="openpyxl") as xw:
            (final_view_df if final_view_df is not None else pd.DataFrame()).to_excel(xw, sheet_name="FINAL_VIEW", index=False)
            (pass_df if pass_df is not None else pd.DataFrame()).to_excel(xw, sheet_name="PASS", index=False)
            (all_df if all_df is not None else pd.DataFrame()).to_excel(xw, sheet_name="ALL_CANDIDATES", index=False)
            (threshold_df if threshold_df is not None else pd.DataFrame()).to_excel(xw, sheet_name="AFTER_THRESHOLD", index=False)

        wb = load_workbook(excel_path)
        for s in wb.sheetnames:
            style_excel_sheet(wb[s])
        wb.save(excel_path)

    except Exception as e:
        (Path(rf) / "midday_excel_error.txt").write_text(repr(e), encoding="utf-8")


# -----------------------------
# LLM helpers
# -----------------------------
def _ensure_llm_col(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy() if df is not None else pd.DataFrame()
    if "llm_insights" not in df.columns:
        df["llm_insights"] = ""
    return df


def _llm_num_or_np(x) -> float | str:
    try:
        v = float(pd.to_numeric(x, errors="coerce"))
        return v if v > 0 else "not provided"
    except Exception:
        return "not provided"


def _apply_llm_explanations(df: pd.DataFrame, *, horizon: str, top_n: int | None) -> pd.DataFrame:
    df = _ensure_llm_col(df)
    if not LLM_ENABLED or df is None or df.empty:
        return df

    n = len(df) if top_n is None else min(int(top_n), len(df))
    idx = df.head(n).index

    def _num_or_none(x):
        v = pd.to_numeric(x, errors="coerce")
        return None if pd.isna(v) else float(v)

    def _int_or_zero(x) -> int:
        v = pd.to_numeric(x, errors="coerce")
        return int(v) if pd.notna(v) else 0

    def _float_or_zero(x) -> float:
        v = pd.to_numeric(x, errors="coerce")
        return float(v) if pd.notna(v) else 0.0

    def _pos_size_or_none():
        v = pd.to_numeric(os.getenv("DEFAULT_POSITION_SIZE_USD", ""), errors="coerce")
        return None if pd.isna(v) or float(v) <= 0 else float(v)

    def _row_payload(r: pd.Series) -> dict:
        return {
            "symbol": str(r.get("symbol", "")).upper().strip(),
            "decision": str(r.get("decision", "")).strip(),

            "score": _int_or_zero(r.get("score")),
            "confidence": _int_or_zero(r.get("confidence")),
            "pct_change": _float_or_zero(r.get("pct_change")),

            "current": _num_or_none(r.get("current")),
            "predicted_price": _num_or_none(r.get("predicted_price")),
            "target_price": _num_or_none(r.get("target_price")),
            "stop_loss": _num_or_none(r.get("stop_loss")),
            "forecast_atr": _num_or_none(r.get("forecast_atr")),

            "forecast_trend": str(r.get("forecast_trend") or ""),
            "forecast_reason": str(r.get("forecast_reason") or ""),

            "news_flag": str(r.get("news_flag") or ""),
            "main_news_title": str(r.get("main_news_title") or ""),
            "reasons": str(r.get("reasons") or ""),

            "horizon": horizon,
            "position_size_usd": _pos_size_or_none(),
            "holding_window": str(r.get("holding_window") or horizon or "intraday"),
            "conf_gate": int(CONF_GATE),

            "stance": str(r.get("stance") or ""),
            "stance_reason": str(r.get("stance_reason") or ""),
        }

    df.loc[idx, "llm_insights"] = df.loc[idx].apply(
        lambda r: str(safe_explain_pick(_row_payload(r)) or "").strip(),
        axis=1
    )
    return df


# -----------------------------
# Deterministic Plan Card (LOCKED FORMAT)
# -----------------------------
def _safe_float(x) -> Optional[float]:
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return None
        return float(x)
    except Exception:
        return None

def _volatility_target_stop(r: pd.Series) -> tuple[Optional[float], Optional[float]]:
    """
    Volatility-aware target/stop:
      target = current + 0.80 * ATR
      stop   = current - 0.60 * ATR
    Falls back to existing target/stop/predicted if ATR missing.
    """
    entry = _safe_float(r.get("current"))
    atr = _safe_float(r.get("forecast_atr"))

    tgt = _safe_float(r.get("target_price"))
    stp = _safe_float(r.get("stop_loss"))

    if entry is not None and entry > 0 and atr is not None and atr > 0:
        return (entry + (0.80 * atr), entry - (0.60 * atr))

    if tgt is None and r.get("predicted_price") is not None:
        tgt = _safe_float(r.get("predicted_price"))

    if stp is None and entry is not None and entry > 0:
        stp = entry * 0.98

    if tgt is None and entry is not None and entry > 0:
        tgt = entry * 1.01

    return tgt, stp

def _entry_range(cur: Optional[float], atr: Optional[float]) -> tuple[Optional[float], Optional[float]]:
    if cur is None or cur <= 0:
        return None, None
    if atr is None or atr <= 0:
        lo = cur * 0.9925
        hi = cur * 1.0075
        return lo, hi
    lo = cur - 0.35 * atr
    hi = cur + 0.35 * atr
    return lo, hi


def _time_window_bucket(conf: int, trend: str) -> str:
    trend = (trend or "").lower()
    if conf >= 7 and trend == "up":
        return "2–3 days"
    if conf >= 6:
        return "1–2 days"
    return "day"


def _plan_card_row(r: pd.Series) -> str:
    cur = _safe_float(r.get("current"))
    tgt = _safe_float(r.get("target_price"))
    stp = _safe_float(r.get("stop_loss"))
    atr = _safe_float(r.get("forecast_atr"))
    conf = int(pd.to_numeric(r.get("confidence"), errors="coerce") or 0)
    trend = str(r.get("forecast_trend") or "")

    lo, hi = _entry_range(cur, atr)
    tw = _time_window_bucket(conf, trend)

    def m(x):
        return "" if x is None else f"{x:.2f}"

    pl_up = pl_dn = rr = None
    up_pct = dn_pct = None

    if cur and tgt and stp and cur > 0:
        pl_up = tgt - cur
        pl_dn = stp - cur
        up_pct = (pl_up / cur) * 100.0
        dn_pct = (pl_dn / cur) * 100.0
        risk = abs(pl_dn)
        rr = (pl_up / risk) if risk > 0 else None

    return "\n".join([
        f"• Entry range: ${m(lo)}–${m(hi)} (cur ${m(cur)})" if (lo is not None and hi is not None) else f"• Entry range: not provided (cur ${m(cur)})",
        f"• Time window: {tw}",
        f"• Levels: target ${m(tgt)} | stop ${m(stp)}",
        f"• P/L: +${m(pl_up)} ({m(up_pct)}%) | ${m(pl_dn)} ({m(dn_pct)}%) | R:R {m(rr)}" if rr is not None else "• P/L: not provided",
        f"• Stance reason: conf {r.get('confidence')} vs gate {CONF_GATE} | news {r.get('news_flag')}",
    ])


# -----------------------------
# FINAL_VIEW builder
# -----------------------------
def _vote_score(r: pd.Series) -> int:
    score = 0
    conf = pd.to_numeric(r.get("confidence"), errors="coerce")
    scr = pd.to_numeric(r.get("score"), errors="coerce")
    pct = pd.to_numeric(r.get("pct_change"), errors="coerce")
    trend = str(r.get("forecast_trend") or "").lower()
    news = str(r.get("news_flag") or "")

    if pd.notna(conf) and int(conf) >= 6:
        score += 3
    if pd.notna(scr) and float(scr) >= 60:
        score += 2
    if pd.notna(pct) and abs(float(pct)) >= 3.0:
        score += 1
    if trend == "up":
        score += 1

    if news == "🔴":
        score -= 2
    elif news == "🟡":
        score -= 1

    return int(score)


def _build_final_view(pass_df: pd.DataFrame, all_df: pd.DataFrame) -> pd.DataFrame:
    cols_min = [
        "symbol", "current", "pct_change",
        "predicted_price", "target_price", "stop_loss",
        "forecast_trend", "forecast_atr", "forecast_reason",
        "score", "score_label", "confidence", "decision",
        "news_flag", "main_news_title", "main_news_link",
        "reasons", "llm_insights",
        "candle_bias",
        "body_pct",
        "upper_wick_pct",
        "lower_wick_pct",
        "range_low",
        "range_high",
        "range_position_pct",
        "range_zone",
        "near_support",
        "near_resistance",
    ]

    def _ensure(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy() if df is not None else pd.DataFrame()
        for c in cols_min:
            if c not in df.columns:
                df[c] = pd.NA
        return df

    pass_df = _ensure(pass_df)
    all_df = _ensure(all_df)

    if not pass_df.empty:
        final = pass_df.head(FINAL_VIEW_TOP_N).copy()
        final["stance"] = "GO"
        final["stance_reason"] = final.apply(
            lambda r: f"conf {r.get('confidence')} vs gate {CONF_GATE} | news {r.get('news_flag')}",
            axis=1
        )
        return final.reset_index(drop=True)

    if all_df.empty:
        return pd.DataFrame(columns=cols_min + ["stance", "stance_reason", "vote_score", "plan_card"])

    tmp = all_df.dropna(subset=["symbol"]).copy()
    tmp["symbol"] = tmp["symbol"].astype(str).str.upper().str.strip()
    tmp = tmp.drop_duplicates(subset=["symbol"])
    tmp["vote_score"] = tmp.apply(_vote_score, axis=1)
    tmp = tmp.sort_values(by=["vote_score", "confidence", "score"], ascending=False)

    final = tmp.head(FINAL_VIEW_TOP_N).copy()
    final["stance"] = "WATCH"
    final["stance_reason"] = final.apply(
        lambda r: f"conf {r.get('confidence')} vs gate {CONF_GATE} | vote {r.get('vote_score')} | news {r.get('news_flag')}",
        axis=1
    )
    return final.reset_index(drop=True)


# -----------------------------
# Helpers
# -----------------------------
def map_score_to_decision(score: int) -> str:
    if score >= SCORE_HIGH:
        return "Strong Buy"
    if score >= SCORE_MEDIUM:
        return "Moderate"
    return "Not Advisable"


def extract_headline_from_html(news_html: str) -> str:
    if not news_html:
        return ""
    news_html = str(news_html).replace("\uFFFC", "").strip()
    m = re.search(r">(.*?)</a>", news_html)
    return (m.group(1).strip() if m else re.sub(r"<.*?>", "", news_html).strip())


def extract_url_from_html(news_html: str) -> str:
    if not news_html:
        return ""
    m = re.search(r'href="([^"]+)"', str(news_html))
    return m.group(1).strip() if m else ""


def news_flag_from_headlines(headlines: List[str]) -> str:
    if not headlines:
        return "🟡"
    score = 0
    for h in headlines:
        t = (h or "").lower()
        if any(w in t for w in POS_WORDS):
            score += 1
        if any(w in t for w in NEG_WORDS):
            score -= 1
    if score >= 1:
        return "🟢"
    if score <= -1:
        return "🔴"
    return "🟡"


def get_market_snapshot() -> dict:
    out = {"trend": "up", "spy_gap_pct": 0.0, "vix": None, "regime": "unknown"}
    try:
        spy = yf.Ticker("SPY").history(period="2d", auto_adjust=False)
        if not spy.empty and len(spy) >= 2:
            prev_close = float(spy["Close"].iloc[-2])
            last_close = float(spy["Close"].iloc[-1])
            out["trend"] = "up" if last_close > prev_close else "down"
            out["spy_gap_pct"] = ((last_close - prev_close) / prev_close) * 100.0
    except Exception:
        pass
    try:
        vix = yf.Ticker("^VIX").history(period="1d", auto_adjust=False)
        if not vix.empty:
            out["vix"] = float(vix["Close"].iloc[-1])
    except Exception:
        pass
    try:
        v = out.get("vix")
        t = out.get("trend")
        if v is None:
            out["regime"] = "unknown"
        elif v >= 25:
            out["regime"] = "risk_off"
        elif t == "up":
            out["regime"] = "risk_on"
        else:
            out["regime"] = "neutral"
    except Exception:
        pass

    return out

def get_volume_ratio(symbol: str, lookback_days: int = 20) -> Optional[float]:
    """
    volume_ratio = today's volume / avg(volume over last N days)
    Uses yfinance. Returns None if unavailable.
    """
    try:
        hist = yf.Ticker(symbol).history(period=f"{max(lookback_days, 5) + 5}d", auto_adjust=False)
        if hist is None or hist.empty or "Volume" not in hist.columns:
            return None

        vol = hist["Volume"].dropna()
        if len(vol) < 3:
            return None

        today_vol = float(vol.iloc[-1])
        # average of prior N days (exclude today)
        prev = vol.iloc[:-1].tail(lookback_days)
        if prev.empty:
            return None
        avg = float(prev.mean())
        if avg <= 0:
            return None
        return today_vol / avg
    except Exception:
        return None

def compute_confidence(score_val: int, pct_change: float, market_trend: str, news_flag: str) -> int:
    score_val = max(0, min(int(score_val), 100))
    base = round((score_val / 100.0) * 10.0)

    move = min(abs(float(pct_change or 0.0)), 8.0)
    vol_penalty = 2 if move >= 6 else (1 if move >= 4 else 0)

    market_bonus = 1 if market_trend == "up" else 0
    news_bonus = 1 if news_flag == "🟢" else (-1 if news_flag == "🔴" else 0)

    conf = int(base + market_bonus + news_bonus - vol_penalty)
    return max(1, min(conf, 10))

def _expected_rr(entry: float | None, target: float | None, stop: float | None) -> float | None:
    try:
        if entry is None or target is None or stop is None:
            return None
        entry = float(entry); target = float(target); stop = float(stop)
        if entry <= 0:
            return None
        reward = target - entry
        risk = entry - stop
        if risk <= 0:
            return None
        return reward / risk
    except Exception:
        return None


def _win_probability(
    *,
    score: int,
    conf: int,
    rr: float | None,
    trend: str,
    news_flag: str,
    market_regime: str,
) -> int:
    """
    Deterministic win probability estimate (0–100).
    Uses only already-known features (no ML, no randomness).
    """
    # base from confidence (conf 1..10)
    p = 35 + (int(conf) * 4)  # conf 5 => 55, conf 6 => 59, conf 7 => 63

    # score contribution (centered around 60)
    p += int((int(score) - 60) * 0.6)  # score 70 => +6, score 50 => -6

    # trend
    t = (trend or "").lower()
    if t == "up":
        p += 3
    elif t == "down":
        p -= 4

    # news flag
    if news_flag == "🟢":
        p += 2
    elif news_flag == "🔴":
        p -= 5

    # market regime
    if market_regime == "risk_on":
        p += 2
    elif market_regime == "risk_off":
        p -= 6

    # rr adjustment (prefer >= 1.2)
    if rr is not None:
        if rr >= 1.8:
            p += 4
        elif rr >= 1.4:
            p += 2
        elif rr < 1.0:
            p -= 4

    return int(max(1, min(99, round(p))))


def _default_target_stop(entry: float, conf: int) -> tuple[float, float]:
    conf = int(conf) if conf is not None else 5
    tgt_pct = 0.015 if conf >= 7 else (0.012 if conf >= 6 else 0.01)
    stp_pct = 0.012 if conf >= 7 else (0.015 if conf >= 6 else 0.02)
    return entry * (1.0 + tgt_pct), entry * (1.0 - stp_pct)


def _sanitize_forecast_levels(

    entry: float,
    conf: int,
    pred: Optional[float],
    tgt: Optional[float],
    stp: Optional[float],
) -> tuple[Optional[float], float, float]:
    if entry is None or entry <= 0:
        dt, ds = _default_target_stop(1.0, conf)
        return pred, float(tgt) if tgt is not None else dt, float(stp) if stp is not None else ds

    if tgt is not None and tgt <= entry:
        tgt = None
    if stp is not None and stp >= entry:
        stp = None

    if tgt is not None and tgt > entry * FORECAST_TARGET_MAX_MULT:
        tgt = None
    if stp is not None and stp < entry * FORECAST_STOP_MIN_MULT:
        stp = None

    if tgt is None or stp is None:
        dt, ds = _default_target_stop(entry, conf)
        if tgt is None:
            tgt = dt
        if stp is None:
            stp = ds

    return pred, float(tgt), float(stp)


def _time_scaled_target_stop(entry: float, base_target: float, base_stop: float, now: datetime) -> tuple[float, float]:
    if entry <= 0:
        return base_target, base_stop

    sess_end_dt = datetime.combine(now.date(), SESSION_END, tzinfo=LOCAL_TZ)
    sess_start_dt = datetime.combine(now.date(), SESSION_START, tzinfo=LOCAL_TZ)

    minutes_left = max(0.0, (sess_end_dt - now).total_seconds() / 60.0)
    total_minutes = max(1.0, (sess_end_dt - sess_start_dt).total_seconds() / 60.0)
    frac_left = min(1.0, minutes_left / total_minutes)

    tgt_pct = max(0.0, (base_target / entry) - 1.0)
    stp_pct = max(0.0, 1.0 - (base_stop / entry)) if base_stop and base_stop > 0 else 0.0

    tgt_pct_scaled = tgt_pct * frac_left
    stp_pct_scaled = stp_pct * frac_left

    tgt_pct_scaled = min(max(tgt_pct_scaled, 0.005), 0.04)
    stp_pct_scaled = min(max(stp_pct_scaled, 0.004), 0.03)

    new_target = entry * (1.0 + tgt_pct_scaled)
    new_stop = entry * (1.0 - stp_pct_scaled)
    return float(new_target), float(new_stop)


# -----------------------------
# Email builder (FINAL_VIEW)
# -----------------------------
def build_midday_alert(final_view_df: pd.DataFrame, run_date: str) -> str:
    if final_view_df is None or final_view_df.empty:
        return f"<h2>⚡ Midday Sudden Movers ({run_date})</h2><p>No qualified picks or watchlist today.</p>"

    df2 = final_view_df.copy()

    def _fmt_money(x):
        try:
            if pd.isna(x):
                return ""
            return f"{float(x):.2f}"
        except Exception:
            return ""

    def _fmt_pct(x):
        try:
            if pd.isna(x):
                return ""
            return f"{float(x):.2f}%"
        except Exception:
            return ""

    def row_html(r):
        sym = _html.escape(str(r.get("symbol", "")))
        cur = _fmt_money(r.get("current"))
        pct = _fmt_pct(r.get("pct_change"))
        pred = _fmt_money(r.get("predicted_price"))
        tgt = _fmt_money(r.get("target_price"))
        stp = _fmt_money(r.get("stop_loss"))

        score = _html.escape(str(r.get("score", "")))
        conf = _html.escape(str(r.get("confidence", "")))
        stance = _html.escape(str(r.get("stance", "")))
        stance_reason = _html.escape(str(r.get("stance_reason", "")))

        title = _html.escape(str(r.get("main_news_title") or ""))
        link = str(r.get("main_news_link") or "").strip() or "#"
        reasons = _html.escape(str(r.get("reasons") or ""))[:400]

        plan = _html.escape(str(r.get("plan_card") or ""))[:900].replace("\n", "<br>")

        ki = _html.escape(str(r.get("key_insight") or ""))[:500].replace("\n", "<br>")
        ins = _html.escape(str(r.get("llm_insights") or ""))[:600].replace("\n", "<br>")

        winp = _html.escape(str(r.get("win_prob", "")))
        rr = r.get("expected_rr")
        rr_txt = ""
        try:
            rr_txt = f"{float(rr):.2f}" if rr is not None and not pd.isna(rr) else ""
        except Exception:
            rr_txt = ""

        return f"""
        <tr>
          <td><b>{sym}</b></td>
          <td>{cur}</td>
          <td>{pct}</td>
          <td>{pred}</td>
          <td>{tgt}</td>
          <td>{stp}</td>
          <td>{score}</td>
          <td>{conf}</td>
          <td>{stance}<br><span style="color:#666;font-size:11px;">{stance_reason}</span></td>
          <td><a href="{link}" target="_blank">{title}</a></td>
          <td style="color:#444;">{reasons}</td>
          <td style="color:#222;white-space:normal;">{plan}</td>
          <td style="color:#555;white-space:normal;">{ki}</td>
          <td style="color:#555;white-space:normal;">{ins}</td>
          <td>{winp}</td>
          <td>{rr_txt}</td>
        </tr>
        """

    rows = "\n".join([row_html(r) for _, r in df2.iterrows()])
    mode_note = "Qualified Picks (GO)" if (
                "stance" in df2.columns and (df2["stance"] == "GO").any()) else "High Vote Watchlist (WATCH)"
    return f"""
    <h2>⚡ Midday Sudden Movers ({run_date})</h2>
    <p><b>Mode:</b> {mode_note}</p>
    <p>Filters: abs(move)≥{SUDDEN_MOVER_PCT_THRESHOLD}%, conf≥{CONF_GATE}, price≤${MAX_PRICE} (elite override allowed).</p>
    <p><b>Win rule (postmarket):</b> target hit anytime after recommendation (intraday high ≥ target).</p>

    <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;font-family:Arial;font-size:13px;">
      <tr style="background:#eee;">
        <th>Symbol</th><th>Price</th><th>%</th><th>Pred</th><th>Target</th><th>Stop</th>
        <th>Score</th><th>Conf</th><th>Stance</th><th>Headline</th><th>Reasons</th><th>Plan Card (Locked)</th><th>Key Insight</th><th>Insights</th><th>Win%</th><th>Exp R:R</th>
      </tr>
      {rows}
    </table>
    """


# -----------------------------
# Logs
# -----------------------------
RECO_LOG_CSV = out_path("recommendations_log.csv", kind="logs")
ensure_csv_exists(RECO_LOG_CSV, [
    "run_ts", "run_date", "mode", "app_env",
    "symbol", "current", "pct_change",
    "decision", "score", "score_label", "confidence",
    "predicted_price", "target_price", "stop_loss",
    "forecast_trend", "forecast_atr",
    "news_flag", "main_news_title", "main_news_link",
    "reasons",
])


def append_recommendations_log(df_reco: pd.DataFrame, now: datetime, mode: str) -> None:
    if df_reco is None or df_reco.empty:
        return

    out = df_reco.copy()

    # ✅ prevent insert collisions if columns already exist
    for c in ["run_ts", "run_date", "mode", "app_env"]:
        if c in out.columns:
            out = out.drop(columns=[c])

    out.insert(0, "run_ts", now.strftime("%Y-%m-%d %H:%M:%S"))
    out.insert(1, "run_date", now.strftime("%Y-%m-%d"))
    out.insert(2, "mode", mode)
    out.insert(3, "app_env", APP_ENV)

    cols_keep = [
        "run_ts", "run_date", "mode", "app_env",
        "symbol", "current", "pct_change",
        "decision", "score", "score_label", "confidence",
        "predicted_price", "target_price", "stop_loss",
        "forecast_trend", "forecast_atr",
        "news_flag", "main_news_title", "main_news_link",
        "reasons",
    ]

    for c in cols_keep:
        if c not in out.columns:
            out[c] = pd.NA

    for c in ["current", "pct_change", "predicted_price", "target_price", "stop_loss", "forecast_atr"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")

    out = out[cols_keep]
    file_exists = os.path.exists(RECO_LOG_CSV) and os.path.getsize(RECO_LOG_CSV) > 0
    out.to_csv(RECO_LOG_CSV, mode="a", header=not file_exists, index=False)


DAILY_LOG_CSV = out_path("daily_stock_log.csv", kind="logs")
ensure_csv_exists(DAILY_LOG_CSV, [
    "run_ts",
    "run_date", "mode", "symbol", "price_category",
    "current", "predicted_price", "target_price", "stop_loss",
    "forecast_trend", "forecast_atr", "forecast_reason",
    "trade_plan", "earnings_risk",
    "decision", "score", "score_label", "confidence",
    "reasons", "llm_insights", "news_flag", "main_news_title", "main_news_link",
    "stance", "stance_reason", "vote_score",
    "plan_card",
    "key_insight",
    "win_prob",
    "ml_win_prob_pct",
    "expected_rr",
    "candle_bias",
    "body_pct",
    "upper_wick_pct",
    "lower_wick_pct",
    "range_low",
    "range_high",
    "range_position_pct",
    "range_zone",
    "near_support",
    "near_resistance",
])

def _clean_text_cell(x) -> str:
    s = "" if x is None else str(x)
    s = s.replace("\uFFFC", "").strip()
    # remove trailing excel-ish refs: 'SHEET'!1:6, A12, 1:4, etc.
    s = re.sub(r"\s*'[^']+'!\d+:\d+\s*$", "", s).strip()
    s = re.sub(r"\s*\b[A-Z]{1,3}\d+\b\s*$", "", s).strip()
    s = re.sub(r"\s*\b\d+:\d+\b\s*$", "", s).strip()
    return s

def append_daily_log(final_view_df: pd.DataFrame, now: datetime, mode: str) -> None:
    """
    Source-of-truth log for postmarket evaluation.
    Writes FINAL_VIEW rows (GO + WATCH) with stable schema (one row per symbol per day+mode).
    """
    if final_view_df is None or final_view_df.empty:
        return

    run_date = now.strftime("%Y-%m-%d")

    cols = [
        "run_ts",
        "run_date", "mode", "symbol", "price_category",
        "current", "predicted_price", "target_price", "stop_loss",
        "forecast_trend", "forecast_atr", "forecast_reason",
        "trade_plan", "earnings_risk",
        "decision", "win_prob", "ml_win_prob_pct", "expected_rr", "score", "score_label", "confidence",
        "reasons", "llm_insights", "news_flag", "main_news_title", "main_news_link",
        "stance", "stance_reason", "vote_score",
        "plan_card", "key_insight",
        "candle_bias",
        "body_pct",
        "upper_wick_pct",
        "lower_wick_pct",
        "range_low",
        "range_high",
        "range_position_pct",
        "range_zone",
        "near_support",
        "near_resistance",
    ]

    out = final_view_df.copy()
    out["run_ts"] = now.strftime("%Y-%m-%d %H:%M:%S")
    out["run_date"] = run_date
    out["mode"] = mode

    for c in cols:
        if c not in out.columns:
            out[c] = pd.NA
    out = out[cols]

    for c in [
        "current", "predicted_price", "target_price", "stop_loss", "forecast_atr",
        "win_prob", "ml_win_prob_pct", "expected_rr",
        "body_pct", "upper_wick_pct", "lower_wick_pct",
        "range_low", "range_high", "range_position_pct",
    ]:
        out[c] = pd.to_numeric(out[c], errors="coerce")

    existing = (
        pd.read_csv(DAILY_LOG_CSV)
        if (os.path.exists(DAILY_LOG_CSV) and os.path.getsize(DAILY_LOG_CSV) > 0)
        else pd.DataFrame(columns=cols)
    )
    existing = existing.reindex(columns=cols)

    merged = pd.concat([existing, out], ignore_index=True)
    merged["symbol"] = merged["symbol"].astype(str).str.upper().str.strip()
    merged = merged.drop_duplicates(subset=["run_date", "mode", "symbol"], keep="last")
    merged.to_csv(DAILY_LOG_CSV, index=False)


def _insight_payload(r: pd.Series) -> dict:
    d = r.to_dict()

    conf = int(pd.to_numeric(d.get("confidence"), errors="coerce") or 0)
    gate = int(CONF_GATE)

    d["confidence"] = conf
    d["conf_gate"] = gate
    d["min_confidence_to_trade"] = gate
    d["is_tradeable"] = conf >= gate

    vs = d.get("vote_score")
    if vs is None or (isinstance(vs, float) and pd.isna(vs)):
        d["vote_score"] = 0
    else:
        d["vote_score"] = int(pd.to_numeric(vs, errors="coerce") or 0)

    # ensure stance exists (FINAL_VIEW has it; others may not)
    if not d.get("stance"):
        d["stance"] = "WATCH"

    return d

# -----------------------------
# Midday runner
# -----------------------------
def run_midday(now: datetime | None = None) -> None:
    now = now or datetime.now(LOCAL_TZ)

    if now.time() < SESSION_START and not IS_LOCAL:
        print("⛔ Midday runner skipped: market not open yet.")
        return
    if now.time() >= SESSION_END and not IS_LOCAL:
        print("⛔ Midday runner skipped: market already closed.")
        return

    mode = "midday"
    rf = run_dir(now, mode)
    run_date = now.strftime("%Y-%m-%d")

    marker = _midday_email_marker(now)
    if marker.exists():
        print("📩 Midday email already sent for this run_date — skipping resend.")
        return

    tickers = fetch_sp500_tickers()
    movers = calculate_top_movers(tickers, top_n=TOP_N)
    raw_df = pd.DataFrame(movers)

    excel_path = out_path(
        f"midday_{now.strftime('%Y%m%d')}_{make_run_id(now)}.xlsx",
        now=now, mode=mode, kind="runs"
    )

    if raw_df.empty or "pct_change" not in raw_df.columns:
        write_midday_excel(
            excel_path,
            final_view_df=pd.DataFrame(),
            pass_df=pd.DataFrame(),
            all_df=pd.DataFrame(),
            threshold_df=pd.DataFrame(),
            rf=Path(rf),
        )
        html = f"<h2>⚡ Midday Sudden Movers ({run_date})</h2><p>No movers returned.</p>"
        if send_email(f"⚡ Sudden Movers Alert ({run_date})", html, attachment_path=excel_path):
            marker.write_text("sent\n", encoding="utf-8")
        return

    raw_df["pct_change"] = pd.to_numeric(raw_df["pct_change"], errors="coerce").fillna(0.0)
    raw_df["current"] = pd.to_numeric(raw_df.get("current"), errors="coerce").fillna(10**9)

    thr_df = raw_df[raw_df["pct_change"].abs() >= SUDDEN_MOVER_PCT_THRESHOLD].copy()

    if MIDDAY_DEBUG:
        thr_df.to_csv(Path(rf) / "midday_after_threshold.csv", index=False)

    # ✅ Liquidity filter (runs AFTER threshold, BEFORE scoring loop)
    if not thr_df.empty:
        liq_cache: dict[str, Optional[float]] = {}
        keep_rows = []

        for _, r in thr_df.iterrows():
            sym = str(r.get("symbol", "")).upper().strip()
            price = float(r.get("current", 0.0) or 0.0)
            if price < MIN_PRICE:
                continue

            if sym not in liq_cache:
                liq_cache[sym] = get_avg_dollar_volume(sym, lookback_days=20)

            adv = liq_cache[sym]
            if adv is None or adv < MIN_AVG_DOLLAR_VOL:
                continue

            keep_rows.append(r)

        thr_df = pd.DataFrame(keep_rows)

    if thr_df.empty:
        write_midday_excel(
            excel_path,
            final_view_df=pd.DataFrame(),
            pass_df=pd.DataFrame(),
            all_df=pd.DataFrame(),
            threshold_df=thr_df,
            rf=Path(rf),
        )
        html = f"""
        <h2>⚡ Midday Sudden Movers ({run_date})</h2>
        <p>No movers passed threshold + liquidity filters (abs(move)≥{SUDDEN_MOVER_PCT_THRESHOLD}%, price≥${MIN_PRICE}, avg $vol≥${int(MIN_AVG_DOLLAR_VOL):,}).</p>
        <p><b>Attachment:</b> Excel included with sheets: FINAL_VIEW (empty), PASS (empty), ALL_CANDIDATES, AFTER_THRESHOLD.</p>
        """
        if send_email(f"⚡ Sudden Movers Alert ({run_date})", html, attachment_path=excel_path):
            marker.write_text("sent\n", encoding="utf-8")
        return

    snapshot = get_market_snapshot()
    market_trend = snapshot.get("trend", "up")
    market_regime = snapshot.get("regime", "unknown")

    scores, labels, reasons_list, confs, decisions = [], [], [], [], []
    titles, links, flags = [], [], []
    preds, tgts, stps, ftrends, fatrs, freasons = [], [], [], [], [], []
    dbg_rows = []

    vol_ratios = []
    vol_cache: dict[str, Optional[float]] = {}

    candle_biases = []
    body_pcts = []
    upper_wick_pcts = []
    lower_wick_pcts = []
    range_lows = []
    range_highs = []
    range_pos_pcts = []
    range_zones = []
    near_supports = []
    near_resistances = []

    for _, row in thr_df.iterrows():
        sym = str(row.get("symbol", "")).upper().strip()
        current = float(row.get("current", 0.0) or 0.0)
        pct_change = float(row.get("pct_change", 0.0) or 0.0)

        pa = compute_price_action_summary(sym, current)

        candle_biases.append(pa.get("candle_bias"))
        body_pcts.append(pa.get("body_pct"))
        upper_wick_pcts.append(pa.get("upper_wick_pct"))
        lower_wick_pcts.append(pa.get("lower_wick_pct"))
        range_lows.append(pa.get("range_low"))
        range_highs.append(pa.get("range_high"))
        range_pos_pcts.append(pa.get("range_position_pct"))
        range_zones.append(pa.get("range_zone"))
        near_supports.append(pa.get("near_support"))
        near_resistances.append(pa.get("near_resistance"))

        # --- Volume confirmation (optional) ---
        vol_ratio = None
        if USE_VOLUME_FILTER:
            if sym in vol_cache:
                vol_ratio = vol_cache[sym]
            else:
                vol_ratio = get_volume_ratio(sym, lookback_days=VOL_LOOKBACK_DAYS)
                vol_cache[sym] = vol_ratio

        score_val, score_label, reasons = get_predictive_score_with_reasons(sym)
        score_val = int(score_val)
        decision = map_score_to_decision(score_val)

        news_items = fetch_news_links(sym, max_articles=1)
        main_item = news_items[0] if news_items else ""
        title = extract_headline_from_html(main_item)
        link = extract_url_from_html(main_item)
        flag = news_flag_from_headlines([title])

        conf = compute_confidence(score_val, pct_change, market_trend, flag)

        if USE_VOLUME_FILTER and vol_ratio is not None:
            if vol_ratio >= VOL_RATIO_MIN:
                conf = min(10, int(conf) + VOL_CONF_BONUS)
            else:
                conf = max(1, int(conf) - VOL_CONF_PENALTY)

        f = forecast_price_levels(sym, current=current, score=score_val, horizon="intraday")
        raw_pred = _safe_float(getattr(f, "predicted_price", None))
        raw_tgt = _safe_float(getattr(f, "target_price", None))
        raw_stp = _safe_float(getattr(f, "stop_loss", None))

        pred_s, tgt_s, stp_s = _sanitize_forecast_levels(current, int(conf), raw_pred, raw_tgt, raw_stp)
        # --- Volatility-aware base target/stop (Step 1) ---
        # Use ATR if available; otherwise keep sanitized levels.
        atr_val = _safe_float(getattr(f, "atr", None))
        base_row = pd.Series({
            "current": current,
            "forecast_atr": atr_val,
            "predicted_price": pred_s,
            "target_price": tgt_s,
            "stop_loss": stp_s,
        })
        vtgt, vstp = _volatility_target_stop(base_row)
        tgt_s = vtgt if vtgt is not None else tgt_s
        stp_s = vstp if vstp is not None else stp_s
        scores.append(score_val)
        labels.append(score_label)
        reasons_list.append(reasons)
        decisions.append(decision)
        confs.append(int(conf))
        vol_ratios.append(vol_ratio if vol_ratio is not None else pd.NA)
        titles.append(title)
        links.append(link)
        flags.append(flag)

        preds.append(pred_s if pred_s is not None else pd.NA)
        tgts.append(tgt_s if tgt_s is not None else pd.NA)
        stps.append(stp_s if stp_s is not None else pd.NA)
        ftrends.append(str(getattr(f, "trend", "")))
        fatrs.append(float(getattr(f, "atr", pd.NA)) if getattr(f, "atr", None) is not None else pd.NA)
        freasons.append(str(getattr(f, "reason", "")))

        if MIDDAY_DEBUG:
            dbg_rows.append({
                "symbol": sym,
                "current": current,
                "pct_change": pct_change,
                "confidence": int(conf),
                "raw_target": raw_tgt,
                "raw_stop": raw_stp,
                "final_target": tgt_s,
                "final_stop": stp_s,
            })

    all_df = thr_df.copy()
    all_df["score"] = scores
    all_df["score_label"] = labels
    all_df["reasons"] = reasons_list
    all_df["decision"] = decisions
    all_df["confidence"] = confs
    all_df["main_news_title"] = titles
    all_df["main_news_link"] = links
    all_df["news_flag"] = flags
    all_df["volume_ratio"] = pd.to_numeric(pd.Series(vol_ratios, index=all_df.index), errors="coerce")

    all_df["predicted_price"] = pd.to_numeric(pd.Series(preds, index=all_df.index), errors="coerce")
    all_df["target_price"] = pd.to_numeric(pd.Series(tgts, index=all_df.index), errors="coerce")
    all_df["stop_loss"] = pd.to_numeric(pd.Series(stps, index=all_df.index), errors="coerce")
    all_df["forecast_trend"] = ftrends
    all_df["forecast_atr"] = fatrs
    all_df["forecast_reason"] = freasons
    all_df["candle_bias"] = candle_biases
    all_df["body_pct"] = pd.to_numeric(pd.Series(body_pcts, index=all_df.index), errors="coerce")
    all_df["upper_wick_pct"] = pd.to_numeric(pd.Series(upper_wick_pcts, index=all_df.index), errors="coerce")
    all_df["lower_wick_pct"] = pd.to_numeric(pd.Series(lower_wick_pcts, index=all_df.index), errors="coerce")
    all_df["range_low"] = pd.to_numeric(pd.Series(range_lows, index=all_df.index), errors="coerce")
    all_df["range_high"] = pd.to_numeric(pd.Series(range_highs, index=all_df.index), errors="coerce")
    all_df["range_position_pct"] = pd.to_numeric(pd.Series(range_pos_pcts, index=all_df.index), errors="coerce")
    all_df["range_zone"] = range_zones
    all_df["near_support"] = near_supports
    all_df["near_resistance"] = near_resistances

    pass_df = all_df[all_df["confidence"] >= CONF_GATE].copy()
    # -----------------------------
    # Market regime filter
    # Avoid trades in risk-off markets unless strong signal
    # -----------------------------
    if market_regime == "risk_off":
        pass_df = pass_df[pass_df["confidence"] >= (CONF_GATE + 1)]
    elif market_regime == "neutral":
        pass
    elif market_regime == "risk_on":
        pass_df = pass_df[pass_df["confidence"] >= CONF_GATE]

    # debug visibility
    if MIDDAY_DEBUG:
        print(
            f"📊 Market regime: {market_regime} | "
            f"filtered_pass_count={len(pass_df)}"
        )
    pass_df = pass_df[
        (pass_df["current"] <= MAX_PRICE) |
        ((pass_df["score"] >= ELITE_SCORE_OVERRIDE) & (pass_df["confidence"] >= ELITE_CONF_OVERRIDE))
    ].copy()
    pass_df = pass_df.sort_values(by=["confidence", "score"], ascending=False)

    # Midday target/stop scaling
    if not pass_df.empty:
        tgt_scaled, stp_scaled = [], []
        for _, r in pass_df.iterrows():
            entry = _safe_float(r.get("current"))
            base_tgt = _safe_float(r.get("target_price"))
            base_stp = _safe_float(r.get("stop_loss"))

            if entry is not None and entry > 0 and base_tgt is not None and base_stp is not None:
                new_tgt, new_stp = _time_scaled_target_stop(entry, base_tgt, base_stp, now)
                tgt_scaled.append(new_tgt)
                stp_scaled.append(new_stp)
            else:
                tgt_scaled.append(base_tgt if base_tgt is not None else pd.NA)
                stp_scaled.append(base_stp if base_stp is not None else pd.NA)

        pass_df["target_price"] = pd.to_numeric(pd.Series(tgt_scaled, index=pass_df.index), errors="coerce")
        pass_df["stop_loss"] = pd.to_numeric(pd.Series(stp_scaled, index=pass_df.index), errors="coerce")

    # LLM explanations (optional; keep for Excel only)
    pass_df = _apply_llm_explanations(pass_df, horizon="midday_pass", top_n=MIDDAY_LLM_TOP_N)
    all_df = _apply_llm_explanations(all_df, horizon="midday_all_candidates", top_n=MIDDAY_LLM_TOP_N)

    # FINAL_VIEW
    final_view_df = _build_final_view(pass_df, all_df)

    # 🔒 Messaging guard: never show buy/preferable if not tradeable by gate
    for df_ in [final_view_df, pass_df, all_df]:
        if df_ is None or df_.empty:
            continue

        for c in ["mover_signal", "day_trade", "week_trade", "month_trade"]:
            if c not in df_.columns:
                df_[c] = ""

        def _guard_row(r: pd.Series) -> pd.Series:
            conf = int(pd.to_numeric(r.get("confidence"), errors="coerce") or 0)
            decision = str(r.get("decision") or "").strip()
            tradeable = conf >= int(CONF_GATE)

            if (not tradeable) or (decision == "Not Advisable"):
                r["mover_signal"] = "❌ AVOID / WATCH"
                r["day_trade"] = "WATCH"
                r["week_trade"] = "WATCH"
                r["month_trade"] = "WATCH"
            return r

        df_.loc[:, :] = df_.apply(_guard_row, axis=1)


    # FINAL_VIEW: deterministic plan_card always present
    if not final_view_df.empty:
        final_view_df["plan_card"] = final_view_df.apply(lambda r: _plan_card_row(r), axis=1)
    else:
        final_view_df = final_view_df.copy()
        final_view_df["plan_card"] = ""

    # --- Win Probability + Expected R:R (deterministic) ---
    final_view_df["expected_rr"] = final_view_df.apply(
        lambda r: _expected_rr(
            _safe_float(r.get("current")),
            _safe_float(r.get("target_price")),
            _safe_float(r.get("stop_loss")),
        ),
        axis=1,
    )

    final_view_df["win_prob"] = final_view_df.apply(
        lambda r: _win_probability(
            score=int(pd.to_numeric(r.get("score"), errors="coerce") or 0),
            conf=int(pd.to_numeric(r.get("confidence"), errors="coerce") or 0),
            rr=_safe_float(r.get("expected_rr")),
            trend=str(r.get("forecast_trend") or ""),
            news_flag=str(r.get("news_flag") or ""),
            market_regime=str(market_regime or "unknown"),
        ),
        axis=1,
    )

    final_view_df["ml_win_prob"] = final_view_df.apply(
        lambda r: predict_win_prob({
            "score": r.get("score"),
            "confidence": r.get("confidence"),
            "current": r.get("current"),
            "target_price": r.get("target_price"),
            "stop_loss": r.get("stop_loss"),
            "decision": r.get("decision"),
            "source_mode": "midday",
            "instrument_type": "stock",
        }),
        axis=1,
    )

    final_view_df["ml_win_prob_pct"] = final_view_df["ml_win_prob"].apply(
        lambda x: round(float(x) * 100, 2) if x is not None and pd.notna(x) else pd.NA
    )

    # Optional: LLM for FINAL_VIEW only (kept but not required)

    # final_view_df = _apply_llm_explanations(
    #     final_view_df,
    #     horizon="midday_final_view",
    #     top_n=None,
    # )

    # Key insight (deterministic)
    for df_ in [final_view_df, pass_df, all_df]:
        if df_ is None or df_.empty:
            continue
        if "key_insight" not in df_.columns:
            df_["key_insight"] = ""
        df_["key_insight"] = df_.apply(lambda r: build_key_insight(_insight_payload(r)), axis=1)


    # scrub trailing artifacts AFTER everything is created
    for df_ in [final_view_df, pass_df, all_df]:
        if df_ is None or df_.empty:
            continue
        for c in ["plan_card", "llm_insights", "key_insight"]:
            if c in df_.columns:
                df_[c] = df_[c].apply(_clean_text_cell)

    if MIDDAY_DEBUG and dbg_rows:
        pd.DataFrame(dbg_rows).to_csv(Path(rf) / "debug_forecast_sanity.csv", index=False)

    final_view_df["expected_rr"] = pd.to_numeric(final_view_df["expected_rr"], errors="coerce")
    final_view_df["win_prob"] = pd.to_numeric(final_view_df["win_prob"], errors="coerce")

    # FINAL_VIEW must be deterministic & never contradict stance
    if "llm_insights" in final_view_df.columns:
        final_view_df["llm_insights"] = ""

    # Logs:
    # - recommendations_log: PASS only (legacy)
    pass_df_to_log = pass_df.copy()
    if "llm_insights" in pass_df_to_log.columns:
        pass_df_to_log = pass_df_to_log.drop(columns=["llm_insights"])
    append_recommendations_log(pass_df_to_log, now, mode="midday")

    # - daily_stock_log: FINAL_VIEW (truth for postmarket)
    append_daily_log(final_view_df, now, mode="midday")

    # Excel
    write_midday_excel(
        excel_path,
        final_view_df=final_view_df,
        pass_df=pass_df,
        all_df=all_df,
        threshold_df=thr_df,
        rf=Path(rf),
    )

    # Email always from FINAL_VIEW (deterministic plan_card)
    if final_view_df.empty:
        html = f"""
        <h2>⚡ Midday Sudden Movers ({run_date})</h2>
        <p>No qualified picks or watchlist today.</p>
        <p><b>Attachment:</b> Excel included with sheets: FINAL_VIEW, PASS, ALL_CANDIDATES, AFTER_THRESHOLD.</p>
        """
        if send_email(f"⚡ Sudden Movers Alert ({run_date})", html, attachment_path=excel_path):
            marker.write_text("sent\n", encoding="utf-8")
        return

    html = build_midday_alert(final_view_df, run_date)
    if send_email(f"⚡ Sudden Movers Alert ({run_date})", html, attachment_path=excel_path):
        marker.write_text("sent\n", encoding="utf-8")

    print(f"✅ Midday complete | final_view={len(final_view_df)} | pass={len(pass_df)} | threshold_rows={len(thr_df)}")


if __name__ == "__main__":
    run_midday()