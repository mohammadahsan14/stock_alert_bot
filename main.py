# main.py
import os
import re
import argparse
import html as _html
from pathlib import Path
from typing import List, Tuple
from datetime import datetime, time, timedelta
from zoneinfo import ZoneInfo

from dotenv import load_dotenv

load_dotenv(dotenv_path=Path(__file__).with_name(".env"))

if os.getenv("DEBUG") == "1":
    print("RESEND_API_KEY loaded:", bool(os.getenv("RESEND_API_KEY")))

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from forecast_engine import forecast_price_levels
from email_sender import send_email as resend_send_email
from top_movers import fetch_sp500_tickers, calculate_top_movers

# ✅ richer scoring API (needs to exist in scoring_engine.py)
from scoring_engine import get_predictive_score_with_reasons

from news_fetcher import fetch_news_links
from price_category import get_price_category
from config import (
    SENDER_EMAIL, RECEIVER_EMAIL,
    TOP_N, SCORE_COLORS, SCORE_HIGH, SCORE_MEDIUM,
    EXPECTED_UPSIDE_HIGH, EXPECTED_UPSIDE_MEDIUM, EXPECTED_DOWN,  # kept for compatibility
)

# ✅ Phase 2: portfolio/performance tracking
from performance_tracker import (
    PortfolioConfig,
    load_open_portfolio,
    save_open_portfolio,
    append_trade_history,
    add_new_positions_from_picks,
    update_and_close_positions,
    portfolio_summary,
)

# -----------------------------
# Output folder
# -----------------------------
OUTPUT_DIR = "outputs"
os.makedirs(OUTPUT_DIR, exist_ok=True)

def out_path(filename: str) -> str:
    """Return a safe file path inside outputs/."""
    return os.path.join(OUTPUT_DIR, filename)

def log_outputs_folder():
    try:
        p = os.path.abspath(OUTPUT_DIR)
        print(f"📁 outputs path: {p}")
        if os.path.exists(p):
            print("📄 outputs files:", os.listdir(p))
        else:
            print("⚠️ outputs folder not found")
    except Exception as e:
        print("⚠️ log_outputs_folder failed:", e)

# Logs inside outputs/
DAILY_LOG_CSV = out_path("daily_stock_log.csv")   # stores ONLY picks sent
PERF_LOG_CSV = out_path("performance_log.csv")    # appended daily after post-market

EXCEL_MAX_ROWS = 10         # SUPPORTING_DATA rows
TRADE_MAX_PICKS = 3         # EMAIL + DAILY_PICKS rows

SUDDEN_MOVER_PCT_THRESHOLD = 3.0

# ✅ Always run in Chicago time (CST/CDT handled by TZ database)
LOCAL_TZ = ZoneInfo("America/Chicago")
POST_MARKET_START = time(15, 10)  # 3:10 PM Chicago (market closes 3:00 PM Chicago)

# -----------------------------
# Phase 2: Market-wide gates (better than "max mover")
# -----------------------------
VIX_SKIP_THRESHOLD = 25.0
SPY_GAP_DOWN_SKIP_PCT = -1.25
SPY_GAP_DOWN_TIGHTEN_PCT = -0.60

# -----------------------------
# Reliability gates (NO TRADE DAY logic)
# -----------------------------
MIN_STRONG_BUY_PICKS = 1
MIN_CONFIDENCE_TO_TRADE = 6

# IMPORTANT: use percentile (p90) of movers instead of max
MAX_ALLOWED_VOLATILITY_P90 = 6.0

MARKET_DOWNSHIFT_BLOCK = True

# -----------------------------
# Earnings filter (gap-risk protection)
# -----------------------------
SKIP_EARNINGS_STOCKS = True
EARNINGS_LOOKAHEAD_DAYS = 3

# Price buckets (used for email grouping only)
PRICE_BUCKETS = [
    ("Ultra Penny ($)", "Ultra Penny Stocks"),
    ("Penny ($)", "Penny Stocks"),
    ("Mid ($$)", "Mid Price Stocks"),
    ("Mid-High ($$$)", "Mid-High Stocks"),
    ("High ($$$$)", "High Price Stocks"),
    ("Unknown", "Unknown"),
]

# -----------------------------
# Map score to decision
# -----------------------------
def map_score_to_decision(score: int) -> str:
    if score >= SCORE_HIGH:
        return "Strong Buy"
    elif score >= SCORE_MEDIUM:
        return "Moderate"
    return "Not Advisable"

# -----------------------------
# Normalize hex color for openpyxl
# -----------------------------
def normalize_color(color: str) -> str:
    if not color:
        color = "#FFFFFF"
    color = color.lstrip("#")
    if len(color) == 6:
        color = "FF" + color
    return color.upper()

# -----------------------------
# Extract headline / url from HTML link string
# -----------------------------
def extract_headline_from_html(news_html: str) -> str:
    if not news_html:
        return ""
    news_html = str(news_html).replace("\uFFFC", "").strip()
    m = re.search(r'>(.*?)</a>', news_html)
    return (m.group(1).strip() if m else re.sub(r"<.*?>", "", news_html).strip())

def extract_url_from_html(news_html: str) -> str:
    if not news_html:
        return ""
    m = re.search(r'href="([^"]+)"', str(news_html))
    return m.group(1).strip() if m else ""

# -----------------------------
# Simple sentiment -> flag
# -----------------------------
POS_WORDS = {"beat", "strong", "growth", "surge", "upgrade", "raises", "record", "profit", "wins", "bull"}
NEG_WORDS = {"miss", "drop", "loss", "cuts", "downgrade", "falls", "weak", "lawsuit", "plunge", "bear"}

def news_flag_from_headlines(headlines: List[str]) -> str:
    if not headlines:
        return "🟡"
    score = 0
    for h in headlines:
        t = (h or "").lower()
        if any(w in t for w in POS_WORDS):
            score += 1
        if any(w in t for w in NEG_WORDS):
            score -= 1
    if score >= 1:
        return "🟢"
    if score <= -1:
        return "🔴"
    return "🟡"

# -----------------------------
# Market snapshot + direction (SPY + VIX)
# -----------------------------
def get_market_snapshot() -> dict:
    """
    Returns:
      {
        "trend": "up"|"down",
        "spy_gap_pct": float,
        "vix": float|None
      }
    """
    out = {"trend": "up", "spy_gap_pct": 0.0, "vix": None}
    try:
        spy = yf.Ticker("SPY").history(period="2d")
        if not spy.empty and len(spy) >= 2:
            prev_close = float(spy["Close"].iloc[-2])
            last_close = float(spy["Close"].iloc[-1])
            out["trend"] = "up" if last_close > prev_close else "down"
            out["spy_gap_pct"] = ((last_close - prev_close) / prev_close) * 100.0
    except Exception:
        pass

    try:
        vix = yf.Ticker("^VIX").history(period="1d")
        if not vix.empty:
            out["vix"] = float(vix["Close"].iloc[-1])
    except Exception:
        pass

    return out

def get_market_direction() -> str:
    return get_market_snapshot().get("trend", "up")

# -----------------------------
# Confidence model (gentler on big movers)
# -----------------------------
def compute_confidence(score_val: int, pct_change: float, market_trend: str, news_flag: str) -> int:
    base = score_val / 10.0  # 0..10
    pct_for_conf = min(abs(pct_change), 5.0)
    vol_adj = max(0.75, 1 - pct_for_conf / 12.0)
    market_adj = 1.05 if market_trend == "up" else 0.95
    news_adj = 1.05 if news_flag == "🟢" else (0.95 if news_flag == "🔴" else 1.0)
    conf = int(round(base * vol_adj * market_adj * news_adj))
    return min(max(conf, 1), 10)

# -----------------------------
# Earnings soon? (gap-risk)
# -----------------------------
def has_earnings_soon(symbol: str, now: datetime, lookahead_days: int = 3) -> bool:
    try:
        t = yf.Ticker(symbol)
        cal = t.calendar
        if cal is None or cal.empty:
            return False

        dt = None
        if "Earnings Date" in cal.index:
            vals = cal.loc["Earnings Date"].values
            if len(vals) > 0:
                dt = vals[0]

        if dt is None:
            for v in cal.values.flatten().tolist():
                if hasattr(v, "to_pydatetime"):
                    dt = v.to_pydatetime()
                    break

        if dt is None or not hasattr(dt, "date"):
            return False

        start = now.date()
        end = (now + timedelta(days=lookahead_days)).date()
        return start <= dt.date() <= end
    except Exception:
        return False

# -----------------------------
# Trade Plan column
# -----------------------------
def assign_trade_plan(risk: str, pct_change: float, market_trend: str, score_val: int) -> str:
    if abs(pct_change) >= 3.0:
        return "Intraday"
    if score_val >= SCORE_HIGH and market_trend == "up" and risk in ("Low", "Medium"):
        return "Swing (3–10 days)"
    return "Intraday"

# -----------------------------
# NO TRADE DAY logic (improved)
# -----------------------------
def should_skip_day(df: pd.DataFrame, market_trend: str, snapshot: dict) -> Tuple[bool, str]:
    if df is None or df.empty:
        return True, "Empty dataset"

    if "current" not in df.columns:
        return True, "Missing prices"

    if df["current"].isna().any() or (df["current"] <= 0).any():
        return True, "Missing/invalid prices"

    vix = snapshot.get("vix")
    if vix is not None and vix >= VIX_SKIP_THRESHOLD:
        return True, f"Risk-off day: VIX too high ({vix:.2f})"

    spy_gap = float(snapshot.get("spy_gap_pct") or 0.0)
    if spy_gap <= SPY_GAP_DOWN_SKIP_PCT:
        return True, f"Risk-off day: SPY weak ({spy_gap:.2f}%)"

    if "pct_change" in df.columns:
        p90 = df["pct_change"].abs().quantile(0.90)
        if pd.notna(p90) and p90 >= MAX_ALLOWED_VOLATILITY_P90:
            if market_trend == "down" or spy_gap <= SPY_GAP_DOWN_TIGHTEN_PCT:
                return True, f"Too volatile (p90 mover {p90:.2f}%) with weak market bias"

    if "decision" in df.columns and "confidence" in df.columns:
        tradeable = df[(df["decision"] == "Strong Buy") & (df["confidence"] >= MIN_CONFIDENCE_TO_TRADE)]
        if len(tradeable) < MIN_STRONG_BUY_PICKS:
            return True, "No Strong Buy picks meeting confidence threshold"

    if MARKET_DOWNSHIFT_BLOCK and market_trend == "down":
        tradeable = df[(df["decision"] == "Strong Buy") & (df["confidence"] >= MIN_CONFIDENCE_TO_TRADE)]
        if len(tradeable) < 2:
            return True, "Market trend down + not enough high-confidence picks"

    return False, ""

# -----------------------------
# Excel styling
# -----------------------------
def style_excel_sheet(sheet):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color=normalize_color("#2F5597"),
        end_color=normalize_color("#2F5597"),
        fill_type="solid",
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    sheet.freeze_panes = "A2"

    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = min(max_length + 2, 55)

    headers = {str(c.value).strip(): idx + 1 for idx, c in enumerate(sheet[1]) if c.value}
    score_label_col = headers.get("score_label")
    decision_col = headers.get("decision")

    for r in range(2, sheet.max_row + 1):
        try:
            if score_label_col:
                cell = sheet.cell(row=r, column=score_label_col)
                label = str(cell.value or "")
                color = SCORE_COLORS.get(label, "#FFFFFF")
                cell.fill = PatternFill(
                    start_color=normalize_color(color),
                    end_color=normalize_color(color),
                    fill_type="solid",
                )
                cell.alignment = center

            if decision_col:
                dcell = sheet.cell(row=r, column=decision_col)
                dval = str(dcell.value or "")
                if dval == "Strong Buy":
                    c = "#92D050"
                elif dval == "Moderate":
                    c = "#FFF2CC"
                else:
                    c = "#F4CCCC"
                dcell.fill = PatternFill(
                    start_color=normalize_color(c),
                    end_color=normalize_color(c),
                    fill_type="solid",
                )
                dcell.alignment = center
        except Exception:
            pass

# -----------------------------
# Email HTML (top picks; includes predicted price + plan)
# -----------------------------
def build_email_html_top_picks(df: pd.DataFrame, run_date: str) -> str:
    if df.empty:
        return f"<h2>📊 Daily Picks – Pre Market ({run_date})</h2><p>No picks available today.</p>"

    html = f"<h2>📊 Daily Picks – Pre Market ({run_date})</h2>"
    html += "<p>Ranked by <b>confidence</b> then <b>score</b>. Full details are in Excel.</p>"

    picks = df.sort_values(by=["confidence", "score"], ascending=False).head(TRADE_MAX_PICKS).copy()

    for cat_code, cat_name in PRICE_BUCKETS:
        cat_df = picks[picks["price_category"] == cat_code]
        if cat_df.empty:
            continue

        html += f"<h3>{_html.escape(cat_name)}</h3>"
        html += "<table style='border-collapse:collapse;font-family:Arial;width:100%;margin-bottom:18px;'>"
        html += """
        <tr style='background:#f2f2f2;'>
          <th style='padding:6px;border:1px solid #ddd;'>Symbol</th>
          <th style='padding:6px;border:1px solid #ddd;'>Price</th>
          <th style='padding:6px;border:1px solid #ddd;'>Predicted</th>
          <th style='padding:6px;border:1px solid #ddd;'>Target</th>
          <th style='padding:6px;border:1px solid #ddd;'>Stop</th>
          <th style='padding:6px;border:1px solid #ddd;'>Plan</th>
          <th style='padding:6px;border:1px solid #ddd;'>Decision</th>
          <th style='padding:6px;border:1px solid #ddd;'>Score</th>
          <th style='padding:6px;border:1px solid #ddd;'>Conf</th>
          <th style='padding:6px;border:1px solid #ddd;'>News</th>
          <th style='padding:6px;border:1px solid #ddd;'>Why</th>
        </tr>
        """
        for _, row in cat_df.iterrows():
            label = str(row.get("score_label", "") or "")
            score_color = SCORE_COLORS.get(label, "#FFFFFF")

            why = _html.escape(str(row.get("reasons", "") or ""))
            if len(why) > 240:
                why = why[:240] + "..."

            html += f"""
            <tr>
              <td style='padding:6px;border:1px solid #ddd;'>{_html.escape(str(row.get('symbol','')))}</td>
              <td style='padding:6px;border:1px solid #ddd;'>{float(row.get('current',0) or 0):.2f}</td>
              <td style='padding:6px;border:1px solid #ddd;'>{float(row.get('predicted_price',0) or 0):.2f}</td>
              <td style='padding:6px;border:1px solid #ddd;'>{float(row.get('target_price',0) or 0):.2f}</td>
              <td style='padding:6px;border:1px solid #ddd;'>{float(row.get('stop_loss',0) or 0):.2f}</td>
              <td style='padding:6px;border:1px solid #ddd;text-align:center;'>{_html.escape(str(row.get('trade_plan','') or ''))}</td>
              <td style='padding:6px;border:1px solid #ddd;text-align:center;'>{_html.escape(str(row.get('decision','') or ''))}</td>
              <td style='padding:6px;border:1px solid #ddd;background:{score_color};text-align:center;'>{_html.escape(label)} ({int(row.get('score',0) or 0)})</td>
              <td style='padding:6px;border:1px solid #ddd;text-align:center;'>{int(row.get('confidence',0) or 0)}</td>
              <td style='padding:6px;border:1px solid #ddd;text-align:center;font-size:16px;'>{_html.escape(str(row.get('news_flag','🟡') or '🟡'))}</td>
              <td style='padding:6px;border:1px solid #ddd;'>{why}</td>
            </tr>
            """
        html += "</table>"

    return html

# -----------------------------
# Send email (Resend wrapper)
# -----------------------------
def send_email(subject: str, html_body: str, attachment_path: str | None = None) -> bool:
    return resend_send_email(
        subject=subject,
        html_body=html_body,
        to_email=RECEIVER_EMAIL,
        from_email=SENDER_EMAIL,
        attachment_path=attachment_path,
    )

# -----------------------------
# Post-market evaluation (reads picks log)
# -----------------------------
def evaluate_post_market_from_log(log_csv: str) -> pd.DataFrame:
    try:
        df = pd.read_csv(log_csv)
    except Exception:
        return pd.DataFrame()

    if df.empty or "symbol" not in df.columns:
        return df

    close_prices = []
    for symbol in df["symbol"].tolist():
        try:
            data = yf.Ticker(symbol).history(period="1d")
            close = float(data["Close"].iloc[-1]) if not data.empty else float("nan")
        except Exception:
            close = float("nan")
        close_prices.append(close)

    df["close_price"] = close_prices
    df["current"] = pd.to_numeric(df.get("current", pd.Series([float("nan")] * len(df))), errors="coerce")
    df["actual_change_pct"] = (df["close_price"] - df["current"]) / df["current"] * 100

    outcomes = []
    for _, row in df.iterrows():
        predicted = row.get("decision", "Not Advisable")
        actual = row.get("actual_change_pct", 0)

        if predicted in ["Strong Buy", "Moderate"] and actual > 0:
            outcomes.append("✅ Correct")
        elif predicted == "Not Advisable" and actual <= 0:
            outcomes.append("✅ Correct")
        else:
            outcomes.append("❌ Incorrect")

    df["outcome"] = outcomes
    return df

# -----------------------------
# Midday alert (sudden movers)
# -----------------------------
def build_midday_alert(df: pd.DataFrame, run_date: str) -> str:
    if df.empty:
        return ""

    html = f"<h2>⚡ Sudden Movers Alert ({run_date})</h2>"
    html += "<table style='border-collapse:collapse;font-family:Arial;width:100%;'>"
    html += """
    <tr style='background:#f2f2f2;'>
      <th style='padding:6px;border:1px solid #ddd;'>Symbol</th>
      <th style='padding:6px;border:1px solid #ddd;'>% Change</th>
      <th style='padding:6px;border:1px solid #ddd;'>Decision</th>
      <th style='padding:6px;border:1px solid #ddd;'>Score</th>
      <th style='padding:6px;border:1px solid #ddd;'>Conf</th>
      <th style='padding:6px;border:1px solid #ddd;'>Why</th>
      <th style='padding:6px;border:1px solid #ddd;'>Main News</th>
    </tr>
    """

    for _, row in df.iterrows():
        link = str(row.get("main_news_link", "") or "")
        title = _html.escape(str(row.get("main_news_title", "") or "—"))
        link_html = f'<a href="{_html.escape(link)}" target="_blank">{title}</a>' if link else title

        why = _html.escape(str(row.get("reasons", "") or ""))
        if len(why) > 200:
            why = why[:200] + "..."

        html += f"""
        <tr>
          <td style='padding:6px;border:1px solid #ddd;'>{_html.escape(str(row.get('symbol','')))}</td>
          <td style='padding:6px;border:1px solid #ddd;text-align:center;'>{float(row.get('pct_change',0) or 0):.2f}%</td>
          <td style='padding:6px;border:1px solid #ddd;text-align:center;'>{_html.escape(str(row.get('decision','') or ''))}</td>
          <td style='padding:6px;border:1px solid #ddd;text-align:center;'>{int(row.get('score',0) or 0)}</td>
          <td style='padding:6px;border:1px solid #ddd;text-align:center;'>{int(row.get('confidence',0) or 0)}</td>
          <td style='padding:6px;border:1px solid #ddd;'>{why}</td>
          <td style='padding:6px;border:1px solid #ddd;'>{link_html}</td>
        </tr>
        """

    html += "</table>"
    return html

# -----------------------------
# Weekly dashboard (last 7 days)
# -----------------------------
def build_weekly_dashboard_html(perf_csv: str, now: datetime) -> str:
    try:
        df = pd.read_csv(perf_csv)
    except Exception:
        return "<p>No performance log found yet.</p>"

    if df.empty or "run_date" not in df.columns:
        return "<p>Performance log is empty.</p>"

    df["run_date"] = pd.to_datetime(df["run_date"], errors="coerce")
    df = df.dropna(subset=["run_date"])

    last7 = df[df["run_date"] >= (now - pd.Timedelta(days=7))].copy()
    if last7.empty:
        return "<p>No trades in last 7 days.</p>"

    total = len(last7)
    correct = int((last7.get("outcome") == "✅ Correct").sum()) if "outcome" in last7.columns else 0
    rate = (correct / total * 100) if total else 0
    avg_move = last7["actual_change_pct"].mean() if "actual_change_pct" in last7.columns else 0

    best = last7.sort_values("actual_change_pct", ascending=False).head(3)[["symbol", "actual_change_pct"]]
    worst = last7.sort_values("actual_change_pct", ascending=True).head(3)[["symbol", "actual_change_pct"]]

    def rows(df2):
        out = ""
        for _, r in df2.iterrows():
            out += f"<li>{_html.escape(str(r['symbol']))}: {float(r['actual_change_pct']):.2f}%</li>"
        return out or "<li>—</li>"

    return f"""
    <h2>📅 Weekly Dashboard (Last 7 Days)</h2>
    <p>
      Trades evaluated: <b>{total}</b><br>
      Win rate: <b>{rate:.2f}%</b><br>
      Avg move: <b>{avg_move:.2f}%</b>
    </p>
    <h3>Top 3 Winners</h3>
    <ul>{rows(best)}</ul>
    <h3>Top 3 Losers</h3>
    <ul>{rows(worst)}</ul>
    """

# -----------------------------
# PRE-MARKET
# -----------------------------
def run_premarket(now: datetime):
    snapshot = get_market_snapshot()
    market_trend = snapshot.get("trend", "up")

    tickers = fetch_sp500_tickers()
    movers = calculate_top_movers(tickers, TOP_N)
    df = pd.DataFrame(movers)

    if df.empty:
        raise RuntimeError("Top movers returned empty dataset. Refusing to run.")

    df = df.head(max(EXCEL_MAX_ROWS, 20)).copy()

    if "pct_change" not in df.columns:
        df["pct_change"] = 0.0
    df["pct_change"] = pd.to_numeric(df["pct_change"], errors="coerce").fillna(0.0)

    if "current" not in df.columns:
        if "price" in df.columns:
            df["current"] = df["price"]
        else:
            raise RuntimeError("Missing 'current' price column from top_movers output.")
    df["current"] = pd.to_numeric(df["current"], errors="coerce")

    # remove invalid price rows early
    df = df[df["current"].notna() & (df["current"] > 0)].copy()
    if df.empty:
        raise RuntimeError("All rows had invalid prices. Refusing to run.")

    scores, labels, reasons_list, confs = [], [], [], []
    main_titles, main_links, flags = [], [], []
    decisions, predicted_prices, targets, stops, categories = [], [], [], [], []
    earnings_risks, trade_plans = [], []

    forecast_reason_list, forecast_trend_list, forecast_atr_list = [], [], []

    for _, row in df.iterrows():
        sym = row["symbol"]
        current = float(row["current"])
        risk = str(row.get("risk", "Medium"))
        pct = float(row.get("pct_change", 0.0))

        score_val, score_label, reasons = get_predictive_score_with_reasons(sym)
        score_val = int(score_val)
        decision = map_score_to_decision(score_val)

        news_items = fetch_news_links(sym, max_articles=3)
        main_item = news_items[0] if news_items else ""
        title = extract_headline_from_html(main_item)
        link = extract_url_from_html(main_item)
        headlines = [extract_headline_from_html(x) for x in news_items if x]
        flag = news_flag_from_headlines(headlines)

        erisk = has_earnings_soon(sym, now, EARNINGS_LOOKAHEAD_DAYS)
        if SKIP_EARNINGS_STOCKS and erisk:
            decision = "Not Advisable"

        conf = compute_confidence(score_val, pct, market_trend, flag)
        tplan = assign_trade_plan(risk=risk, pct_change=pct, market_trend=market_trend, score_val=score_val)

        scores.append(score_val)
        labels.append(score_label)
        reasons_list.append(reasons)
        confs.append(conf)

        main_titles.append(title)
        main_links.append(link)
        flags.append(flag)
        decisions.append(decision)

        try:
            fc = forecast_price_levels(sym, current=current, score=score_val)
        except Exception:
            fc = None

        if fc is None:
            predicted_prices.append(current)
            targets.append(current)
            stops.append(current * 0.97)
            forecast_reason_list.append("Forecast unavailable (fallback used)")
            forecast_trend_list.append("unknown")
            forecast_atr_list.append(0.0)
        else:
            predicted_prices.append(float(getattr(fc, "predicted_price", current)))
            targets.append(float(getattr(fc, "target_price", getattr(fc, "predicted_price", current))))
            stops.append(float(getattr(fc, "stop_loss", current * 0.97)))
            forecast_reason_list.append(str(getattr(fc, "reason", "")))
            forecast_trend_list.append(str(getattr(fc, "trend", "")))
            forecast_atr_list.append(float(getattr(fc, "atr", 0.0)))

        categories.append(get_price_category(current) if current > 0 else "Unknown")
        earnings_risks.append("YES" if erisk else "NO")
        trade_plans.append(tplan)

    df["score"] = scores
    df["score_label"] = labels
    df["reasons"] = reasons_list
    df["confidence"] = confs
    df["decision"] = decisions

    df["predicted_price"] = predicted_prices
    df["target_price"] = targets
    df["stop_loss"] = stops

    df["forecast_trend"] = forecast_trend_list
    df["forecast_atr"] = forecast_atr_list
    df["forecast_reason"] = forecast_reason_list

    df["price_category"] = categories
    df["news_flag"] = flags
    df["main_news_title"] = main_titles
    df["main_news_link"] = main_links
    df["earnings_risk"] = earnings_risks
    df["trade_plan"] = trade_plans
    df["run_date"] = now.strftime("%Y-%m-%d")

    # ✅ NO TRADE DAY check (improved)
    skip, reason = should_skip_day(df, market_trend, snapshot)

    df_support = df.sort_values(by=["confidence", "score"], ascending=False).head(
        max(EXCEL_MAX_ROWS, TRADE_MAX_PICKS)
    ).copy()

    daily_picks = (
        df_support[(df_support["decision"] == "Strong Buy") & (df_support["confidence"] >= MIN_CONFIDENCE_TO_TRADE)]
        .sort_values(by=["confidence", "score"], ascending=False)
        .head(TRADE_MAX_PICKS)
        .copy()
    )
    if daily_picks.empty:
        daily_picks = df_support.sort_values(by=["confidence", "score"], ascending=False).head(
            min(TRADE_MAX_PICKS, len(df_support))
        ).copy()

    # ✅ store picks sent for post-market evaluation
    daily_picks.to_csv(DAILY_LOG_CSV, index=False)

    daily_cols = [
        "run_date", "symbol", "price_category",
        "current", "predicted_price", "target_price", "stop_loss",
        "forecast_trend", "forecast_atr", "forecast_reason",
        "trade_plan", "earnings_risk",
        "decision", "score", "score_label", "confidence",
        "news_flag", "main_news_title", "main_news_link",
        "reasons",
    ]
    support_cols = [
        "run_date", "symbol", "price_category", "current", "pct_change",
        "predicted_price", "target_price", "stop_loss",
        "forecast_trend", "forecast_atr",
        "trade_plan", "earnings_risk",
        "decision", "score", "score_label", "confidence",
        "news_flag", "main_news_title", "main_news_link",
        "reasons",
    ]

    daily_picks_out = daily_picks[[c for c in daily_cols if c in daily_picks.columns]].copy()
    supporting = df_support[[c for c in support_cols if c in df_support.columns]].copy()

    perf = pd.DataFrame([{
        "run_date": now.strftime("%Y-%m-%d"),
        "picks_sent": len(daily_picks_out),
        "note": "Post-market updates win-rate (based on picks sent).",
        "market_trend": market_trend,
        "spy_gap_pct": float(snapshot.get("spy_gap_pct") or 0.0),
        "vix": snapshot.get("vix"),
    }])

    pre_excel = out_path(f"stock_watchlist_{now.strftime('%Y%m%d')}.xlsx")
    with pd.ExcelWriter(pre_excel, engine="openpyxl") as writer:
        daily_picks_out.to_excel(writer, sheet_name="DAILY_PICKS", index=False)
        supporting.to_excel(writer, sheet_name="SUPPORTING_DATA", index=False)
        perf.to_excel(writer, sheet_name="MODEL_PERFORMANCE", index=False)

    wb = load_workbook(pre_excel)
    for s in wb.sheetnames:
        style_excel_sheet(wb[s])
    wb.save(pre_excel)

    # ✅ Phase 2: Update portfolio + close positions + add today's picks (ONLY IF NOT SKIPPED)
    # This ensures "no trade day" doesn't add new positions.
    psummary_text = ""
    if not skip:
        try:
            cfg = PortfolioConfig()
            open_df = load_open_portfolio(cfg)

            # 1) Update/close existing
            remaining_open, closed_today = update_and_close_positions(cfg, open_df, asof=now)

            # 2) Save closed trades to history
            if closed_today is not None and not closed_today.empty:
                append_trade_history(cfg, closed_today)

            # 3) Add today's picks into portfolio (respect max open slots)
            remaining_open, added_today = add_new_positions_from_picks(
                cfg=cfg,
                open_df=remaining_open,
                picks_df=daily_picks_out,
                run_date=now,
            )

            # 4) Save updated open file
            save_open_portfolio(cfg, remaining_open)

            # 5) Summary for potential email footer
            psummary_text = portfolio_summary(remaining_open, closed_today)
        except Exception as e:
            # If portfolio logic fails, don't break the daily email.
            psummary_text = f"Portfolio update failed: {e}"

    if skip:
        html = f"""
        <h2>🛑 NO TRADE DAY – Pre Market ({now.strftime('%Y-%m-%d')})</h2>
        <p><b>Reason:</b> {_html.escape(reason)}</p>
        <p><b>Market:</b> trend={_html.escape(market_trend)}, SPY={float(snapshot.get("spy_gap_pct") or 0.0):.2f}%, VIX={(snapshot.get("vix") if snapshot.get("vix") is not None else "n/a")}</p>
        <p>Excel is attached for review, but no picks were sent for trading today.</p>
        """
        send_email(f"🛑 NO TRADE DAY – Pre Market ({now.strftime('%Y-%m-%d')})", html, attachment_path=pre_excel)
        return

    email_html = build_email_html_top_picks(daily_picks_out, now.strftime("%Y-%m-%d"))
    if psummary_text:
        email_html += f"<hr><h3>📁 Portfolio</h3><pre>{_html.escape(psummary_text)}</pre>"

    send_email(f"📈 Daily Stock Alert – Pre Market ({now.strftime('%Y-%m-%d')})", email_html, attachment_path=pre_excel)

# -----------------------------
# MIDDAY
# -----------------------------
def run_midday(now: datetime):
    tickers = fetch_sp500_tickers()
    movers = calculate_top_movers(tickers, TOP_N)
    df = pd.DataFrame(movers)
    if df.empty or "pct_change" not in df.columns:
        return

    df["pct_change"] = pd.to_numeric(df["pct_change"], errors="coerce").fillna(0.0)
    df = df[df["pct_change"].abs() >= SUDDEN_MOVER_PCT_THRESHOLD].copy()
    if df.empty:
        return

    snapshot = get_market_snapshot()
    market_trend = snapshot.get("trend", "up")

    scores, labels, reasons_list, confs, decisions = [], [], [], [], []
    titles, links = [], []

    for _, row in df.iterrows():
        sym = row["symbol"]

        score_val, score_label, reasons = get_predictive_score_with_reasons(sym)
        score_val = int(score_val)
        decision = map_score_to_decision(score_val)

        news_items = fetch_news_links(sym, max_articles=1)
        main_item = news_items[0] if news_items else ""
        title = extract_headline_from_html(main_item)
        link = extract_url_from_html(main_item)

        flag = news_flag_from_headlines([title])
        conf = compute_confidence(score_val, float(row.get("pct_change", 0.0)), market_trend, flag)

        scores.append(score_val)
        labels.append(score_label)
        reasons_list.append(reasons)
        decisions.append(decision)
        confs.append(conf)
        titles.append(title)
        links.append(link)

    df["score"] = scores
    df["score_label"] = labels
    df["reasons"] = reasons_list
    df["decision"] = decisions
    df["confidence"] = confs
    df["main_news_title"] = titles
    df["main_news_link"] = links

    df = df[df["confidence"] >= 7].sort_values(by=["confidence", "score"], ascending=False)
    if df.empty:
        return

    html = build_midday_alert(df, now.strftime("%Y-%m-%d"))
    send_email(f"⚡ Sudden Movers Alert ({now.strftime('%Y-%m-%d')})", html)

# -----------------------------
# POST-MARKET
# -----------------------------
def run_postmarket(now: datetime):
    if now.time() < POST_MARKET_START:
        print("⏳ Post-market skipped (too early).")
        return

    # ✅ Phase 2: update/close open positions and keep summary for email
    psummary_text = ""
    try:
        cfg = PortfolioConfig()
        open_df = load_open_portfolio(cfg)
        remaining_open, closed_today = update_and_close_positions(cfg, open_df, asof=now)
        if closed_today is not None and not closed_today.empty:
            append_trade_history(cfg, closed_today)
        save_open_portfolio(cfg, remaining_open)
        psummary_text = portfolio_summary(remaining_open, closed_today)
    except Exception as e:
        psummary_text = f"Portfolio update failed: {e}"

    df = evaluate_post_market_from_log(DAILY_LOG_CSV)
    if df.empty:
        html = f"""
        <h2>📊 Post-Market Summary ({now.strftime('%Y-%m-%d')})</h2>
        <p>No picks log found / empty picks log — cannot evaluate today.</p>
        """
        if psummary_text:
            html += f"<h3>📁 Portfolio Summary</h3><pre>{_html.escape(psummary_text)}</pre>"
        send_email(f"📊 Post-Market Summary ({now.strftime('%Y-%m-%d')})", html)
        return

    correct = int((df["outcome"] == "✅ Correct").sum()) if "outcome" in df.columns else 0
    total = len(df)
    rate = (correct / total * 100) if total else 0

    summary_html = f"""
    <h2>📊 Post-Market Summary ({now.strftime('%Y-%m-%d')})</h2>
    <p><b>Picks evaluated:</b> {total}<br>
       <b>Correct:</b> {correct}<br>
       <b>Incorrect:</b> {total - correct}<br>
       <b>Success Rate:</b> {rate:.2f}%</p>
    """
    if psummary_text:
        summary_html += f"<h3>📁 Portfolio Summary</h3><pre>{_html.escape(psummary_text)}</pre>"

    post_excel = out_path(f"post_market_{now.strftime('%Y%m%d')}.xlsx")
    df.to_excel(post_excel, index=False)
    wb = load_workbook(post_excel)
    for s in wb.sheetnames:
        style_excel_sheet(wb[s])
    wb.save(post_excel)

    send_email(f"📊 Post-Market Stock Summary ({now.strftime('%Y-%m-%d')})", summary_html, attachment_path=post_excel)

    # Append to perf log
    out = df.copy()
    if "run_date" not in out.columns:
        out["run_date"] = now.strftime("%Y-%m-%d")

    keep = ["run_date", "symbol", "decision", "score", "confidence", "current", "close_price", "actual_change_pct", "outcome"]
    out = out[[c for c in keep if c in out.columns]]

    try:
        if os.path.exists(PERF_LOG_CSV):
            prev = pd.read_csv(PERF_LOG_CSV)
            merged = pd.concat([prev, out], ignore_index=True)
            merged.to_csv(PERF_LOG_CSV, index=False)
        else:
            out.to_csv(PERF_LOG_CSV, index=False)
    except Exception as e:
        print("⚠️ Perf log append failed:", e)

    if now.weekday() == 4:
        dash = build_weekly_dashboard_html(PERF_LOG_CSV, now)
        if psummary_text:
            dash += f"<h3>📁 Current Portfolio</h3><pre>{_html.escape(psummary_text)}</pre>"
        send_email(f"📅 Weekly Trading Dashboard ({now.strftime('%Y-%m-%d')})", dash)

# -----------------------------
# Main
# -----------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["premarket", "midday", "postmarket"], default="premarket")
    args = parser.parse_args()

    # ✅ always in Chicago TZ
    now = datetime.now(LOCAL_TZ)

    if args.mode == "premarket":
        run_premarket(now)
    elif args.mode == "midday":
        run_midday(now)
    else:
        run_postmarket(now)

if __name__ == "__main__":
    main()