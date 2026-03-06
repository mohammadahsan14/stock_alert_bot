# strategy_performance.py
# ============================================================
# Strategy Performance Dashboard (Deterministic, Log-Driven)
# - Reads outputs/<APP_ENV>/logs/daily_stock_log.csv (required)
# - Optionally merges evaluation results if any of these exist:
#     performance_log.csv | postmarket_results.csv | evaluation_log.csv
# - Produces:
#     Console summary
#     outputs/<APP_ENV>/reports/strategy_report_<YYYYMMDD_HHMMSS>.xlsx
#
# Usage:
#   APP_ENV=prod python strategy_performance.py
#   APP_ENV=local python strategy_performance.py --days 10
#   python strategy_performance.py --app-env prod --start 2026-03-01 --end 2026-03-05
#
# Notes:
# - This file does NOT change trading logic.
# - It tolerates missing columns and missing evaluation logs.
# ============================================================

from __future__ import annotations

import argparse
import os
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Tuple, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# -----------------------------
# Config / paths
# -----------------------------
EVAL_CANDIDATE_FILES = [
    "performance_log.csv",
    "postmarket_results.csv",
    "evaluation_log.csv",
]

REQUIRED_DAILY_LOG = "daily_stock_log.csv"


@dataclass
class Paths:
    base_dir: Path
    logs_dir: Path
    reports_dir: Path
    daily_log: Path
    eval_log: Optional[Path]


def _resolve_paths(app_env: str) -> Paths:
    here = Path(__file__).resolve().parent
    base_dir = here / "outputs" / app_env
    logs_dir = base_dir / "logs"
    reports_dir = base_dir / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)

    daily_log = logs_dir / REQUIRED_DAILY_LOG

    eval_log = None
    for name in EVAL_CANDIDATE_FILES:
        p = logs_dir / name
        if p.exists() and p.stat().st_size > 0:
            eval_log = p
            break

    return Paths(
        base_dir=base_dir,
        logs_dir=logs_dir,
        reports_dir=reports_dir,
        daily_log=daily_log,
        eval_log=eval_log,
    )


# -----------------------------
# Helpers
# -----------------------------
def _to_dt(s) -> Optional[pd.Timestamp]:
    try:
        if s is None:
            return None
        v = pd.to_datetime(s, errors="coerce")
        return None if pd.isna(v) else v
    except Exception:
        return None


def _to_float(x) -> Optional[float]:
    try:
        v = pd.to_numeric(x, errors="coerce")
        if pd.isna(v):
            return None
        return float(v)
    except Exception:
        return None


def _to_int(x) -> Optional[int]:
    try:
        v = pd.to_numeric(x, errors="coerce")
        if pd.isna(v):
            return None
        return int(v)
    except Exception:
        return None


def _ensure_cols(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    df = df.copy()
    for c in cols:
        if c not in df.columns:
            df[c] = pd.NA
    return df


def _normalize_symbol(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "symbol" in df.columns:
        df["symbol"] = df["symbol"].astype(str).str.upper().str.strip()
    return df


def _style_excel(path: Path) -> None:
    """
    Simple consistent formatting:
    - blue header with white bold font
    - freeze top row
    - reasonable column widths
    """
    wb = load_workbook(path)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5597", end_color="FF2F5597", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        ws.freeze_panes = "A2"

        # set column widths based on content length (capped)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    wb.save(path)


def _date_filter(df: pd.DataFrame, start: Optional[str], end: Optional[str], days: Optional[int]) -> pd.DataFrame:
    """
    Filters by run_date if present, else by run_ts.
    start/end are YYYY-MM-DD inclusive for start, inclusive for end.
    """
    df = df.copy()
    if df.empty:
        return df

    # Build a canonical date column
    if "run_date" in df.columns and df["run_date"].notna().any():
        dt = pd.to_datetime(df["run_date"], errors="coerce")
    else:
        dt = pd.to_datetime(df.get("run_ts"), errors="coerce")

    df["_dt"] = dt

    if days is not None and days > 0:
        cutoff = pd.Timestamp(datetime.now() - timedelta(days=days))
        df = df[df["_dt"] >= cutoff]

    if start:
        s = pd.Timestamp(start)
        df = df[df["_dt"] >= s]

    if end:
        # inclusive end-of-day
        e = pd.Timestamp(end) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
        df = df[df["_dt"] <= e]

    df = df.drop(columns=["_dt"], errors="ignore")
    return df


def _mode_filter(df: pd.DataFrame, mode: Optional[str]) -> pd.DataFrame:
    if not mode:
        return df
    if "mode" not in df.columns:
        return df
    return df[df["mode"].astype(str).str.lower() == mode.lower()].copy()


def _safe_group_mean(df: pd.DataFrame, col: str) -> float:
    if col not in df.columns or df.empty:
        return float("nan")
    v = pd.to_numeric(df[col], errors="coerce")
    return float(v.mean()) if v.notna().any() else float("nan")


def _safe_group_sum(df: pd.DataFrame, mask: pd.Series) -> int:
    try:
        return int(mask.sum())
    except Exception:
        return 0


# -----------------------------
# Evaluation merging (optional)
# -----------------------------
def _load_eval(eval_path: Path) -> pd.DataFrame:
    df = pd.read_csv(eval_path)
    df = _normalize_symbol(df)

    # --- Map postmarket/performance_log schema to dashboard schema ---
    # performance_log.csv uses source_mode, target_hit, stop_hit
    if "mode" not in df.columns and "source_mode" in df.columns:
        df["mode"] = df["source_mode"]

    if "hit_target" not in df.columns and "target_hit" in df.columns:
        df["hit_target"] = df["target_hit"]

    if "hit_stop" not in df.columns and "stop_hit" in df.columns:
        df["hit_stop"] = df["stop_hit"]

    # Try to normalize common evaluation schemas into:
    # run_date, mode, symbol, outcome, hit_target, hit_stop, r_multiple (optional)
    df = _ensure_cols(df, ["run_date", "mode", "symbol", "outcome", "hit_target", "hit_stop", "r_multiple"])

    # If outcome missing but hit_target/hit_stop present, derive outcome.
    if df["outcome"].isna().all():
        ht = df["hit_target"].astype(str).str.lower().isin(["1", "true", "yes", "y"])
        hs = df["hit_stop"].astype(str).str.lower().isin(["1", "true", "yes", "y"])
        out = pd.Series(pd.NA, index=df.index)
        out[ht] = "WIN"
        out[hs] = "LOSS"
        out[(~ht) & (~hs)] = "NOT_HIT"
        df["outcome"] = out

    df["outcome"] = df["outcome"].astype(str).str.upper().str.strip()

    # Ensure run_date is present if possible
    if df["run_date"].isna().all() and "run_ts" in df.columns:
        dt = pd.to_datetime(df["run_ts"], errors="coerce")
        df["run_date"] = dt.dt.strftime("%Y-%m-%d")

    # r_multiple numeric if exists
    if "r_multiple" in df.columns:
        df["r_multiple"] = pd.to_numeric(df["r_multiple"], errors="coerce")

    return df


def _merge_daily_with_eval(daily: pd.DataFrame, eval_df: pd.DataFrame) -> pd.DataFrame:
    daily = daily.copy()
    eval_df = eval_df.copy()

    daily = _ensure_cols(daily, ["run_date", "mode", "symbol"])
    eval_df = _ensure_cols(eval_df, ["run_date", "mode", "symbol", "outcome", "r_multiple"])

    daily["run_date"] = daily["run_date"].astype(str)
    eval_df["run_date"] = eval_df["run_date"].astype(str)

    # Deduplicate eval on (run_date, mode, symbol) keep last
    eval_df = eval_df.drop_duplicates(subset=["run_date", "mode", "symbol"], keep="last")

    merged = daily.merge(
        eval_df[["run_date", "mode", "symbol", "outcome", "r_multiple"]],
        on=["run_date", "mode", "symbol"],
        how="left",
    )
    return merged


# -----------------------------
# Analytics
# -----------------------------
def _build_summary_tables(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Returns:
      1) headline KPIs (single-row)
      2) breakdown by (mode, stance, market_regime)
      3) breakdown by confidence bucket
    """
    df = df.copy()

    df = _ensure_cols(df, [
        "run_date", "mode", "symbol",
        "stance", "confidence", "score", "score_label",
        "market_regime", "expected_rr", "win_prob",
        "outcome", "r_multiple",
    ])

    # Normalize stance
    df["stance"] = df["stance"].astype(str).str.upper().str.strip()
    df.loc[df["stance"].isin(["", "NAN", "NONE"]), "stance"] = "UNKNOWN"

    # Normalize market_regime
    df["market_regime"] = df["market_regime"].astype(str).str.lower().str.strip()
    df.loc[df["market_regime"].isin(["", "nan", "none"]), "market_regime"] = "unknown"

    # Confidence numeric
    df["confidence"] = pd.to_numeric(df["confidence"], errors="coerce")

    # Outcome normalization
    df["outcome"] = df["outcome"].astype(str).str.upper().str.strip()
    df.loc[df["outcome"].isin(["", "NAN", "NONE"]), "outcome"] = pd.NA

    has_eval = df["outcome"].notna().any()

    total_rows = len(df)
    unique_symbols = df["symbol"].nunique(dropna=True)

    go_count = int((df["stance"] == "GO").sum())
    watch_count = int((df["stance"] == "WATCH").sum())

    avg_conf = _safe_group_mean(df, "confidence")
    avg_score = _safe_group_mean(df, "score")
    avg_rr = _safe_group_mean(df, "expected_rr")
    avg_winp = _safe_group_mean(df, "win_prob")

    # Outcomes if evaluation exists
    win = loss = not_hit = evaluated = 0
    win_rate = float("nan")
    avg_r_mult = float("nan")

    if has_eval:
        evaluated = int(df["outcome"].notna().sum())
        win = int((df["outcome"] == "WIN").sum())
        loss = int((df["outcome"] == "LOSS").sum())
        not_hit = int((df["outcome"] == "NOT_HIT").sum())
        win_rate = (win / evaluated * 100.0) if evaluated > 0 else float("nan")

        if "r_multiple" in df.columns:
            rm = pd.to_numeric(df["r_multiple"], errors="coerce")
            avg_r_mult = float(rm.mean()) if rm.notna().any() else float("nan")

    kpis = pd.DataFrame([{
        "rows": total_rows,
        "unique_symbols": unique_symbols,
        "GO_rows": go_count,
        "WATCH_rows": watch_count,
        "avg_confidence": avg_conf,
        "avg_score": avg_score,
        "avg_expected_rr": avg_rr,
        "avg_win_prob": avg_winp,
        "has_evaluation": bool(has_eval),
        "evaluated_rows": evaluated,
        "wins": win,
        "losses": loss,
        "not_hit": not_hit,
        "win_rate_%": win_rate,
        "avg_r_multiple": avg_r_mult,
    }])

    # Breakdown by mode/stance/regime
    grp = df.groupby(["mode", "stance", "market_regime"], dropna=False)
    by_msr = grp.agg(
        rows=("symbol", "count"),
        unique_symbols=("symbol", "nunique"),
        avg_conf=("confidence", "mean"),
        avg_score=("score", "mean"),
        avg_expected_rr=("expected_rr", "mean"),
        avg_win_prob=("win_prob", "mean"),
        evaluated=("outcome", lambda x: x.notna().sum()),
        wins=("outcome", lambda x: (x == "WIN").sum()),
        losses=("outcome", lambda x: (x == "LOSS").sum()),
        not_hit=("outcome", lambda x: (x == "NOT_HIT").sum()),
    ).reset_index()

    def _wr(row):
        ev = row.get("evaluated", 0)
        w = row.get("wins", 0)
        return (w / ev * 100.0) if ev and ev > 0 else float("nan")

    by_msr["win_rate_%"] = by_msr.apply(_wr, axis=1)

    # Confidence bucket table
    def bucket(c):
        if pd.isna(c):
            return "NA"
        c = int(c)
        if c <= 3:
            return "1-3"
        if c <= 5:
            return "4-5"
        if c <= 7:
            return "6-7"
        return "8-10"

    df["conf_bucket"] = df["confidence"].apply(bucket)

    by_conf = df.groupby(["conf_bucket", "mode", "stance"], dropna=False).agg(
        rows=("symbol", "count"),
        unique_symbols=("symbol", "nunique"),
        avg_score=("score", "mean"),
        avg_expected_rr=("expected_rr", "mean"),
        avg_win_prob=("win_prob", "mean"),
        evaluated=("outcome", lambda x: x.notna().sum()),
        wins=("outcome", lambda x: (x == "WIN").sum()),
        losses=("outcome", lambda x: (x == "LOSS").sum()),
        not_hit=("outcome", lambda x: (x == "NOT_HIT").sum()),
    ).reset_index()
    by_conf["win_rate_%"] = by_conf.apply(_wr, axis=1)

    return kpis, by_msr.sort_values(["mode", "stance", "market_regime"]), by_conf.sort_values(["conf_bucket", "mode", "stance"])


# -----------------------------
# Reporting
# -----------------------------
def _write_excel_report(
    out_path: Path,
    *,
    kpis: pd.DataFrame,
    by_msr: pd.DataFrame,
    by_conf: pd.DataFrame,
    daily_rows: pd.DataFrame,
    eval_path: Optional[Path],
    daily_path: Path,
) -> None:
    meta = pd.DataFrame([{
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "daily_log": str(daily_path),
        "eval_log_used": str(eval_path) if eval_path else "",
    }])

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        meta.to_excel(xw, sheet_name="META", index=False)
        kpis.to_excel(xw, sheet_name="KPI", index=False)
        by_msr.to_excel(xw, sheet_name="BY_MODE_STANCE_REGIME", index=False)
        by_conf.to_excel(xw, sheet_name="BY_CONF_BUCKET", index=False)

        # keep last: raw
        daily_rows.to_excel(xw, sheet_name="RAW_ROWS", index=False)

    _style_excel(out_path)


def _print_console(kpis: pd.DataFrame, eval_path: Optional[Path]) -> None:
    row = kpis.iloc[0].to_dict()

    print("\n==============================")
    print("📊 Strategy Performance Report")
    print("==============================")
    print(f"Rows: {int(row.get('rows', 0))} | Unique symbols: {int(row.get('unique_symbols', 0))}")
    print(f"GO: {int(row.get('GO_rows', 0))} | WATCH: {int(row.get('WATCH_rows', 0))}")
    print(f"Avg confidence: {row.get('avg_confidence', float('nan')):.2f} | Avg score: {row.get('avg_score', float('nan')):.2f}")
    print(f"Avg expected R:R: {row.get('avg_expected_rr', float('nan')):.2f} | Avg win_prob: {row.get('avg_win_prob', float('nan')):.2f}")

    if eval_path:
        print(f"\nEvaluation log detected: {eval_path.name}")
    else:
        print("\nEvaluation log detected: (none)")

    if bool(row.get("has_evaluation", False)):
        print(
            f"Evaluated: {int(row.get('evaluated_rows', 0))} | "
            f"Wins: {int(row.get('wins', 0))} | Losses: {int(row.get('losses', 0))} | Not hit: {int(row.get('not_hit', 0))}"
        )
        wr = row.get("win_rate_%", float("nan"))
        print(f"Win rate: {wr:.2f}%")
        arm = row.get("avg_r_multiple", float("nan"))
        if arm == arm:  # not NaN
            print(f"Avg R multiple: {arm:.2f}")
    else:
        print("Outcomes: not available yet (waiting for postmarket evaluation log).")


# -----------------------------
# Main
# -----------------------------
def main() -> int:
    parser = argparse.ArgumentParser(description="Historical performance dashboard from logs (daily + optional eval).")
    parser.add_argument("--app-env", default=os.getenv("APP_ENV", "prod"), help="APP_ENV to read from outputs/<APP_ENV>/")
    parser.add_argument("--mode", default="", help="Filter mode: premarket | midday | postmarket (optional)")
    parser.add_argument("--days", type=int, default=0, help="Lookback N days (optional). If >0, overrides start/end if not provided.")
    parser.add_argument("--start", default="", help="Start date YYYY-MM-DD (optional)")
    parser.add_argument("--end", default="", help="End date YYYY-MM-DD (optional)")
    parser.add_argument("--export-raw-max", type=int, default=2000, help="Max rows to include in RAW_ROWS sheet")
    args = parser.parse_args()

    app_env = (args.app_env or "prod").strip().lower()
    paths = _resolve_paths(app_env)

    if not paths.daily_log.exists() or paths.daily_log.stat().st_size == 0:
        raise FileNotFoundError(f"Missing required daily log: {paths.daily_log}")

    daily = pd.read_csv(paths.daily_log)
    daily = _normalize_symbol(daily)

    # Ensure core columns exist
    daily = _ensure_cols(daily, ["run_ts", "run_date", "mode", "symbol", "stance"])

    # Normalize run_date formatting (prevents merge mismatches)
    daily["run_date"] = pd.to_datetime(daily["run_date"], errors="coerce").dt.strftime("%Y-%m-%d")
    # If run_date still missing, derive from run_ts
    if daily["run_date"].isna().all() and "run_ts" in daily.columns:
        daily["run_date"] = pd.to_datetime(daily["run_ts"], errors="coerce").dt.strftime("%Y-%m-%d")

    # Normalize mode text
    daily["mode"] = daily["mode"].astype(str).str.lower().str.strip()

    # Filter by date and mode
    daily = _date_filter(daily, start=args.start or None, end=args.end or None, days=(args.days if args.days > 0 else None))
    daily = _mode_filter(daily, args.mode or None)

    # Optional evaluation merge
    eval_df = None
    merged = daily
    if paths.eval_log:
        try:
            eval_df = _load_eval(paths.eval_log)
            # Normalize eval fields for merge consistency
            if "mode" in eval_df.columns:
                eval_df["mode"] = eval_df["mode"].astype(str).str.lower().str.strip()

            if "run_date" in eval_df.columns:
                eval_df["run_date"] = pd.to_datetime(eval_df["run_date"], errors="coerce").dt.strftime("%Y-%m-%d")
            eval_df = _date_filter(eval_df, start=args.start or None, end=args.end or None, days=(args.days if args.days > 0 else None))
            eval_df = _mode_filter(eval_df, args.mode or None)
            merged = _merge_daily_with_eval(daily, eval_df)


        except Exception as e:
            # soft-fail: keep report working without eval
            print(f"⚠️ Failed to load/merge eval log ({paths.eval_log.name}): {e}")
            merged = daily

    # Limit RAW rows for Excel
    raw_rows = merged.copy()
    if len(raw_rows) > int(args.export_raw_max):
        raw_rows = raw_rows.sort_values(by=["run_date", "mode", "symbol"], ascending=False).head(int(args.export_raw_max))

    # Build tables
    kpis, by_msr, by_conf = _build_summary_tables(merged)

    # Console
    _print_console(kpis, paths.eval_log)

    # Export report
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = paths.reports_dir / f"strategy_report_{ts}.xlsx"
    _write_excel_report(
        out_xlsx,
        kpis=kpis,
        by_msr=by_msr,
        by_conf=by_conf,
        daily_rows=raw_rows,
        eval_path=paths.eval_log,
        daily_path=paths.daily_log,
    )

    print(f"\n✅ Report written: {out_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())